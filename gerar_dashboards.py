\
# -*- coding: utf-8 -*-
"""
Regenera os dois dashboards HTML (Equipes e Frota) a partir das planilhas fonte.
Le os dados brutos (nao depende de tabela dinamica / pivot refresh) e substitui,
via regex, os trechos dinamicos dos arquivos HTML existentes -- preservando
todo o layout/CSS/JS ja construidos.

Fontes:
  - Diario de Bordo (Extreme Wind).xlsx        -> dashboard-resumo-equipes.html
  - Controle de Frota - ... - ATUAL.xlsb.xlsx  -> dashboard-frota-manutencao.html

Uso: python3 gerar_dashboards.py
"""
import os
import re
import sys
import time
import json
import shutil
import zipfile
import datetime
import unicodedata
import subprocess
import tempfile
from collections import defaultdict, Counter
import openpyxl
from openpyxl.utils import range_boundaries


def wait_for_file(path, tries=4, delay_seconds=5):
    """Tenta aguardar um arquivo Dropbox 'somente online' terminar de baixar.
    Faz ate `tries` checagens de os.path.isfile, com pausa entre elas.
    Retorna True se o arquivo aparecer, False se continuar ausente apos todas
    as tentativas (nesse caso o chamador deve seguir tratando como erro real)."""
    for attempt in range(tries):
        if os.path.isfile(path):
            return True
        if attempt < tries - 1:
            time.sleep(delay_seconds)
    return os.path.isfile(path)

# Extrai os ranges de celulas mescladas direto do XML da planilha (via zip),
# sem precisar abrir o workbook no modo "completo" do openpyxl (que carrega
# estilos/formatacao de todas as abas e e MUITO mais lento em arquivos
# grandes). Permite usar read_only=True (rapido) e ainda resolver mesclas.
def get_merged_ranges_fast(path, sheet_name):
    with zipfile.ZipFile(path) as z:
        wb_xml = z.read('xl/workbook.xml').decode('utf-8')
        rels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        m = re.search(r'<sheet[^>]*name="%s"[^>]*r:id="([^"]+)"' % re.escape(sheet_name), wb_xml)
        if not m:
            m = re.search(r'<sheet[^>]*r:id="([^"]+)"[^>]*name="%s"' % re.escape(sheet_name), wb_xml)
        if not m:
            return []
        rid = m.group(1)
        m2 = re.search(r'<Relationship[^>]*Id="%s"[^>]*Target="([^"]+)"' % re.escape(rid), rels_xml)
        if not m2:
            return []
        target = m2.group(1)
        sheet_path = 'xl/' + target if not target.startswith('/') else target.lstrip('/')
        sheet_xml = z.read(sheet_path).decode('utf-8')
    ranges = []
    for mm in re.finditer(r'<mergeCell ref="([^"]+)"', sheet_xml):
        min_col, min_row, max_col, max_row = range_boundaries(mm.group(1))
        ranges.append((min_row, min_col, max_row, max_col))
    return ranges

# Le uma aba inteira para um dict {(linha,coluna): valor} via iter_rows, que e
# a forma rapida de acessar celulas em modo read_only (acesso aleatorio direto
# via ws.cell(row=,column=) e MUITO lento nesse modo).
def sheet_to_grid(ws):
    grid = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for c in row:
            if c.value is not None:
                grid[(c.row, c.column)] = c.value
    return grid

# Caminhos derivados da localizacao do proprio script (independente do prefixo
# de sessao/sandbox, que muda a cada execucao agendada). Este script deve
# morar em ".../SITE E APP (DASHBOARDS)/_scripts/gerar_dashboards.py". As
# pastas Dropbox conectadas separadamente ("CONTROLE DAS EQUIPES GERAL",
# "BOLETOS LOCADORAS", "02 - ENTREGA MATERIAIS" etc.) podem aparecer como
# irmas de "SITE E APP (DASHBOARDS)" diretamente sob mnt/, OU como irmas de
# uma pasta intermediaria "4 - MANUFATURA" (o layout de mount muda conforme a
# sessao/conta) -- por isso procuramos em AMBOS os niveis, nao assumimos
# profundidade fixa.
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SITE_DIR = os.path.dirname(_SCRIPT_DIR)                    # .../SITE E APP (DASHBOARDS)
_CANDIDATE_ROOTS = [
    os.path.dirname(SITE_DIR),                  # mnt/ direto (irmas soltas)
    os.path.dirname(os.path.dirname(SITE_DIR)),  # mnt/ com "4 - MANUFATURA" no meio
]


def _find_sibling(folder_name, filename=None):
    """Procura 'folder_name' (opcionalmente + filename) nos possiveis mount
    roots candidatos, retornando o primeiro caminho que existir. Se nenhum
    existir, devolve o caminho montado a partir do primeiro candidato (para
    a mensagem de erro ficar clara sobre onde tentou procurar)."""
    for root in _CANDIDATE_ROOTS:
        path = os.path.join(root, folder_name, filename) if filename else os.path.join(root, folder_name)
        if os.path.exists(path):
            return path
    return os.path.join(_CANDIDATE_ROOTS[0], folder_name, filename) if filename else os.path.join(_CANDIDATE_ROOTS[0], folder_name)


def _find_fixed_subpath(*parts):
    """BUG corrigido (04-09/08/2026): BOLETOS_XLSX e ENVIOS_XLSX nao vivem
    direto na raiz -- ficam varios niveis abaixo dela (ex.: '13 - LOGISTICA/
    CONTROLE DE VEICULOS/Zz-PRINCIPAIS/BOLETOS LOCADORAS/...' e '3 - SUPRIMENTOS/
    01 - ALMOXARIFADO/02 - ENTREGA MATERIAIS/...'). _find_sibling so procura UM
    nivel abaixo de cada candidate root, entao para esses dois casos ele nunca
    encontrava o arquivo e caia no path de fallback inexistente -> FileNotFoundError
    (ou, com o fix de origem/None desta mesma leva, d_bo/d_en=None e a secao
    correspondente ficava "preservada" indefinidamente em vez de atualizar de
    verdade). Aqui procuramos o caminho completo (todos os parts) sob cada
    candidate root, ao inves de apenas um nivel.

    BUG corrigido (25/08/2026): em sessoes onde o Cowork conecta SO a pasta
    final (ex.: "BOLETOS LOCADORAS" ou "02 - ENTREGA MATERIAIS") como pasta de
    workspace de primeiro nivel -- sem as pastas intermediarias "13 -
    LOGISTICA/..." ou "3 - SUPRIMENTOS/..." existirem no mount -- o caminho
    completo acima nunca existe em nenhum candidate root, mesmo com o arquivo
    presente. Agora tambem tentamos a pasta final + arquivo (ultimos dois
    "parts") direto em cada candidate root, cobrindo esse layout de mount."""
    for root in _CANDIDATE_ROOTS:
        path = os.path.join(root, *parts)
        if os.path.exists(path):
            return path
    if len(parts) >= 2:
        for root in _CANDIDATE_ROOTS:
            path = os.path.join(root, *parts[-2:])
            if os.path.exists(path):
                return path
    return os.path.join(_CANDIDATE_ROOTS[-1], *parts)


EQUIPES_XLSX = _find_sibling("CONTROLE DAS EQUIPES GERAL", "Diário de Bordo (Extreme Wind).xlsx")

# VAGAS MESTRE: planilha com BANCO DE DADOS de colaboradores (status, qualificacao, contrato, etc)
# Localizada dinamicamente em 2 - RH & DP / 2 - MASTER / 2026 -- procura o arquivo mais recente de agosto/setembro
def _find_vagas_mestre_xlsx():
    """Procura pelo arquivo MASTER de RH mais recente (ex: '08 - MASTER AGOSTO 18.08.xlsx').
    Preferencia por Agosto > Julho > arquivos anteriores no ano. Sem ano=None."""
    import glob as glob_
    candidates = []
    for root in _CANDIDATE_ROOTS:
        for pattern in [
            os.path.join(root, "2 - RH & DP", "2 - MASTER", "2026", "08 - MASTER AGOSTO*.xlsx"),
            os.path.join(root, "2 - RH & DP", "2 - MASTER", "2026", "07 - MASTER JULHO*.xlsx"),
            os.path.join(root, "2 - RH & DP", "2 - MASTER", "2026", "*.xls*"),
        ]:
            files = glob_.glob(pattern)
            candidates.extend(files)
    if candidates:
        candidates.sort(key=lambda f: os.path.getmtime(f), reverse=True)
        return candidates[0]
    return None

VAGAS_MESTRE_XLSX = _find_vagas_mestre_xlsx()

# Fila de lancamentos futuros de mobilizacao/desmobilizacao (aplicada por
# aplicar_pendencias.py quando a data chega). Um item conta como "em
# desmobilizacao" no dashboard a partir de "data_fim_projeto" (data em que a
# SAIDA foi efetivamente decidida/confirmada -- pode ser hoje mesmo, mesmo que
# a viagem/finalizacao fisica so aconteca depois), ate ser aplicado na
# planilha. Ou seja, data_fim_projeto NAO precisa ser a data de encerramento
# de um parque/cliente -- e simplesmente "a partir de quando essa saida deve
# refletir no headcount planejado".
PENDENCIAS_JSON = _find_sibling("CONTROLE DAS EQUIPES GERAL", "pendencias_mobilizacao.json")


def load_em_desmobilizacao():
    """Devolve dict {mat(str): item} dos tecnicos cuja saida ja foi decidida
    (data_fim_projeto <= hoje) mas que ainda nao foram aplicados na planilha."""
    if not os.path.isfile(PENDENCIAS_JSON):
        return {}
    try:
        with open(PENDENCIAS_JSON, encoding="utf-8") as f:
            fila = json.load(f)
    except Exception:
        return {}
    hoje = datetime.date.today()
    out = {}
    for item in fila.get("pendencias", []):
        if item.get("aplicado"):
            continue
        try:
            fim_projeto = datetime.date.fromisoformat(item["data_fim_projeto"])
        except Exception:
            continue
        if fim_projeto <= hoje:
            out[str(item["mat"])] = item
    return out
# Planilha de Frota foi movida para dentro de SITE E APP (DASHBOARDS)/Checklist (Sr. Fernando)
# (antes vinha da pasta "Joao Diniz- EM USO", que parou de ser atualizada).
def _find_frota_xlsx():
    """O Dropbox por vezes cria uma copia de nome diferente quando ha conflito
    de sincronizacao (ex.: "...ATUAL.xlsb (Copia em conflito de <dispositivo>
    <data>).xlsx"), fazendo o nome exato esperado abaixo nao existir por um
    tempo (mesmo com o arquivo real presente na pasta, so com outro nome).
    Corrigido em 22/08/2026: se o nome exato nao existir, cai para o arquivo
    mais recente (maior mtime) na mesma pasta cujo nome comece com "Controle
    de Frota - TESTE DE FORMULAÇÃO" -- cobre tanto o nome oficial quanto
    variantes de conflito, sempre preferindo a versao mais nova."""
    pasta = os.path.join(SITE_DIR, "Checklist (Sr. Fernando)")
    nome_oficial = "Controle de Frota - TESTE DE FORMULAÇÃO (BACK UP) - ATUAL.xlsb.xlsx"
    caminho_oficial = os.path.join(pasta, nome_oficial)
    if os.path.isfile(caminho_oficial):
        return caminho_oficial
    try:
        candidatos = [
            os.path.join(pasta, f) for f in os.listdir(pasta)
            if f.startswith("Controle de Frota - TESTE DE FORMULAÇÃO") and f.lower().endswith(".xlsx")
        ]
    except OSError:
        return caminho_oficial
    if not candidatos:
        return caminho_oficial
    candidatos.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidatos[0]


FROTA_XLSX = _find_frota_xlsx()

EQUIPES_HTML = SITE_DIR + "/dashboard-resumo-equipes.html"
FROTA_HTML = SITE_DIR + "/dashboard-frota-manutencao.html"
INDEX_HTML = SITE_DIR + "/index.html"
BOLETOS_HTML = SITE_DIR + "/dashboard-boletos.html"

# Planilha de Boletos vive VARIOS niveis abaixo da raiz (nao e irma direta):
# "13 - LOGISTICA/CONTROLE DE VEÍCULOS/Zz-PRINCIPAIS/BOLETOS LOCADORAS/...".
# BUG corrigido (09/08/2026): _find_sibling so procura 1 nivel abaixo de cada
# candidate root, entao NUNCA encontrava este arquivo -- caia sempre no path
# de fallback inexistente, e o main() tratava isso como fonte indisponivel
# (d_bo=None) toda vez que rodava, nao so ocasionalmente. Confirmado o path
# real via busca no Dropbox em 09/08/2026.
BOLETOS_XLSX = _find_fixed_subpath(
    "13 - LOGISTICA", "CONTROLE DE VEÍCULOS", "Zz-PRINCIPAIS",
    "BOLETOS LOCADORAS", "01. CONTROLE DE CONTAS A PAGAR UNIDAS.xlsx",
)

# Planilha de Envios Logisticos (requisicoes de materiais) vive em
# "3 - SUPRIMENTOS/01 - ALMOXARIFADO/02 - ENTREGA MATERIAIS/...", mesmo
# problema de profundidade de BOLETOS_XLSX (ver comentario acima). A aba
# "DASHBOARD (ROBSON)" dentro deste arquivo e atualizada por outra rotina
# automatizada e nao e lida por este script.
ENVIOS_XLSX = _find_fixed_subpath(
    "3 - SUPRIMENTOS", "01 - ALMOXARIFADO", "02 - ENTREGA MATERIAIS",
    "Controle_Logistico_Requisicoes.xlsx",
)
ENVIOS_HTML = SITE_DIR + "/dashboard-envios-logisticos.html"

# Planilha de Reembolso vive na pasta Dropbox "REEMBOLSO SEMANAL", irma no mesmo
# padrao das outras (BOLETOS_XLSX, ENVIOS_XLSX). O nome do arquivo e da subpasta
# leva o ANO (ex: "REEMBOLSOS 2026/Formulário de Reembolso Master (2026).xlsx"),
# e o padrao ja existe para 2025 tambem -- assumimos que a convencao se repete
# a cada ano novo (subpasta "REEMBOLSOS <ano>" + arquivo "... Master (<ano>).xlsx").
# Se a Extreme Wind mudar essa convencao de nome, este caminho vai parar de
# achar o arquivo e o dashboard vai logar [ERRO] (nao quebra os outros).
#
# BUG corrigido (25/08/2026): desde que "SITE E APP (DASHBOARDS)" virou pasta
# de primeiro nivel dentro de "02 - EXTREME WIND" (13/08/2026), SITE_DIR passou
# a ser "mnt/17 - SITE E APP (DASHBOARDS)". _find_sibling so olha
# dirname(SITE_DIR) ("mnt/") e dirname(dirname(SITE_DIR)) (um nivel ACIMA de
# "mnt/" inteiro) -- nunca testa "mnt/4 - MANUFATURA/", que e onde
# "REEMBOLSO SEMANAL" continua vivendo de fato. Adicionamos esse caminho como
# candidate root explicito.
def _reembolso_xlsx_path():
    candidatos_base = [
        _find_sibling("REEMBOLSO SEMANAL"),
        os.path.join(os.path.dirname(SITE_DIR), "4 - MANUFATURA", "REEMBOLSO SEMANAL"),
    ]
    ano = datetime.date.today().year
    for base in candidatos_base:
        caminho = os.path.join(base, f"REEMBOLSOS {ano}", f"Formulário de Reembolso Master ({ano}).xlsx")
        if os.path.isfile(caminho):
            return caminho
    return os.path.join(candidatos_base[0], f"REEMBOLSOS {ano}", f"Formulário de Reembolso Master ({ano}).xlsx")


REEMBOLSO_XLSX = _reembolso_xlsx_path()
REEMBOLSO_HTML = SITE_DIR + "/dashboard-reembolso.html"

# "Nossas Demandas" e uma copia consolidada (nao uma fonte propria) -- agrega
# os itens ja marcados como pendentes/urgentes em cada um dos 4 dashboards
# acima, num unico lugar. Nao le nenhuma planilha; so reaproveita os dicts
# ja construidos por build_equipes_data/build_frota_data/build_boletos_data/
# build_envios_data dentro de main().
DEMANDAS_HTML = SITE_DIR + "/dashboard-nossas-demandas.html"

# Controle de Efetivo: painel de gestao de disponibilidade, BDH, convocacoes e historico
EFETIVO_HTML = SITE_DIR + "/dashboard-controle-efetivo.html"

# O site publicado (GitHub Pages, acessado pelo app mobile via capacitor.config.json
# "server.url") serve os arquivos de "Mobile/www/", NAO os desta pasta raiz --
# sao copias separadas. O deploy (Mobile/deploy-pages.yml) so publica quando
# alguem faz "git push" da pasta www/ pro repositorio -- este script nao tem
# acesso ao git/GitHub, entao so consegue manter as DUAS copias locais iguais;
# o push pro GitHub continua manual.
MOBILE_WWW_DIR = os.path.join(SITE_DIR, "Mobile", "www")


def sync_to_mobile_www():
    """Copia os HTMLs (e so eles -- assets/manifest/sw.js nao mudam) da pasta
    raiz pra Mobile/www/, se essa pasta existir."""
    if not os.path.isdir(MOBILE_WWW_DIR):
        return
    import shutil
    for fname in ("index.html", "dashboard-resumo-equipes.html", "dashboard-frota-manutencao.html", "dashboard-boletos.html", "dashboard-envios-logisticos.html", "dashboard-reembolso.html", "dashboard-nossas-demandas.html", "dashboard-controle-efetivo.html"):
        src = os.path.join(SITE_DIR, fname)
        if os.path.isfile(src):
            shutil.copyfile(src, os.path.join(MOBILE_WWW_DIR, fname))
    log(f"  Sincronizado com {MOBILE_WWW_DIR} (falta apenas o \"git push\" pro GitHub Pages publicar).")


LOG = []
def log(msg):
    LOG.append(msg)
    print(msg)

MESES_PT = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']

def fmt_date_br(d):
    if isinstance(d, datetime.datetime):
        return d.strftime('%d/%m/%Y')
    return str(d)

def js_str(s):
    return json.dumps(s, ensure_ascii=False)

def rhu(x):
    """Arredondamento 'metade para cima' (0.5 -> 1), como se espera no uso
    cotidiano (Python round() usa arredondamento bancario, que confunde
    quem esta conferindo os numeros a olho)."""
    import math
    return int(math.floor(x + 0.5))

# ============================================================
# EQUIPES
# ============================================================

def build_equipes_data():
    wb = openpyxl.load_workbook(EQUIPES_XLSX, data_only=True, read_only=True)

    # ---- CONTROLE DE EQUIPES (ativos hoje / real vs plano) ----
    ws = wb['CONTROLE DE EQUIPES']
    grid = sheet_to_grid(ws)
    merged_val = {}
    for (min_row, min_col, max_row, max_col) in get_merged_ranges_fast(EQUIPES_XLSX, 'CONTROLE DE EQUIPES'):
        tl = grid.get((min_row, min_col))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_val[(r, c)] = tl

    def getv(r, c):
        if (r, c) in merged_val:
            return merged_val[(r, c)]
        return grid.get((r, c))

    em_desmob_por_mat = load_em_desmobilizacao()
    em_desmobilizacao = []

    # Agrupamos por (PARQUE, CLIENTE) -- nao so por PARQUE -- porque um mesmo
    # parque fisico pode ter mais de um cliente/projeto rodando ao mesmo tempo
    # (ex: TRAIRI/CE tem uma equipe GE/EVAIR e uma equipe SIEMENS/WILDNER
    # separadas). Agrupar so por parque escondia isso: mostrava "TRAIRI/CE"
    # como uma linha unica com o cliente errado (o primeiro em ordem
    # alfabetica) somando o headcount dos dois clientes junto.
    parques = defaultdict(lambda: {"real": 0, "plano": 0, "cliente": set(), "suop": set(), "equipes": set()})
    tecnicos_ativos = []
    for r in range(4, ws.max_row + 1):
        parque = getv(r, 8)   # H
        if not parque:
            continue
        cliente = getv(r, 9)   # I
        suop = getv(r, 11)     # K
        nome = grid.get((r, 3))  # C (nao mesclado)
        equipe = grid.get((r, 13))  # M
        mat = grid.get((r, 2))       # MAT
        mat_str = str(mat) if mat is not None else None
        cliente_norm = cliente.strip() if isinstance(cliente, str) and cliente.strip() else "-"
        key = (parque, cliente_norm)
        if nome and mat_str in em_desmob_por_mat:
            # saida ja decidida (data_fim_projeto <= hoje) mas ainda nao
            # aplicada na planilha (viagem/finalizacao em andamento) -- nao
            # conta como ativo NEM como headcount planejado a partir de agora,
            # entra so na lista separada "em desmobilizacao". Note que a vaga
            # NAO volta a contar em "plano" (a posicao sai do planejamento
            # assim que a saida e decidida, nao so quando o vago abre na
            # planilha).
            pend = em_desmob_por_mat[mat_str]
            em_desmobilizacao.append({
                "mat": mat_str,
                "nome": nome.strip(),
                "parque": parque.strip(),
                "cliente": cliente_norm,
                "data_finalizacao": pend.get("data_finalizacao_tecnico"),
            })
        else:
            # vago (posicao a preencher) ou tecnico ativo: ambos contam para
            # o headcount planejado ("plano"); so o tecnico ativo conta tambem
            # como "real".
            parques[key]["plano"] += 1
        if nome and mat_str not in em_desmob_por_mat:
            parques[key]["real"] += 1
            acumulado = getv(r, 6)       # ACUMULADO (dias em campo)
            tecnicos_ativos.append({
                "mat": mat_str,
                "nome": nome.strip(),
                "acumulado": acumulado if isinstance(acumulado, (int, float)) else None,
                "parque": parque.strip(),
                "suop": suop.strip() if isinstance(suop, str) and suop.strip() else "-",
            })
        if cliente:
            parques[key]["cliente"].add(cliente)
        if suop:
            parques[key]["suop"].add(suop)
        if equipe:
            parques[key]["equipes"].add(equipe)
    tecnicos_ativos = [t for t in tecnicos_ativos if isinstance(t["acumulado"], (int, float)) and t["acumulado"] > 90]
    tecnicos_ativos.sort(key=lambda t: t["acumulado"], reverse=True)  # so>90 dias, maior->menor
    em_desmobilizacao.sort(key=lambda t: t["data_finalizacao"] or "")

    parques_list = []
    for (parque_nome, cliente_chave), v in parques.items():
        if v["real"] == 0 and v["plano"] == 0:
            # combo parque+cliente sem nenhum tecnico ativo NEM vaga planejada
            # (ex: TRAIRI/CE+GE depois que os 3 tecnicos da equipe Evair
            # foram todos pra "em desmobilizacao") -- projeto encerrado, nao
            # deve continuar contando no total de parques ("X/Y"), senao Y
            # fica inflado com uma linha fantasma de headcount zero.
            continue
        parques_list.append({
            "nome": parque_nome.strip(),
            "cliente": sorted(v["cliente"])[0] if v["cliente"] else cliente_chave,
            "suop": sorted(v["suop"])[0] if v["suop"] else "-",
            "real": v["real"],
            "plano": v["plano"],
            "nsub": len(v["equipes"]) or 1,
        })

    tot_real = sum(p["real"] for p in parques_list)
    tot_plano = sum(p["plano"] for p in parques_list)
    tot_real_fisico = tot_real + len(em_desmobilizacao)
    parques_ativos = sum(1 for p in parques_list if p["real"] > 0)
    parques_total = len(parques_list)

    # ---- DADOS (historico mob/desmob/cancel) ----
    wsd = wb['DADOS']
    gridd = sheet_to_grid(wsd)
    rows = []
    for r in range(4, wsd.max_row + 1):
        desc = gridd.get((r, 7))
        data = gridd.get((r, 5))
        parque = gridd.get((r, 10))
        cliente = gridd.get((r, 11))
        motivo = gridd.get((r, 12))
        if desc not in ("MOBILIZAÇÃO", "DESMOBILIZAÇÃO", "CANCELAMENTO"):
            continue
        rows.append(dict(desc=desc, data=data, parque=parque, cliente=cliente, motivo=motivo))

    c = Counter(r["desc"] for r in rows)
    n_mob, n_desmob, n_canc = c["MOBILIZAÇÃO"], c["DESMOBILIZAÇÃO"], c["CANCELAMENTO"]
    n_total = n_mob + n_desmob + n_canc

    parque_agg = defaultdict(lambda: Counter())
    for r in rows:
        if r["parque"]:
            parque_agg[r["parque"]][r["desc"]] += 1
    ranking_parque = []
    for p, cnt in parque_agg.items():
        tot = sum(cnt.values())
        ranking_parque.append((p, cnt["MOBILIZAÇÃO"], cnt["DESMOBILIZAÇÃO"], cnt["CANCELAMENTO"], tot))
    ranking_parque.sort(key=lambda x: -x[4])
    ranking_parque = ranking_parque[:10]

    cliente_agg = defaultdict(lambda: Counter())
    for r in rows:
        if r["cliente"]:
            cliente_agg[r["cliente"]][r["desc"]] += 1
    ranking_cliente = []
    for p, cnt in cliente_agg.items():
        tot = sum(cnt.values())
        ranking_cliente.append((p, cnt["MOBILIZAÇÃO"], cnt["DESMOBILIZAÇÃO"], cnt["CANCELAMENTO"], tot))
    ranking_cliente.sort(key=lambda x: -x[4])

    month_agg = defaultdict(lambda: Counter())
    for r in rows:
        d = r["data"]
        if not isinstance(d, datetime.datetime):
            continue
        month_agg[(d.year, d.month)][r["desc"]] += 1
    months_sorted = sorted(month_agg.keys())
    meses_labels = [f"{MESES_PT[m-1]}/{str(y)[2:]}" for (y, m) in months_sorted]
    mob_arr = [month_agg[k]["MOBILIZAÇÃO"] for k in months_sorted]
    desmob_arr = [month_agg[k]["DESMOBILIZAÇÃO"] for k in months_sorted]
    canc_arr = [month_agg[k]["CANCELAMENTO"] for k in months_sorted]

    def bucket(motivo):
        if not motivo:
            return "OUTRO"
        m = motivo.upper()
        if m.startswith("SOLICITAÇÃO CLIENTE"):
            return "CLIENTE"
        return "EXTREME"  # inclui "SOLICITAÇÃO EXTREME" e "SOLICITAÇÃO LEAN WAY" (parceiro operacional)

    motivo_month = defaultdict(lambda: Counter())
    for r in rows:
        d = r["data"]
        if not isinstance(d, datetime.datetime):
            continue
        motivo_month[(d.year, d.month)][bucket(r["motivo"])] += 1
    motivo_cliente_arr = [motivo_month[k]["CLIENTE"] for k in months_sorted]
    motivo_extreme_arr = [motivo_month[k]["EXTREME"] for k in months_sorted]

    return dict(
        parques=parques_list, tot_real=tot_real, tot_plano=tot_plano,
        tot_real_fisico=tot_real_fisico, em_desmobilizacao=em_desmobilizacao,
        parques_ativos=parques_ativos, parques_total=parques_total,
        tecnicos_ativos=tecnicos_ativos,
        n_mob=n_mob, n_desmob=n_desmob, n_canc=n_canc, n_total=n_total,
        ranking_parque=ranking_parque, ranking_cliente=ranking_cliente,
        meses_labels=meses_labels, mob_arr=mob_arr, desmob_arr=desmob_arr, canc_arr=canc_arr,
        motivo_cliente_arr=motivo_cliente_arr, motivo_extreme_arr=motivo_extreme_arr,
    )


def render_equipes(html, d):
    hoje = datetime.date.today().strftime('%d/%m/%Y')

    def sub_once(pattern, replacement, text, flags=re.S):
        new_text, n = re.subn(pattern, replacement, text, count=1, flags=flags)
        if n == 0:
            log(f"  [aviso] padrao nao encontrado (equipes): {pattern[:60]}...")
        return new_text

    # subtitulo
    html = sub_once(r'(Diario de Bordo - dados consolidados em )\d{2}/\d{2}/\d{4}', lambda m: m.group(1) + hoje, html)
    html = sub_once(r'(Painel 1: CONTROLE DE EQUIPES\. Painel 2: DADOS/RESUMOS \(aba DASHBOARD original\)\. )\d{2}/\d{2}/\d{4}', lambda m: m.group(1) + hoje, html)

    # "Tecnicos ativos" = presenca fisica atual (inclui quem ja esta em
    # desmobilizacao mas ainda nao viajou). "Planejado (headcount)" ja
    # desconta quem tem saida decidida. O deficit compara o headcount-alvo
    # (ja ajustado) com quem efetivamente conta como ativo daqui pra frente.
    deficit = d["tot_real"] - d["tot_plano"]
    n_em_desmob = len(d.get("em_desmobilizacao", []))
    if n_em_desmob:
        datas = sorted({t["data_finalizacao"] for t in d["em_desmobilizacao"] if t.get("data_finalizacao")})
        datas_br = ", ".join(datetime.datetime.strptime(dd, "%Y-%m-%d").strftime("%d/%m/%Y") for dd in datas)
        tecnicos_sub = f'{n_em_desmob} em desmobilizacao ate {datas_br}'
        desmob_kpi = f'''
  <div class="kpi"><div class="label">Em desmobilizacao</div><div class="value amber">{n_em_desmob}</div><div class="sub">saida ja decidida, viagem ate {datas_br}</div></div>'''
    else:
        tecnicos_sub = "em campo agora"
        desmob_kpi = ""
    kpis_ativos = f'''<div class="kpi"><div class="label">Tecnicos ativos</div><div class="value">{d["tot_real_fisico"]}</div><div class="sub">{tecnicos_sub}</div></div>
  <div class="kpi"><div class="label">Planejado (headcount)</div><div class="value">{d["tot_plano"]}</div><div class="sub">meta de equipes</div></div>
  <div class="kpi"><div class="label">Deficit total</div><div class="value red">{deficit}</div><div class="sub">abaixo do plano</div></div>
  <div class="kpi"><div class="label">Parques ativos</div><div class="value">{d["parques_ativos"]}<span class="sub" style="font-size:16px;"> / {d["parques_total"]}</span></div><div class="sub">{d["parques_total"]-d["parques_ativos"]} parques vazios</div></div>{desmob_kpi}
</div>'''
    html = sub_once(r'<div class="kpi"><div class="label">Tecnicos ativos</div>.*?(?=\n<div class="grid)', kpis_ativos, html)

    def js_tecnicos(items):
        parts = []
        for t in items:
            mat_js = js_str(t["mat"]) if t["mat"] is not None else "null"
            acum_js = str(t["acumulado"]) if t["acumulado"] is not None else "null"
            parts.append("  {mat:%s, nome:%s, acumulado:%s, parque:%s, suop:%s}" % (
                mat_js, js_str(t["nome"]), acum_js, js_str(t["parque"]), js_str(t["suop"])))
        return "[\n" + ",\n".join(parts) + "\n]"

    new_tecnicos_js = "var tecnicosAtivos = " + js_tecnicos(d["tecnicos_ativos"]) + ";"
    html = sub_once(r'var tecnicosAtivos = \[.*?\];', new_tecnicos_js, html)

    pct = lambda n: round(100*n/d["n_total"], 1) if d["n_total"] else 0
    kpis_hist = f'''<div class="kpi"><div class="label">Mobilizacoes</div><div class="value">{d["n_mob"]}</div><div class="sub">{pct(d["n_mob"])}% dos eventos</div></div>
  <div class="kpi"><div class="label">Desmobilizacoes</div><div class="value">{d["n_desmob"]}</div><div class="sub">{pct(d["n_desmob"])}% dos eventos</div></div>
  <div class="kpi"><div class="label">Cancelamentos</div><div class="value amber">{d["n_canc"]}</div><div class="sub">{pct(d["n_canc"])}% dos eventos</div></div>
  <div class="kpi"><div class="label">Total de eventos</div><div class="value">{d["n_total"]}</div><div class="sub">jan/24 a {d["meses_labels"][-1].lower() if d["meses_labels"] else "?"}</div></div>'''
    html = sub_once(r'<div class="kpi"><div class="label">Mobilizacoes</div>.*?jan/24 a jul/26</div></div>', kpis_hist, html)

    def js_arr_objs(items):
        parts = []
        for p in items:
            parts.append("  {nome:%s, cliente:%s, suop:%s, real:%d, plano:%d}" % (
                js_str(p["nome"]), js_str(p["cliente"]), js_str(p["suop"]), p["real"], p["plano"]))
        return "[\n" + ",\n".join(parts) + "\n]"

    new_parques_js = "var parques = " + js_arr_objs(d["parques"]) + ";"
    html = sub_once(r'var parques = \[.*?\];', new_parques_js, html)
    html = sub_once(r'var emDesmobQtd = \d+;', f'var emDesmobQtd = {len(d.get("em_desmobilizacao", []))};', html)

    def js_ranking(items):
        parts = [f"  [{js_str(p[0])},{p[1]},{p[2]},{p[3]}]" for p in items]
        return "[\n" + ",\n".join(parts) + "\n]"

    html = sub_once(r'var rankingParque = \[.*?\];', "var rankingParque = " + js_ranking(d["ranking_parque"]) + ";", html)
    html = sub_once(r'var rankingCliente = \[.*?\];', "var rankingCliente = " + js_ranking(d["ranking_cliente"]) + ";", html)

    # Grafico "Evolucao mensal de eventos" foi removido (nao existe mais mob/desmob/canc no HTML).
    # "mesesTodos"/"motivoClienteTodos"/"motivoExtremeTodos" alimentam o grafico "Solicitacoes por
    # motivo", que exibe so a janela dos ultimos 12 meses + mes atual (fatiada em JS a partir destes
    # arrays completos) -- por isso continuamos atualizando os arrays "Todos" inteiros aqui.
    html = sub_once(r"var mesesTodos = \[.*?\];", "var mesesTodos = [" + ",".join(js_str(m) for m in d["meses_labels"]) + "];", html)
    html = sub_once(r"var motivoClienteTodos = \[.*?\];", "var motivoClienteTodos = [" + ",".join(str(x) for x in d["motivo_cliente_arr"]) + "];", html)
    html = sub_once(r"var motivoExtremeTodos = \[.*?\];", "var motivoExtremeTodos = [" + ",".join(str(x) for x in d["motivo_extreme_arr"]) + "];", html)

    # ---- Busca por tecnico (3a aba) ----
    tecnicos_busca = d.get("tecnicos_busca")
    if tecnicos_busca is not None:
        html = sub_once(r"var TECNICOS_DATA = \[.*?\];\n?", "var TECNICOS_DATA = " + json.dumps(tecnicos_busca, ensure_ascii=False) + ";\n", html)
        html = sub_once(r"var CLIENTES_TEC = \[.*?\];", "var CLIENTES_TEC = " + json.dumps(CLIENTES_TEC, ensure_ascii=False) + ";", html)

    return html


# ============================================================
# BUSCA POR TECNICO (aba "RESUMOS TECNICOS" da planilha Diario de Bordo,
# recalculada aqui a partir da aba bruta "DADOS" -- mesma logica da planilha:
# pareia cada DESMOBILIZACAO com a MOBILIZACAO mais recente da mesma matricula
# (linhas anteriores, em ordem da planilha) pra achar dias mobilizado, e
# marca observacao negativa quando MOTIVO+OBSERVACOES bate com uma das
# palavras-chave da lista AQ1:AQ14 da planilha.
# ============================================================

# "PROBLEMA" e "DESLIGAMENTO" foram removidas dessa lista a pedido do Robson --
# "PROBLEMA" pegava falsos positivos genericos demais (ex: "solicitou desmob
# para resolver problemas pessoais", que NAO e historico negativo, so um
# desligamento por motivo pessoal legitimo), e "DESLIGAMENTO" sozinho tambem
# aparece como categoria administrativa de rotina (ex: "FOLGA + DESLIGAMENTO")
# sem indicar nenhuma conduta ruim. So entram palavras que descrevem conduta
# genuinamente negativa (abandono, indisciplina, briga/conflito interpessoal,
# comportamento, desconfianca, etc.) -- ver tambem o filtro por tipo de evento
# logo abaixo (so avalia linhas de DESMOBILIZACAO/CANCELAMENTO, nunca de
# MOBILIZACAO, porque "mobilizado para atender/substituir" nunca e historico
# negativo do tecnico que esta sendo mobilizado).
NEGATIVE_KEYWORDS_TEC = [
    "DESCONFIAN", "COMPORTAMENT", "NÃO RETORNOU", "BRIGA",
    "INDISCIPLINA", "ADVERTÊNCIA", "SUSPEITA", "FURTO", "AGRESS",
    "EMBRIAG", "DEMITID", "INEFICI", "INTERPESSOAL", "DESENTENDIMENTO",
    "ABANDON", "PARALIS",
]

# Palavras genericas que podem aparecer logo depois de "ABANDONO DE/DO/DA"
# sem se referir a uma PESSOA (ex: "ABANDONO DE PROJETO", "ABANDONO DE
# EMPREGO") -- nesses casos o abandono e do proprio tecnico da linha, entao
# a observacao negativa continua valendo pra ele normalmente.
NOMES_GENERICOS_ABANDONO = {
    "PROJETO", "EMPREGO", "EQUIPE", "TRABALHO", "PARQUE", "ATIVIDADE",
    "CONTRATO", "POSTO", "FUNCAO", "FUNCÃO", "CARGO", "OBRA",
}


def _strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def _abandono_de_outro_tecnico(texto, nome_proprio, primeiros_nomes_conhecidos):
    """Detecta o BUG reportado pelo Robson: uma linha de DESMOBILIZACAO as
    vezes explica a causa como "DEVIDO ABANDONO DO <NOME-DE-OUTRO-TECNICO>"
    (ex: "DESMOBILIZACAO A PEDIDO DO CLIENTE DEVIDO ABANDONO DO LUCENÍLIO
    N2") -- o tecnico desta linha (ex: Felipe Serpa) so foi afetado
    indiretamente pela saida de outra pessoa, ele NAO tem culpa nenhuma, entao
    isso nao pode virar uma observacao negativa dele. Se compararmos o nome
    citado com a lista de primeiros-nomes de todos os tecnicos da planilha e
    ele bater com um nome DIFERENTE do proprio tecnico da linha (e nao for uma
    palavra generica tipo PROJETO/EMPREGO), tratamos como "abandono de
    terceiro" e suprimimos a observacao desta linha.
    So olha o padrao "ABANDONO DE/DO/DA <PALAVRA>" (nao usa o padrao mais
    generico "TECNICO <NOME>", que aparece o tempo todo em texto de
    MOBILIZACAO/substituicao legitimo e geraria falsos positivos)."""
    m = re.search(r"ABANDON[OA]\s+D[AOE]\s+([A-ZÀ-Ü]+)", texto)
    if not m:
        return False
    candidato = _strip_accents(m.group(1))
    if candidato in NOMES_GENERICOS_ABANDONO:
        return False
    proprio_primeiro = _strip_accents(nome_proprio.upper().split()[0]) if nome_proprio else ""
    if candidato == proprio_primeiro:
        return False
    return candidato in primeiros_nomes_conhecidos
# Ordem fixa das colunas de cliente na tabela "por ano e por cliente", igual
# ao cabecalho da linha 18 da aba RESUMOS TECNICOS.
CLIENTES_TEC = ["GE", "SIEMENS", "NORDEX / LWS", "VESTAS", "ELETROBRAS",
                "EÓLICAS BABILÔNIA", "LM", "SINOMA BLADES"]


def _norm_nome_tec(s):
    """Normaliza nome pra usar como chave de agrupamento (mesma pessoa pode
    aparecer com matriculas diferentes na planilha -- ex: recontratacao gera
    um novo MAT para o mesmo tecnico). So colapsa espacos/maiusculas, nao
    mexe em acentos (nomes com grafias realmente diferentes NAO devem ser
    unificados)."""
    return re.sub(r"\s+", " ", s.strip().upper())


def _split_days_by_year(start, end):
    """Divide o intervalo [start, end) em dias corridos por ano-civil. Usado
    tanto para ciclos MOBILIZACAO->DESMOBILIZACAO ja fechados quanto para o
    streak em aberto (mobilizado ate hoje, sem DESMOBILIZACAO ainda) -- em
    ambos os casos o periodo pode atravessar a virada do ano (ex: mobilizado
    em nov/2025, ainda ativo em 2026), e os dias de cada lado precisam contar
    no ano certo em vez de tudo ir pro ano do inicio ou do fim do intervalo."""
    out = defaultdict(int)
    if not start or not end or end <= start:
        return out
    cur = start
    while cur < end:
        proximo_ano = datetime.datetime(cur.year + 1, 1, 1)
        fim_trecho = min(end, proximo_ano)
        out[cur.year] += (fim_trecho - cur).days
        cur = fim_trecho
    return out


def build_tecnicos_data():
    wb = openpyxl.load_workbook(EQUIPES_XLSX, data_only=True, read_only=True)
    ws = wb['DADOS']
    grid = sheet_to_grid(ws)

    hoje = datetime.datetime.now()
    rows = []
    for r in range(4, ws.max_row + 1):
        mat = grid.get((r, 2))
        nome = grid.get((r, 3))
        if mat is None or not nome:
            continue
        classificacao = grid.get((r, 4))
        data = grid.get((r, 5))
        desc = grid.get((r, 7))
        parque = grid.get((r, 10))
        cliente = grid.get((r, 11))
        motivo = grid.get((r, 12))
        obs = grid.get((r, 14))
        rows.append(dict(
            mat=str(mat).strip(), nome=str(nome).strip(),
            classificacao=str(classificacao).strip().upper() if isinstance(classificacao, str) else "",
            data=data if isinstance(data, datetime.datetime) else None,
            desc=str(desc).strip().upper() if isinstance(desc, str) else "",
            parque=str(parque).strip() if isinstance(parque, str) else "",
            cliente=str(cliente).strip() if isinstance(cliente, str) else "",
            motivo=str(motivo).strip() if isinstance(motivo, str) else "",
            obs=str(obs).strip() if isinstance(obs, str) else "",
        ))

    # Conjunto de primeiros-nomes de TODOS os tecnicos da planilha, usado so
    # pra diferenciar "ABANDONO DO <PROJETO/EMPREGO>" (auto-causado, fica) de
    # "ABANDONO DO <NOME-DE-OUTRO-TECNICO>" (bug reportado pelo Robson --
    # nao pode virar observacao negativa de quem so foi afetado de tabela).
    primeiros_nomes_conhecidos = set()
    for row in rows:
        if row["nome"]:
            partes = row["nome"].upper().split()
            if partes:
                primeiros_nomes_conhecidos.add(_strip_accents(partes[0]))

    # ---- Passo 1: agrega por MATRICULA (pareamento MOBILIZACAO/DESMOBILIZACAO
    # so faz sentido dentro da mesma matricula, igual a formula original da
    # planilha, que usa AB=matricula como criterio do MAXIFS). ----
    tecnicos = {}
    last_mob = {}  # mat -> data da MOBILIZACAO mais recente ja vista (varredura em ordem da planilha)
    mob_aberta = {}  # mat -> {"data":..., "cliente":...} da MOBILIZACAO ainda sem DESMOBILIZACAO seguinte (streak atual em andamento)
    for row in rows:
        mat = row["mat"]
        t = tecnicos.setdefault(mat, {
            "mat": mat, "nome": row["nome"],
            "dias_ano": defaultdict(int),
            "dias_ano_cliente": defaultdict(lambda: defaultdict(int)),
            "negativos": [],
            "classificacao_atual": "",
        })
        t["nome"] = row["nome"]  # ultimo nome visto (nomes nao deveriam mudar, mas por seguranca)
        if row["classificacao"] in ("N1", "N2", "N3"):
            t["classificacao_atual"] = row["classificacao"]

        if row["desc"] == "MOBILIZAÇÃO":
            last_mob[mat] = row["data"]
            # assume em aberto ate ver uma DESMOBILIZACAO depois -- guardamos o
            # cliente tambem, pra poder atribuir os dias do streak em aberto
            # (ver bug corrigido abaixo) ao cliente certo em "dias_ano_cliente".
            mob_aberta[mat] = {"data": row["data"], "cliente": row["cliente"] or "-"}
        elif row["desc"] == "DESMOBILIZAÇÃO":
            mob_data = last_mob.get(mat)
            if mob_data and row["data"] and row["data"] >= mob_data:
                # BUG corrigido: antes, o ciclo inteiro (mobilizacao ate
                # desmobilizacao) era jogado 100% no ano da DESMOBILIZACAO --
                # um tecnico mobilizado em nov/2025 e desmobilizado em fev/2026
                # aparecia com 0 dias em 2025 e TODOS os dias em 2026 (ou
                # vice-versa se o corte fosse no sentido contrario). Agora
                # dividimos os dias corridos pelo ano-civil de cada trecho.
                cliente = row["cliente"] or "-"
                for ano, dias in _split_days_by_year(mob_data, row["data"]).items():
                    if dias <= 0:
                        continue
                    t["dias_ano"][ano] += dias
                    t["dias_ano_cliente"][ano][cliente] += dias
            mob_aberta[mat] = None  # fechou o ciclo -- nao esta mais "em campo agora"

        # So avalia historico negativo em linhas de DESMOBILIZACAO/CANCELAMENTO --
        # uma linha de MOBILIZACAO descreve o motivo de OUTRA pessoa ter saido
        # (o tecnico sendo mobilizado aqui e o substituto, nao quem tem o
        # historico ruim), entao nunca deve gerar observacao negativa para ele.
        if row["desc"] in ("DESMOBILIZAÇÃO", "CANCELAMENTO"):
            texto = (row["motivo"] + " " + row["obs"]).upper()
            tem_keyword = any(kw in texto for kw in NEGATIVE_KEYWORDS_TEC)
            # BUG corrigido (pedido do Robson): "DESMOBILIZACAO ... DEVIDO
            # ABANDONO DO LUCENÍLIO N2" na linha do Felipe Serpa/Oseias nao e
            # culpa deles -- e do Lucenilio, citado pelo nome no texto. So
            # aplica a observacao negativa nesta linha se NAO for esse caso.
            eh_culpa_de_outro = _abandono_de_outro_tecnico(texto, row["nome"], primeiros_nomes_conhecidos)
            if tem_keyword and not eh_culpa_de_outro:
                t["negativos"].append({
                    "data": fmt_date_br(row["data"]) if row["data"] else "",
                    "data_sort": row["data"].strftime("%Y-%m-%d") if row["data"] else "",
                    "parque": row["parque"],
                    "motivo": (row["motivo"] + (" - " + row["obs"] if row["obs"] else "")).strip(" -"),
                })

    # Dias "em carreira" pra ranking: soma dos ciclos ja fechados (dias_ano)
    # mais, se a matricula ainda esta em campo agora (ultima MOBILIZACAO sem
    # DESMOBILIZACAO seguinte), os dias corridos desde essa mobilizacao ate
    # hoje -- senao um tecnico mobilizado ha 200 dias e ainda em campo
    # apareceria com "0 dias" no ranking so por nao ter sido desmobilizado
    # ainda.
    #
    # BUG corrigido: esses dias do streak em aberto entravam no total do
    # ranking (dias_total_mat, abaixo) mas NUNCA em dias_ano/dias_ano_cliente
    # -- por isso um tecnico ainda mobilizado (sem DESMOBILIZACAO lancada)
    # aparecia certo no ranking geral mas com "0 dias em 2026" no historico
    # por ano da busca por tecnico (ex.: Felipe de Lima Serpa, mobilizado
    # em 27/12/2025 e ainda ativo -- 806 dias corretos no ranking, 0 em 2026
    # no detalhe). Agora o streak em aberto tambem e somado em dias_ano,
    # dividido por ano-civil igual aos ciclos ja fechados.
    for mat, t in tecnicos.items():
        dias_fechados = sum(t["dias_ano"].values())
        aberta = mob_aberta.get(mat)
        dias_streak = 0
        if aberta:
            for ano, dias in _split_days_by_year(aberta["data"], hoje).items():
                if dias <= 0:
                    continue
                t["dias_ano"][ano] += dias
                t["dias_ano_cliente"][ano][aberta["cliente"]] += dias
                dias_streak += dias
        t["dias_total_mat"] = dias_fechados + max(0, dias_streak)

    # ---- Passo 2: unifica por NOME (mesmo tecnico pode ter mais de uma
    # matricula ao longo do tempo -- ex: desligamento/recontratacao). A busca
    # e feita por nome, entao o resultado precisa somar os dados de todas as
    # matriculas daquela pessoa num unico registro, em vez de mostrar
    # "2 tecnicos" separados para quem e a mesma pessoa. ----
    grupos = {}
    for mat, t in tecnicos.items():
        chave = _norm_nome_tec(t["nome"])
        g = grupos.setdefault(chave, {
            "nome": t["nome"], "mats": set(),
            "dias_ano": defaultdict(int),
            "dias_ano_cliente": defaultdict(lambda: defaultdict(int)),
            "negativos": [],
            "dias_total": 0,
            "classificacao": "",
        })
        g["mats"].add(mat)
        for ano, dias in t["dias_ano"].items():
            g["dias_ano"][ano] += dias
        for ano, por_cliente in t["dias_ano_cliente"].items():
            for cliente, dias in por_cliente.items():
                g["dias_ano_cliente"][ano][cliente] += dias
        g["negativos"].extend(t["negativos"])
        g["dias_total"] += t["dias_total_mat"]
        if t["classificacao_atual"]:
            g["classificacao"] = t["classificacao_atual"]  # matricula mais recente processada "vence"

    out = []
    for chave, g in grupos.items():
        negativos = sorted(g["negativos"], key=lambda x: x["data_sort"])
        for n in negativos:
            n.pop("data_sort", None)
        mats_sorted = sorted(g["mats"], key=lambda m: (0, int(m)) if m.isdigit() else (1, m))
        out.append({
            "mat": mats_sorted[0], "mats": mats_sorted, "nome": g["nome"],
            "dias_ano": {str(ano): dias for ano, dias in sorted(g["dias_ano"].items())},
            "dias_ano_cliente": {str(ano): dict(cl) for ano, cl in sorted(g["dias_ano_cliente"].items())},
            "negativos": negativos,
            "dias_total": g["dias_total"],
            "classificacao": g["classificacao"] or None,
        })
    out.sort(key=lambda t: t["nome"])
    n_unificados = sum(1 for t in out if len(t["mats"]) > 1)
    log(f"  Tecnicos (busca): {len(out)} pessoas mapeadas ({n_unificados} unificadas por terem mais de 1 matricula), "
        f"{sum(len(t['negativos']) for t in out)} observacoes negativas no total.")
    return out


# ============================================================
# FROTA
# ============================================================

CATS_HIST = ["CHECKLIST", "REVISÃO", "ESTEPE EXTRA", "MEDIÇÃO PNEUS", "RASTREADOR"]
# Categoria completa (10 colunas do bloco semanal, na ordem em que aparecem na
# planilha) -- usada na tabela "Conformidade por Categoria" e no historico
# completo. CATS_HIST (5 categorias) continua sendo o criterio oficial da
# "Media Geral", igual a planilha ja usa.
CATS_ALL = ["KM PONTO DE TROCA", "SALDO EXTRA", "CHECKLIST", "REVISÃO", "ESTEPE EXTRA",
            "MEDIÇÃO PNEUS", "RASTREADOR", "CONSUMO", "DOC POSTADO", "DOC ENTREGUE"]
CATS_ALL_LABELS = ["KM Ponto de Troca", "Saldo Extra", "Checklist", "Revisão", "Estepe Extra",
                    "Medição Pneus", "Rastreador + Bloqueador", "Consumo KM/L",
                    "Doc Postado (MOB)", "Doc Entregue Cliente"]


def parse_resumo_oficial(grid, max_row):
    """Le as tabelas "RESUMO" que a propria planilha calcula dentro de cada
    bloco semanal (coluna J = 'RESUMO', linhas 'WEEK NN' com os 5 percentuais
    classicos + Media Geral, ja calculados por formula na planilha). Blocos
    mais recentes repetem/corrigem semanas ja vistas em blocos anteriores --
    varremos em ordem crescente e deixamos a ULTIMA ocorrencia prevalecer,
    assumindo que e a versao mais atualizada/corrigida.
    Cobre normalmente as semanas mais recentes (a planilha so passou a manter
    essa tabela a partir de um certo ponto); semanas mais antigas nao aparecem
    aqui e ficam por conta do calculo bruto (contagem OK/N-OK) em block_pct."""
    out = {}
    for r in range(1, max_row + 1):
        if grid.get((r, 10)) != 'RESUMO':
            continue
        hdr = r + 1
        if grid.get((hdr, 10)) != 'SEMANA':
            continue
        rr = hdr + 1
        while True:
            lbl = grid.get((rr, 10))
            if not isinstance(lbl, str):
                break
            m = re.match(r'^WEEK\s*(\d+)$', lbl.strip().upper())
            if not m:
                break
            wk = int(m.group(1))
            vals = {
                "checklist": grid.get((rr, 11)), "revisao": grid.get((rr, 12)),
                "estepe": grid.get((rr, 13)), "rastreador": grid.get((rr, 14)),
                "pneus": grid.get((rr, 15)), "geral": grid.get((rr, 16)),
            }
            if all(isinstance(v, (int, float)) for v in vals.values()):
                out[wk] = {k: round(v * 100, 1) for k, v in vals.items()}
            rr += 1
    return out


def classificar_acoes_por_categoria(acao_rows):
    """Classifica cada linha do Plano de Acao em Emergencial / Programada /
    Administrativa por palavra-chave no texto (acao+obs+tipo), regra definida
    pelo Robson em 18/08/2026 (documentada no REGRAS_AUTOMACAO.md, mas nunca
    implementada no codigo antes de 22/08/2026 -- confirmado com o Robson
    antes de escrever esta funcao):
      - Emergencial: menciona pneu, checklist, revisao ou freio.
      - Programada: menciona km, saldo (estourado/contrato) ou "40 mil"/"40k".
      - Administrativa: tudo que sobrar (nao casa com nenhuma das duas acima).
    Quando uma acao menciona motivos das DUAS categorias ao mesmo tempo (ex.:
    "40K DE KM RODADOS/SALDO ESTOURADO/MEDIÇÃO PNEU N/OK"), e desmembrada em
    uma linha por categoria -- cada linha desmembrada mostra o prefixo
    "[Emergencial: <motivo>]" ou "[Programada: <motivo>]" no campo "tipo"
    exibido, para ficar claro qual motivo aquela linha representa.
    Isto e usado so como FALLBACK por linha (a planilha nao guarda a
    categoria oficial por linha, so o total agregado Emergencial/Programado
    que already é lido em outro lugar como n_emerg_sheet/n_prog_sheet)."""
    RE_EMERG = re.compile(r'pneu|checklist|revis[aã]o|freio', re.I)
    RE_PROG = re.compile(r'\bkm\b|saldo|40\s*mil|40k', re.I)

    out = []
    for a in acao_rows:
        texto = f'{a.get("acao","")} {a.get("obs","")} {a.get("tipo","")}'
        motivos_emerg = sorted(set(m.group(0).strip().upper() for m in RE_EMERG.finditer(texto)))
        motivos_prog = sorted(set(m.group(0).strip().upper() for m in RE_PROG.finditer(texto)))

        if motivos_emerg and motivos_prog:
            row_e = dict(a)
            row_e["categoria"] = "Emergencial"
            row_e["tipo"] = f'[Emergencial: {", ".join(motivos_emerg)}] {a.get("tipo","")}'.strip()
            out.append(row_e)
            row_p = dict(a)
            row_p["categoria"] = "Programada"
            row_p["tipo"] = f'[Programada: {", ".join(motivos_prog)}] {a.get("tipo","")}'.strip()
            out.append(row_p)
        elif motivos_emerg:
            row = dict(a)
            row["categoria"] = "Emergencial"
            out.append(row)
        elif motivos_prog:
            row = dict(a)
            row["categoria"] = "Programada"
            out.append(row)
        else:
            row = dict(a)
            row["categoria"] = "Administrativa"
            out.append(row)
    return out


def build_frota_data():
    wb = openpyxl.load_workbook(FROTA_XLSX, data_only=True, read_only=True)
    ws = wb['BASE DASHBOARD']
    grid = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for c in row:
            if c.value is not None:
                grid[(c.row, c.column)] = c.value

    def get(r, c):
        return grid.get((r, c))

    # localizar blocos semanais pela label "WK NN" ou "WEEK NN" na coluna 20
    # (as duas grafias aparecem na planilha, ex.: "WK 33" e "WEEK 34" --
    # corrigido em 22/08/2026: antes so reconhecia "WK", entao um bloco
    # rotulado "WEEK NN" ficava invisivel ao detector e seu conteudo era lido
    # como se ainda pertencesse ao ultimo bloco "WK" valido anterior, fazendo
    # o dashboard "congelar" numa semana antiga mesmo com uma semana mais
    # nova ja lancada na planilha -- ver REGRAS_AUTOMACAO.md).
    # a mesma label pode aparecer 2x dentro do mesmo bloco (uma segunda mencao
    # mais abaixo) -- so conta como novo bloco quando o texto da label muda.
    WK_LABEL_RE = re.compile(r'^(?:WK|WEEK)\s*(\d+)$', re.I)
    raw_hits = []
    for (r, c), v in grid.items():
        if c == 20 and isinstance(v, str) and WK_LABEL_RE.match(v.strip()):
            raw_hits.append((r, v.strip()))
    raw_hits.sort(key=lambda x: x[0])
    block_starts = []
    block_labels = []
    last_label = None
    for r, lbl in raw_hits:
        if lbl != last_label:
            block_starts.append(r)
            block_labels.append(lbl)
            last_label = lbl
    block_ends = block_starts[1:] + [ws.max_row + 1]
    n_blocks = len(block_starts)
    # numeramos sequencialmente a partir do NUMERO DO ULTIMO BLOCO (ancora confiavel:
    # e a semana corrente, com data de "ATUALIZAÇÃO" explicita ao lado) -- mesmo
    # criterio que o dashboard original ja usava para as semanas WK12-WK14 com
    # rotulo ambiguo na planilha (inferencia pela posicao sequencial).
    last_label = block_labels[-1]
    m_last = WK_LABEL_RE.match(last_label)
    last_wk_num = int(m_last.group(1)) if m_last else None
    if last_wk_num is None or last_wk_num < n_blocks:
        # rotulo do ultimo bloco tambem suspeito -- nao ha ancora confiavel.
        log(f"  [risco] rotulo do bloco mais recente ({last_label!r}) nao pode ser usado como ancora. Assumindo numeracao sequencial a partir de WK08.")
        week_numbers = list(range(8, 8 + n_blocks))
    else:
        week_numbers = list(range(last_wk_num - n_blocks + 1, last_wk_num + 1))
    uncertain_weeks = set()
    for wk, lbl in zip(week_numbers, block_labels):
        m = WK_LABEL_RE.match(lbl)
        raw_num = int(m.group(1)) if m else None
        if raw_num != wk:
            uncertain_weeks.add(wk)
            log(f"  [risco] bloco na linha correspondente a WK{wk:02d} tinha rotulo original {lbl!r} (inconsistente) -- numero inferido pela posicao.")

    def find_headers(start, end):
        catrow = parquerow = None
        for r in range(start, min(start + 20, end)):
            for c in range(1, 25):
                v = get(r, c)
                if isinstance(v, str):
                    vu = v.strip().upper()
                    if vu.startswith("CHECKLIST") and catrow is None:
                        catrow = r
                    if vu == "PARQUE" and parquerow is None:
                        parquerow = r
        return catrow, parquerow

    def block_pct(start, end):
        catrow, parquerow = find_headers(start, end)
        if not catrow or not parquerow:
            return None, 0
        catcols = {}
        for c in range(1, 25):
            v = get(catrow, c)
            if isinstance(v, str):
                vu = v.strip().upper()
                for cat in CATS_ALL:
                    if vu.startswith(cat) and cat not in catcols:
                        catcols[cat] = c
        parque_col = None
        for c in range(1, 12):
            v = get(parquerow, c)
            if isinstance(v, str) and v.strip().upper() == "PARQUE":
                parque_col = c
                break
        if not parque_col:
            return None, 0
        oks = {c: 0 for c in CATS_ALL}
        tots = {c: 0 for c in CATS_ALL}
        nveh = 0
        STATUS_TOKENS = {"OK", "N/OK", "ATENÇÃO", "ATENCAO", "-"}
        scan_start = max(catrow, parquerow) + 1
        for r in range(scan_start, end):
            pval = get(r, parque_col)
            if not isinstance(pval, str) or not pval.strip():
                continue
            statuses = [get(r, cc) for cc in catcols.values()]
            if not any(isinstance(s, str) and s.strip().upper() in STATUS_TOKENS for s in statuses):
                continue
            nveh += 1
            for cat, cc in catcols.items():
                v = get(r, cc)
                if isinstance(v, str):
                    vv = v.strip().upper()
                    if vv == "OK":
                        oks[cat] += 1; tots[cat] += 1
                    elif vv in ("N/OK", "ATENÇÃO", "ATENCAO"):
                        tots[cat] += 1
        pct = {c: (round(100 * oks[c] / tots[c], 1) if tots[c] else None) for c in CATS_ALL}
        return pct, nveh

    resumo_oficial = parse_resumo_oficial(grid, ws.max_row)

    hist = []
    for wk, start, end in zip(week_numbers, block_starts, block_ends):
        pct, nveh = block_pct(start, end)
        cats_vals = [pct[c] for c in CATS_ALL] if pct else [None] * len(CATS_ALL)
        media_vals = [pct[c] for c in CATS_HIST] if pct else [None] * 5
        media_bruta = round(sum(v for v in media_vals if v is not None) / max(1, len([v for v in media_vals if v is not None])), 1) if pct and any(v is not None for v in media_vals) else None
        oficial = resumo_oficial.get(wk)
        if oficial is not None:
            # A propria planilha ja calcula esses 5 + a Media Geral numa
            # tabela "RESUMO" por formula -- preferimos esse valor ao nosso
            # recalculo bruto (que pode divergir por causa de arredondamento/
            # denominador diferente). KM Troca/Saldo Extra/Consumo/Docs nao
            # tem tabela oficial, entao continuam vindo da contagem bruta.
            idx = {c: i for i, c in enumerate(CATS_ALL)}
            cats_vals[idx["CHECKLIST"]] = oficial["checklist"]
            cats_vals[idx["REVISÃO"]] = oficial["revisao"]
            cats_vals[idx["ESTEPE EXTRA"]] = oficial["estepe"]
            cats_vals[idx["MEDIÇÃO PNEUS"]] = oficial["pneus"]
            cats_vals[idx["RASTREADOR"]] = oficial["rastreador"]
            media = oficial["geral"]
            fonte = "oficial"
        else:
            media = media_bruta
            fonte = "bruto"
        hist.append(dict(wk=wk, cats=cats_vals, geral=media, nveh=nveh, fonte=fonte))

    # semanas com poucos veiculos (<10) tratadas como baixa confianca
    for h in hist:
        if h["nveh"] < 10:
            uncertain_weeks.add(h["wk"])

    # ---- semana mais recente: tabela de veiculos + plano de acao ----
    last_start, last_end = block_starts[-1], ws.max_row + 1
    # cabecalho ATUALIZACAO + data
    upd_date = None
    for r in range(last_start, min(last_start+10, last_end)):
        for c in range(1, 25):
            v = get(r, c)
            if isinstance(v, str) and v.strip().upper().startswith("ATUALIZAÇÃO"):
                # data normalmente algumas colunas a frente na mesma linha
                for cc in range(c, c+8):
                    dv = get(r, cc)
                    if isinstance(dv, datetime.datetime):
                        upd_date = dv
                break

    veh_rows = []
    for r in range(last_start, last_end):
        placa = get(r, 3)
        parque = get(r, 7)
        if not isinstance(placa, str) or not placa.strip():
            continue
        if not isinstance(parque, str) or not parque.strip():
            continue
        locadora = get(r, 4) or ""
        setor = get(r, 5)
        data_inicio = get(r, 8)
        prd = get(r, 9)
        cliente = get(r, 10) or ""
        vals = [get(r, cc) for cc in range(11, 21)]
        if not any(isinstance(v, str) and v.strip().upper() in ("OK", "N/OK", "ATENÇÃO", "-", "N/A") for v in vals):
            continue
        meses_str = f"{prd:.2f}".replace('.', ',') if isinstance(prd, (int, float)) else ""
        km_troca, saldo_extra, checklist, revisao, estepe, pneus, rastreador, consumo, doc_post, doc_ent = vals
        consumo_str = f"{consumo:.2f}".replace('.', ',') if isinstance(consumo, (int, float)) else str(consumo)
        veh_rows.append([
            placa.strip(), locadora.strip(), str(setor) if setor is not None else "",
            parque.strip(), cliente.strip() if isinstance(cliente,str) else str(cliente),
            fmt_date_br(data_inicio) if data_inicio else "", meses_str,
            km_troca or "", saldo_extra or "", checklist or "", revisao or "",
            estepe or "", pneus or "", rastreador or "", consumo_str,
            doc_post or "", doc_ent or "",
        ])

    n_localiza = sum(1 for v in veh_rows if v[1].upper().startswith("LOCALIZA"))
    n_unidas = sum(1 for v in veh_rows if v[1].upper().startswith("UNIDAS"))
    n_total_veic = len(veh_rows)

    # Mesmas 5 categorias oficiais usadas na "Media Geral" do grafico de
    # tendencia (CATS_HIST / parse_resumo_oficial) -- Checklist, Revisao,
    # Estepe Extra, Pneus, Rastreador+Bloqueador. Antes o card de Conformidade
    # Geral usava 9 colunas (incluindo Km Troca, Saldo Extra e Documentacao),
    # o que fazia o numero do card divergir do numero do grafico/planilha
    # oficial pra mesma semana -- alinhado a pedido do Robson.
    STATUS_COLS_IDX = [9, 10, 11, 12, 13]  # checklist, revisao, estepe, pneus, rastreador (indices em veh_rows)
    n_ok_geral = 0
    n_tot_geral = 0
    for v in veh_rows:
        for idx in STATUS_COLS_IDX:
            val = (v[idx] or "").strip().upper()
            if val in ("OK",):
                n_ok_geral += 1; n_tot_geral += 1
            elif val in ("N/OK", "ATENÇÃO", "ATENCAO"):
                n_tot_geral += 1
    conformidade_geral = round(100*n_ok_geral/n_tot_geral, 0) if n_tot_geral else None

    # Média de KM: media da coluna N ("Ultimo km registrado") da aba ACOMPANHAMENTO
    # (antes usava, por engano, a coluna R/"consumo" do bloco semanal em BASE DASHBOARD).
    kms = []
    if 'ACOMPANHAMENTO' in wb.sheetnames:
        ws_ac = wb['ACOMPANHAMENTO']
        for row in ws_ac.iter_rows(min_row=6, max_row=ws_ac.max_row, min_col=3, max_col=14):
            status = row[0].value   # coluna C
            placa = row[1].value    # coluna D
            n_val = row[11].value   # coluna N
            if not isinstance(placa, str) or not placa.strip():
                continue
            if isinstance(status, str) and status.strip().upper() not in ("ATIVO",):
                continue
            if isinstance(n_val, (int, float)):
                kms.append(n_val)
    media_km = round(sum(kms)/len(kms)) if kms else None

    # ---- plano de acao ----
    # A planilha organiza o Plano de Ação em secoes marcadas por linhas-divisoras
    # na coluna W (23): "PLANO DE AÇÃO WKxx" (cabecalho, com os totais oficiais de
    # EMERGENCIAL/PROGRAMADO gravados nas celulas ao lado), "AÇÕES ADMINISTRATIVAS",
    # "AÇÕES RESOLVIDAS" e "AÇÕES PARA MOBILIZAÇÃO/DESMOBILIZAÇÃO". Usamos essas
    # divisoras para classificar cada acao em Emergencial / Programado / Administrativa
    # em vez de inferir pelo texto do status (heuristica antiga, sujeita a erro).
    DIVIDER_LABELS = {
        "AÇÃO",
        "AÇÕES PARA MOBILIZAÇÃO/DESMOBILIZAÇÃO",
        "AÇÕES ADMINISTRATIVAS",
        "AÇÕES RESOLVIDAS",
    }

    def is_divider(acao_u):
        return acao_u in DIVIDER_LABELS or acao_u.startswith("PLANO DE AÇÃO")

    plano_row = admin_row = resolvidas_row = desmob_row = None
    for r in range(last_start, last_end):
        v = get(r, 23)
        if not isinstance(v, str):
            continue
        vu = v.strip().upper()
        if vu.startswith("PLANO DE AÇÃO") and plano_row is None:
            plano_row = r
        elif vu == "AÇÕES ADMINISTRATIVAS" and admin_row is None:
            admin_row = r
        elif vu == "AÇÕES RESOLVIDAS" and resolvidas_row is None:
            resolvidas_row = r
        elif vu == "AÇÕES PARA MOBILIZAÇÃO/DESMOBILIZAÇÃO" and desmob_row is None:
            desmob_row = r

    # limite da secao "pendente" (Emergencial+Programado): tudo entre o cabecalho
    # do plano e o inicio das Administrativas (ou Resolvidas, se nao houver bloco
    # administrativo separado nesta semana).
    pending_end = admin_row or resolvidas_row or last_end
    # limite da secao Administrativas: entre seu cabecalho e Resolvidas/Desmob.
    admin_end = resolvidas_row or desmob_row or last_end
    # limite da secao Resolvidas: entre seu cabecalho e Desmob (ou fim do bloco).
    resolvidas_end = desmob_row or last_end

    acao_rows = []
    admin_section_rows = 0
    resolvidas_section_rows = 0
    for r in range(last_start, last_end):
        acao = get(r, 23)
        if not isinstance(acao, str) or not acao.strip():
            continue
        acao_u = acao.strip().upper()
        if is_divider(acao_u):
            continue
        tipo = get(r, 24) or ""
        parque_a = get(r, 26) or get(r,25) or ""
        quem = get(r, 27) or ""
        status = get(r, 28)
        obs = get(r, 29) or ""
        if isinstance(status, datetime.datetime):
            status_str = fmt_date_br(status)
        else:
            status_str = str(status) if status is not None else ""
        acao_rows.append(dict(acao=acao.strip(), tipo=str(tipo).strip(), parque=str(parque_a).strip(),
                               quem=str(quem).strip(), status=status_str, obs=str(obs).strip()))
        if admin_row is not None and admin_row < r < admin_end:
            admin_section_rows += 1
        if resolvidas_row is not None and resolvidas_row < r < resolvidas_end:
            resolvidas_section_rows += 1

    acao_rows = classificar_acoes_por_categoria(acao_rows)

    # totais oficiais de Emergencial/Programado: procurados como rotulo + numero
    # na mesma linha do cabecalho do plano (ex.: coluna AA="EMERGENCIAL", AB=6).
    n_emerg_sheet = n_prog_sheet = None
    if plano_row is not None:
        for r in range(plano_row, min(plano_row + 3, pending_end)):
            for c in range(20, 31):
                v = get(r, c)
                if not isinstance(v, str):
                    continue
                vu = v.strip().upper()
                if vu == "EMERGENCIAL" and n_emerg_sheet is None:
                    for cc in range(c + 1, c + 4):
                        nv = get(r, cc)
                        if isinstance(nv, (int, float)):
                            n_emerg_sheet = int(nv); break
                elif vu == "PROGRAMADO" and n_prog_sheet is None:
                    for cc in range(c + 1, c + 4):
                        nv = get(r, cc)
                        if isinstance(nv, (int, float)):
                            n_prog_sheet = int(nv); break

    if n_emerg_sheet is not None and n_prog_sheet is not None:
        n_emerg, n_prog = n_emerg_sheet, n_prog_sheet
    else:
        # fallback: heuristica antiga por status, restrita a secao "pendente"
        log("  [aviso] totais oficiais de Emergencial/Programado nao encontrados na planilha -- usando heuristica por status.")
        n_emerg = sum(1 for a in acao_rows if a["status"] == "!!!" or "URGENTE" in a["status"].upper())
        n_prog = sum(1 for a in acao_rows if a["status"] not in ("!!!",) and a["status"] != "" and not re.search(r'conclu', a["status"], re.I))

    n_admin = admin_section_rows
    n_resolvidas = resolvidas_section_rows
    n_ok_acao = sum(1 for a in acao_rows if re.search(r'conclu', a["status"], re.I))

    saldo_km_rows = build_saldo_km_data(wb)

    return dict(
        week_label=f"WK {week_numbers[-1]:02d}", upd_date=fmt_date_br(upd_date) if upd_date else "",
        n_total_veic=n_total_veic, n_localiza=n_localiza, n_unidas=n_unidas,
        conformidade_geral=conformidade_geral, n_acoes=len(acao_rows), n_emerg=n_emerg, n_prog=n_prog, n_admin=n_admin, n_resolvidas=n_resolvidas,
        media_km=media_km, veh_rows=veh_rows, acao_rows=acao_rows,
        hist=hist, uncertain_weeks=sorted(uncertain_weeks),
        saldo_km_rows=saldo_km_rows,
    )


def build_saldo_km_data(wb):
    """Le o grafico 'GERENCIAMENTO DE SALDO KM' da aba GRÁFICOS (chart4 do
    arquivo original, series GRÁFICOS!$C$6:$C$20 = parque/cliente e
    GRÁFICOS!$D$6:$D$20 = saldo de km). Descarta linhas com #N/A ou sem
    parque preenchido (veiculos sem parque ativo vinculado) -- pedido
    explicito do Robson em 22/08/2026: mostrar so os veiculos/parques
    realmente ativos no grafico do app/site, nao os "#N/D" da planilha."""
    if 'GRÁFICOS' not in wb.sheetnames:
        log("  [aviso] aba GRÁFICOS nao encontrada -- grafico de Saldo de KM nao sera gerado.")
        return []
    ws = wb['GRÁFICOS']
    rows = []
    for r in range(6, 21):
        parque = ws.cell(row=r, column=3).value
        saldo = ws.cell(row=r, column=4).value
        if not isinstance(parque, str) or not parque.strip():
            continue
        if isinstance(saldo, str):
            continue  # "#N/A"/"#N/D" chegam como string; valor numerico real vem como int/float
        if not isinstance(saldo, (int, float)):
            continue
        rows.append((parque.strip(), float(saldo)))
    return rows


def _catmull_rom_path(points):
    """Gera o atributo 'd' de um path suavizado (curvas C) passando pelos
    pontos dados, usando o mesmo criterio (Catmull-Rom -> Bezier, extremos
    presos) identificado no grafico desenhado a mao que este substitui."""
    n = len(points)
    if n == 1:
        x, y = points[0]
        return f"M{x:.1f},{y:.1f}"
    d = f"M{points[0][0]:.1f},{points[0][1]:.1f}"
    for i in range(n - 1):
        p_prev = points[i - 1] if i > 0 else points[i]
        p0, p1 = points[i], points[i + 1]
        p_next = points[i + 2] if i + 2 < n else points[i + 1]
        cp1 = (p0[0] + (p1[0] - p_prev[0]) / 6, p0[1] + (p1[1] - p_prev[1]) / 6)
        cp2 = (p1[0] - (p_next[0] - p0[0]) / 6, p1[1] - (p_next[1] - p0[1]) / 6)
        d += f" C{cp1[0]:.1f},{cp1[1]:.1f} {cp2[0]:.1f},{cp2[1]:.1f} {p1[0]:.1f},{p1[1]:.1f}"
    return d


def render_trend_svg(weeks_vals):
    """weeks_vals: lista de (label_wk, valor%) das ultimas semanas (com valor
    nao-None), na ordem cronologica. Reproduz o grafico 'Tendencia da Media
    Geral' (viewBox 640x220) que antes era editado a mao toda semana."""
    n = len(weeks_vals)
    x0, x1 = 70, 600
    xs = [x0 + i * (x1 - x0) / max(1, n - 1) for i in range(n)]

    def y(v):
        # 100% -> y=20, cada 5% -> 35px (mesma escala do grid fixo 80-100%)
        yy = 20 + (100 - v) * 7
        return max(20, min(160, yy))

    def color(v):
        if v >= 90:
            return "#12A874"
        if v >= 70:
            return "#D89A2B"
        return "#E0483C"

    pts = [(xs[i], y(weeks_vals[i][1])) for i in range(n)]
    line_d = _catmull_rom_path(pts)
    area_d = line_d + f" L{pts[-1][0]:.1f},160 L{pts[0][0]:.1f},160 Z"

    points_svg = ""
    for i, (wk_lbl, v) in enumerate(weeks_vals):
        px, py = pts[i]
        ty = py - 12
        points_svg += (f'      <circle cx="{px:.1f}" cy="{py:.1f}" r="5.5" fill="#0F2A40" '
                       f'stroke="{color(v)}" stroke-width="2.5" filter="url(#ptShadow)"></circle>\n'
                       f'      <text x="{px:.1f}" y="{ty:.1f}" fill="#EAF1F7" font-size="13.5" '
                       f'font-weight="700" text-anchor="middle">{rhu(v)}%</text>\n\n')
    xlabels_svg = "".join(
        f'      <text x="{xs[i]:.1f}" y="182">{weeks_vals[i][0]}</text>\n' for i in range(n)
    )

    return f'''<svg viewBox="0 0 640 220" class="trend-chart">
    <defs>
      <lineargradient id="areaGrad" x1="0" y1="0" x2="0" y2="1">
        <stop offset="0%" stop-color="#00679A" stop-opacity="0.38"></stop>
        <stop offset="100%" stop-color="#00679A" stop-opacity="0"></stop>
      </lineargradient>
      <filter id="ptShadow" x="-50%" y="-50%" width="200%" height="200%">
        <fedropshadow dx="0" dy="1" stdDeviation="1.5" flood-color="#000" flood-opacity="0.35"></fedropshadow>
      </filter>
    </defs>

    <!-- gridlines -->
    <g stroke="#1E4363" stroke-width="1">
      <line x1="70" y1="20" x2="600" y2="20"></line>
      <line x1="70" y1="55" x2="600" y2="55"></line>
      <line x1="70" y1="90" x2="600" y2="90"></line>
      <line x1="70" y1="125" x2="600" y2="125"></line>
      <line x1="70" y1="160" x2="600" y2="160"></line>
    </g>
    <g fill="#8FA9C0" font-size="11" text-anchor="end" font-family="Segoe UI,Arial,sans-serif">
      <text x="60" y="24">100%</text>
      <text x="60" y="59">95%</text>
      <text x="60" y="94">90%</text>
      <text x="60" y="129">85%</text>
      <text x="60" y="164">80%</text>
    </g>

    <!-- axes -->
    <line x1="70" y1="20" x2="70" y2="160" stroke="#3A567A" stroke-width="1.2"></line>
    <line x1="70" y1="160" x2="600" y2="160" stroke="#3A567A" stroke-width="1.2"></line>

    <!-- meta line at 90% -->
    <line x1="70" y1="90" x2="600" y2="90" stroke="#D89A2B" stroke-width="1.6" stroke-dasharray="7,5"></line>
    <text x="600" y="84" fill="#D89A2B" font-size="11.5" font-weight="700" text-anchor="end" font-family="Segoe UI,Arial,sans-serif">Meta 90%</text>

    <!-- area fill under the smoothed trend curve -->
    <path d="{area_d}" fill="url(#areaGrad)"></path>

    <!-- smoothed trend line -->
    <path d="{line_d}" fill="none" stroke="#00679A" stroke-width="3" stroke-linecap="round"></path>

    <!-- points -->
    <g font-family="Segoe UI,Arial,sans-serif">
{points_svg.rstrip()}
    </g>

    <!-- x labels -->
    <g fill="#8FA9C0" font-size="12" text-anchor="middle" font-family="Segoe UI,Arial,sans-serif">
{xlabels_svg.rstrip()}
    </g>
  </svg>'''


def render_saldo_km_svg(rows):
    """Grafico de barras verticais (mobile-friendly) do Saldo de KM por
    parque/cliente, lido de build_saldo_km_data(). Regra de cor definida
    pelo Robson em 22/08/2026: verde (--ok) quando o saldo estiver entre 0 e
    +4.000 km (inclusive); vermelho (--nok) quando negativo OU acima de
    +4.000 km positivos."""
    if not rows:
        return '<div class="note">Sem dados de Saldo de KM por parque nesta atualização (todos os veículos da aba GRÁFICOS estavam sem parque ativo vinculado).</div>'

    GREEN, RED = "#12A874", "#E0483C"

    def color(v):
        return GREEN if 0 <= v <= 4000 else RED

    n = len(rows)
    # layout: barras verticais, categorias no eixo X (rotativas p/ caber no
    # mobile), largura total cresce com o numero de parques mas o viewBox
    # e sempre normalizado 0-W para caber tanto no desktop quanto no celular
    # (a tag <svg> em si e responsiva via CSS, igual aos outros graficos).
    bar_w = 46
    gap = 22
    padL, padR = 60, 20
    padT, padB = 26, 78
    W = padL + padR + n * bar_w + (n - 1) * gap
    H = 320

    vals = [v for _, v in rows]
    v_max = max(4000, max(vals) if vals else 0)
    v_min = min(0, min(vals) if vals else 0)
    span = max(1.0, v_max - v_min)
    chart_h = H - padT - padB
    zero_y = padT + chart_h * (v_max - 0) / span

    def y_of(v):
        return padT + chart_h * (v_max - v) / span

    bars_svg = []
    labels_svg = []
    vallabels_svg = []
    for i, (parque, v) in enumerate(rows):
        x = padL + i * (bar_w + gap)
        y_top = y_of(max(v, 0))
        y_bot = y_of(min(v, 0))
        h = max(2, y_bot - y_top)
        cx = x + bar_w / 2
        label_y = zero_y + 14 if v < 0 else y_top - 8
        val_txt = f'{v:,.0f}'.replace(",", ".")
        bars_svg.append(
            f'      <rect x="{x:.1f}" y="{y_top:.1f}" width="{bar_w}" height="{h:.1f}" '
            f'rx="4" fill="{color(v)}"></rect>'
        )
        vallabels_svg.append(
            f'      <text x="{cx:.1f}" y="{label_y:.1f}" fill="#EAF1F7" font-size="11.5" '
            f'font-weight="700" text-anchor="middle">{val_txt}</text>'
        )
        labels_svg.append(
            f'      <text x="{cx:.1f}" y="{zero_y + 34:.1f}" fill="#8FA9C0" font-size="10.5" '
            f'text-anchor="end" transform="rotate(-38 {cx:.1f} {zero_y + 34:.1f})">{parque}</text>'
        )

    return f'''<svg viewBox="0 0 {W} {H}" class="trend-chart saldo-km-chart" style="height:auto; max-height:340px;">
    <!-- linha zero -->
    <line x1="{padL - 10}" y1="{zero_y:.1f}" x2="{W - padR + 10}" y2="{zero_y:.1f}" stroke="#3A567A" stroke-width="1.2"></line>
    <text x="{padL - 16}" y="{zero_y + 4:.1f}" fill="#8FA9C0" font-size="10.5" text-anchor="end" font-family="Segoe UI,Arial,sans-serif">0</text>

    <!-- linha de referencia +4.000 km (limite da faixa verde) -->
    <line x1="{padL - 10}" y1="{y_of(4000):.1f}" x2="{W - padR + 10}" y2="{y_of(4000):.1f}" stroke="#D89A2B" stroke-width="1.4" stroke-dasharray="6,5"></line>
    <text x="{W - padR + 10}" y="{y_of(4000) - 6:.1f}" fill="#D89A2B" font-size="10.5" font-weight="700" text-anchor="end" font-family="Segoe UI,Arial,sans-serif">+4.000 km</text>

    <g font-family="Segoe UI,Arial,sans-serif">
{chr(10).join(bars_svg)}
{chr(10).join(vallabels_svg)}
{chr(10).join(labels_svg)}
    </g>
  </svg>'''


def render_frota(html, d):
    def sub_once(pattern, replacement, text, flags=re.S):
        new_text, n = re.subn(pattern, replacement, text, count=1, flags=flags)
        if n == 0:
            log(f"  [aviso] padrao nao encontrado (frota): {pattern[:60]}...")
        return new_text

    def sub_nth(pattern, replacements, text, flags=re.S):
        """Substitui cada ocorrencia do padrao (em ordem de aparicao no
        arquivo) pelo texto correspondente na lista 'replacements'. Mais
        robusto que sub_once quando ha 2+ trechos parecidos no arquivo e o
        texto atual de cada um pode variar (edicao manual previa, por ex.)."""
        matches = list(re.finditer(pattern, text, flags))
        if len(matches) != len(replacements):
            log(f"  [aviso] esperava {len(replacements)} ocorrencia(s) do padrao, encontrei {len(matches)}: {pattern[:60]}...")
        out, last_end = [], 0
        for i, m in enumerate(matches):
            out.append(text[last_end:m.start()])
            out.append(replacements[i] if i < len(replacements) else m.group(0))
            last_end = m.end()
        out.append(text[last_end:])
        return "".join(out)

    html = sub_once(r'Semana \d+ \(WK \d+\)(?: · Atualizado \d{2}/\d{2}/\d{4})?',
                     f'{d["week_label"].replace("WK","Semana").replace("  "," ")} (WK {d["week_label"].split()[1]})', html)

    # ---- titulos e badges dos paineis "Detalhamento" e "Plano de Acao" --
    # antes ficavam com o rotulo de semana e as contagens Emergencial/
    # Programado/Administrativa "congelados" do ultimo texto editado a mao
    # (ex: ainda "Semana 30" e "6/10/4" varias semanas depois da planilha
    # ja estar em WK32) porque nao havia sub_once para esses trechos --
    # apenas o badge-week do topo e o card "Conformidade Geral" eram
    # recalculados. Corrigido para reescrever os 3 pontos a cada execucao.
    week_num_str = d["week_label"].split()[1]
    html = sub_once(r'Veículos — Detalhamento \(Semana \d+\)',
                     f'Veículos — Detalhamento (Semana {week_num_str})', html)
    html = sub_once(r'Plano de Ação — Semana \d+',
                     f'Plano de Ação — Semana {week_num_str}', html)
    html = sub_once(
        r'<div class="loc-badge" style="background:rgba\(224,72,60,\.18\); color:var\(--nok\)"><span class="n">\d+</span><span class="l">Emergencial</span></div>\s*'
        r'<div class="loc-badge" style="background:rgba\(216,154,43,\.24\); color:var\(--warn\)"><span class="n">\d+</span><span class="l">Programado</span></div>\s*'
        r'<div class="loc-badge" style="background:rgba\(92,122,147,\.24\); color:var\(--na\)"><span class="n">\d+</span><span class="l">Administrativa</span></div>',
        f'<div class="loc-badge" style="background:rgba(224,72,60,.18); color:var(--nok)"><span class="n">{d["n_emerg"]}</span><span class="l">Emergencial</span></div>\n'
        f'      <div class="loc-badge" style="background:rgba(216,154,43,.24); color:var(--warn)"><span class="n">{d["n_prog"]}</span><span class="l">Programado</span></div>\n'
        f'      <div class="loc-badge" style="background:rgba(92,122,147,.24); color:var(--na)"><span class="n">{d["n_admin"]}</span><span class="l">Administrativa</span></div>',
        html)

    # ---- nota de rodape final (bloco/data de atualizacao) -- tambem ficava
    # congelada ("bloco da Semana 30", "ATUALIZACAO ainda mostra 14/07/2026")
    # porque nao havia sub_once para ela; agora reflete o bloco/data reais.
    upd_date_str = d["upd_date"] or "(nao encontrada no bloco)"
    html = sub_once(
        r'Fonte: planilha "Controle de Frota" — aba BASE (?:DESHBOARD|DASHBOARD), bloco da Semana \d+\. '
        r'"Meses" calculado a partir da Data Início até a data de atualização '
        r'\(a célula "ATUALIZAÇÃO" da planilha ainda mostra \d{2}/\d{2}/\d{4} — '
        r'se você já rodou o checklist desta semana, vale atualizar essa célula na planilha\)\.',
        f'Fonte: planilha "Controle de Frota" — aba BASE DASHBOARD, bloco da Semana {week_num_str}. '
        f'"Meses" calculado a partir da Data Início até a data de atualização '
        f'(célula "ATUALIZAÇÃO" do bloco atual: {upd_date_str}).',
        html)

    kpi1 = f'''<div class="kpi">
    <div class="label">Total de Veículos</div>
    <div class="value">{d["n_total_veic"]}</div>
    <div class="locadoras">
      <div class="loc-badge loc-localiza"><span class="n">{d["n_localiza"]}</span><span class="l">Localiza</span></div>
      <div class="loc-badge loc-unidas"><span class="n">{d["n_unidas"]}</span><span class="l">Unidas</span></div>
    </div>
  </div>'''
    html = sub_once(r'<div class="kpi">\n    <div class="label">Total de Veículos</div>.*?</div>\n  </div>', kpi1, html)

    conf_txt = f'{rhu(d["conformidade_geral"])}%' if d["conformidade_geral"] is not None else "-"
    if d["conformidade_geral"] is None:
        conf_color = "var(--muted)"
    elif d["conformidade_geral"] >= 90:
        conf_color = "var(--ok)"
    elif d["conformidade_geral"] >= 70:
        conf_color = "var(--warn)"
    else:
        conf_color = "var(--nok)"
    n_pend_total = d["n_prog"] + d["n_emerg"] + d["n_admin"]
    # Layout padrao: 3 badges (Emergenciais/Programadas/Administrativas), sem
    # o badge "Resol" -- confirmado como o novo padrao do dashboard.
    kpi2 = f'''<div class="kpi">
    <div class="label">Conformidade Geral &amp; Ações Pendentes</div>
    <div class="value" style="color:{conf_color}">{conf_txt}</div>
    <div class="foot">{d["week_label"]} · {n_pend_total} ações pendentes no total</div>
    <div class="locadoras" style="display:grid; grid-template-columns:1fr 1fr 1fr; gap:8px;">
      <div class="loc-badge" style="background:rgba(224,72,60,.18); color:var(--nok)"><span class="n">{d["n_emerg"]}</span><span class="l">Emergenciais</span></div>
      <div class="loc-badge" style="background:rgba(216,154,43,.24); color:var(--warn)"><span class="n">{d["n_prog"]}</span><span class="l">Programadas</span></div>
      <div class="loc-badge" style="background:rgba(92,122,147,.24); color:var(--na)"><span class="n">{d["n_admin"]}</span><span class="l">Administrativas</span></div>
    </div>
  </div>'''
    html = sub_once(r'<div class="kpi">\n    <div class="label">Conformidade Geral.*?</div>\n  </div>', kpi2, html)

    if d["media_km"] is not None:
        km_color = "var(--nok)" if d["media_km"] >= 30000 else "var(--ok)"
        km_fmt = f'{d["media_km"]:,}'.replace(',', '.')
        html = sub_once(r'data-km="\d+"', f'data-km="{d["media_km"]}"', html)
        html = sub_once(r'(<div class="value" id="kmValue" data-km="\d+" style="color: )[^;]+(;">)[\d.]+(</div>)',
                         lambda m: m.group(1) + km_color + m.group(2) + km_fmt + m.group(3), html)

    html = sub_once(r'\d+ de \d+ veículos exibidos', f'{d["n_total_veic"]} de {d["n_total_veic"]} veículos exibidos', html)

    def esc(s):
        return str(s).replace('"', '&quot;')

    rows_js = "[\n" + ",\n".join(
        "[" + ",".join(js_str(str(x)) for x in row) + "]" for row in d["veh_rows"]
    ) + ",\n]"
    html = sub_once(r'const rows = \[.*?\n\];', "const rows = " + rows_js + ";", html)

    acao_js_items = []
    for a in d["acao_rows"]:
        acao_js_items.append(
            "  { acao:%s, tipo:%s, parque:%s, quem:%s, status:%s, obs:%s, categoria:%s }" % (
                js_str(a["acao"]), js_str(a["tipo"]), js_str(a["parque"]), js_str(a["quem"]),
                js_str(a["status"]), js_str(a["obs"]), js_str(a.get("categoria", "Administrativa")))
        )
    acao_js = "[\n" + ",\n".join(acao_js_items) + "\n]"
    html = sub_once(r'const acaoData = \[.*?\n\];', "const acaoData = " + acao_js + ";", html)

    hist_items = []
    for h in d["hist"]:
        cats_str = ",".join("null" if v is None else str(v) for v in h["cats"])
        geral_str = "null" if h["geral"] is None else str(h["geral"])
        hist_items.append(f'  {{ wk:"WK{h["wk"]:02d}", geral:{geral_str}, cats:[{cats_str}] }}')
    hist_js = "[\n" + ",\n".join(hist_items) + ",\n]"
    html = sub_once(r'const histData = \[.*?\n\];', "const histData = " + hist_js + ";", html)

    # nota de rodape do historico -- reescrita por completo a cada execucao
    # (nunca "acrescentada" ao texto anterior) para garantir idempotencia:
    # rodar o script varias vezes seguidas produz sempre o mesmo resultado.
    uncertain = d["uncertain_weeks"]
    base_note = (
        f'Fonte: aba BASE DASHBOARD, {len(d["hist"])} blocos semanais identificados '
        f'(WK{d["hist"][0]["wk"]:02d}–WK{d["hist"][-1]["wk"]:02d}). '
        '"Média Geral" calculada como a média de Checklist, Revisão, Estepe Extra, Rastreador+Bloqueador '
        'e Medição Pneus por bloco semanal, direto dos dados brutos (nao depende de tabela dinamica/pivot).'
    )
    risk_bits = []
    if uncertain:
        wk_str = ", ".join(f"WK{w:02d}" for w in sorted(uncertain))
        risk_bits.append(
            "Semanas com rótulo original inconsistente na planilha, número inferido pela posição "
            "sequencial (confirme antes de usar externamente): " + wk_str + "."
        )
    low_n = sorted(h["wk"] for h in d["hist"] if h["nveh"] < 10 and h["wk"] not in uncertain)
    if low_n:
        risk_bits.append(
            "Semanas com poucos veículos registrados no bloco (dado possivelmente incompleto): "
            + ", ".join(f"WK{w:02d}" for w in low_n) + "."
        )
    full_note = base_note + ((" " + " ".join(risk_bits)) if risk_bits else "")

    # ---- grafico "Tendencia da Media Geral" (topo) -- ultimas semanas ----
    # Antes era editado a mao toda semana; agora e gerado a partir do
    # historico calculado acima (RESUMO oficial quando existe, senao bruto).
    WINDOW_N = 5
    hist_window = [h for h in d["hist"][-WINDOW_N:] if h["geral"] is not None]
    if len(hist_window) >= 2:
        weeks_vals = [(f'WK {h["wk"]:02d}', h["geral"]) for h in hist_window]
        trend_svg = render_trend_svg(weeks_vals)
        html = sub_once(r'<svg viewBox="0 0 640 220" class="trend-chart">\s*<defs>.*?</svg>', trend_svg, html)
    else:
        log("  [aviso] dados insuficientes para regenerar o grafico de tendencia (topo).")

    # ---- grafico "Gerenciamento de Saldo de KM por Parque" ----
    saldo_km_svg = render_saldo_km_svg(d.get("saldo_km_rows") or [])
    html = sub_once(r'<svg viewBox="0 0 [\d ]+" class="trend-chart(?: saldo-km-chart)?"[^>]*>\s*<!-- linha zero -->.*?</svg>', saldo_km_svg, html)
    log(f"  Saldo de KM: {len(d.get('saldo_km_rows') or [])} parques ativos no grafico.")

    # ---- tabela "Conformidade por Categoria" -- mesmas semanas do grafico ----
    conf_window = d["hist"][-WINDOW_N:]
    if conf_window:
        wk_lo, wk_hi = conf_window[0]["wk"], conf_window[-1]["wk"]
        # No mobile so as 2 ultimas semanas (vigente + anterior) ficam visiveis
        # na tabela (as demais sao escondidas via CSS), entao o titulo precisa
        # de uma versao curta so com essas 2 pro mobile -- o desktop continua
        # mostrando a janela cheia de WINDOW_N semanas.
        last2 = conf_window[-2:] if len(conf_window) >= 2 else conf_window
        wk_lo2, wk_hi2 = last2[0]["wk"], last2[-1]["wk"]
        html = sub_once(
            r'<h2>Conformidade por Categoria — Tendência Semanal '
            r'<span class="wk-full">\(WK \d+–\d+\)</span>'
            r'<span class="wk-mobile">\(WK \d+–\d+\)</span></h2>'
            r'|<h2>Conformidade por Categoria — Tendência Semanal \(WK \d+–\d+\)</h2>',
            '<h2>Conformidade por Categoria — Tendência Semanal '
            f'<span class="wk-full">(WK {wk_lo:02d}–{wk_hi:02d})</span>'
            f'<span class="wk-mobile">(WK {wk_lo2:02d}–{wk_hi2:02d})</span></h2>', html)

        def cellclass(v):
            if v is None:
                return "nd"
            if v >= 90:
                return "cell-ok"
            if v >= 70:
                return "cell-warn"
            return "cell-nok"

        def cellval(v):
            return "–" if v is None else f"{rhu(v)}%"

        n_weeks = len(conf_window)
        thead_cells = "".join(f"<th>WK {h['wk']:02d}</th>" for h in conf_window)
        thead = f'<tr>\n        <th>Categoria</th>{thead_cells}\n      </tr>'
        html = sub_once(r'<tr>\s*<th>Categoria</th>.*?</tr>', thead, html)

        body_rows = []
        for cat_idx, label in enumerate(CATS_ALL_LABELS):
            cells = ""
            for i, h in enumerate(conf_window):
                v = h["cats"][cat_idx]
                cur = " cur" if i == n_weeks - 1 else ""
                cells += f'<td class="{cellclass(v)}{cur}">{cellval(v)}</td>'
            body_rows.append(f'      <tr><td>{label}</td>{cells}</tr>')
        tbody = "\n".join(body_rows)
        html = sub_once(r'<tbody>\s*<tr><td>KM Ponto de Troca</td>.*?</tbody>',
                         "<tbody>\n" + tbody + "\n    </tbody>", html)

        geral_cells = ""
        for i, h in enumerate(conf_window):
            v = h["geral"]
            cur = " cur" if i == n_weeks - 1 else ""
            geral_cells += f'<td class="{cellclass(v)}{cur}">{cellval(v)}</td>'
        html = sub_once(r'<tr><td>Média Geral</td>.*?</tr>', f'<tr><td>Média Geral</td>{geral_cells}</tr>', html)

    # ---- titulo do historico completo (JS-renderizado a partir de histData) ----
    html = sub_once(
        r'<h2 style="margin-bottom:0">Histórico Semanal Completo — WK \d+ a WK \d+</h2>',
        f'<h2 style="margin-bottom:0">Histórico Semanal Completo — WK {d["hist"][0]["wk"]:02d} a WK {d["hist"][-1]["wk"]:02d}</h2>',
        html)

    # ---- nota de rodape do historico completo (segunda ocorrencia do padrao
    # "Fonte: aba BASE DESHBOARD" no arquivo -- distinta da nota tratada acima,
    # por isso usa um prefixo proprio para ser encontrada de forma unica e
    # continuar idempotente em execucoes futuras) ----
    oficial_weeks = sorted(h["wk"] for h in d["hist"] if h.get("fonte") == "oficial")
    bruto_weeks = sorted(h["wk"] for h in d["hist"] if h.get("fonte") == "bruto")
    cat_note = (
        'Fonte: aba BASE DASHBOARD, tabela "RESUMO" (cálculo oficial por fórmula da própria planilha) '
        'para Checklist, Revisão, Estepe Extra, Medição Pneus, Rastreador+Bloqueador e Média Geral'
        + (f', disponível nas semanas WK{oficial_weeks[0]:02d}–WK{oficial_weeks[-1]:02d}' if oficial_weeks else '')
        + '. KM Ponto de Troca, Saldo Extra, Consumo, Doc Postado e Doc Entregue não têm tabela oficial na '
          'planilha e são sempre calculados por contagem bruta OK/N-OK de cada veículo no bloco semanal'
        + (f'; esse mesmo cálculo bruto também cobre Checklist/Revisão/Estepe/Pneus/Rastreador nas semanas '
           f'sem tabela RESUMO (WK{bruto_weeks[0]:02d}–WK{bruto_weeks[-1]:02d}).' if bruto_weeks else '.')
    )
    if uncertain:
        wk_str = ", ".join(f"WK{w:02d}" for w in sorted(uncertain))
        cat_note += (' Semanas com rótulo original inconsistente na planilha, número inferido pela posição '
                     'sequencial (confirme antes de usar externamente): ' + wk_str + '.')

    # As duas notas de rodape ("Fonte: aba BASE DESHBOARD...") sao substituidas
    # juntas, por ordem de aparicao no arquivo (1a = painel de Plano de Acao,
    # 2a = painel Historico Semanal Completo) -- assim funciona independente
    # do texto atual de cada uma (inclusive apos edicao manual previa).
    html = sub_nth(
        r'<div class="note">\s*Fonte: aba BASE (?:DESHBOARD|DASHBOARD).*?</div>',
        [f'<div class="note">\n      {full_note}\n    </div>',
         f'<div class="note">\n      {cat_note}\n    </div>'],
        html)

    return html


def render_index(html, d_eq=None, d_fr=None, d_bo=None, d_en=None, d_re=None):
    def sub_once(pattern, replacement, text, flags=re.S):
        new_text, n = re.subn(pattern, replacement, text, count=1, flags=flags)
        if n == 0:
            log(f"  [aviso] padrao nao encontrado (index): {pattern[:60]}...")
        return new_text

    hoje = datetime.date.today().strftime('%d/%m/%Y')
    # menuDemandasLabel agora e texto fixo "NOSSAS DEMANDAS (HOJE)" -- nao
    # precisa mais de substituicao de data a cada execucao.
    html = sub_once(r'<div class="badge" id="topBadge">Atualizado \d{2}/\d{2}/\d{4}</div>',
                     f'<div class="badge" id="topBadge">Atualizado {hoje}</div>', html)

    if d_eq is not None:
        # Corrigido em 22/08/2026: o card real do card flutuante "Equipes"
        # (dentro do bloco do mapa) usa os rotulos "Ativos"/"Desmob."/"Parques"
        # com estilos inline especificos (font-size reduzido pra caber no
        # card pequeno de 115px), nao "Ativos hoje"/"Em Desmob." como o
        # regex antigo esperava -- por isso esse card NUNCA era atualizado
        # (o [aviso] "padrao nao encontrado (index)" nos logs era exatamente
        # este ponto). Ver tambem REGRAS_AUTOMACAO.md.
        n_em_desmob = len(d_eq.get("em_desmobilizacao", []))
        card_eq = (
            '<div class="kpirow" style="gap:4px 6px;">\n'
            f'              <div><div class="n" style="font-size:12px;">{d_eq["tot_real"]}</div><div class="l" style="font-size:7px;">Ativos</div></div>\n'
            f'              <div><div class="n" style="color:var(--warn); font-size:12px;">{n_em_desmob}</div><div class="l" style="font-size:7px;">Desmob.</div></div>\n'
            f'              <div><div class="n" style="font-size:12px;">{d_eq["parques_ativos"]}/{d_eq["parques_total"]}</div><div class="l" style="font-size:7px;">Parques</div></div>\n'
            '            </div>'
        )
        html = sub_once(
            r'<div class="kpirow" style="gap:4px 6px;">\s*<div><div class="n" style="font-size:12px;">\d+</div><div class="l" style="font-size:7px;">Ativos</div></div>.*?<div class="l" style="font-size:7px;">Parques</div></div>\s*</div>',
            card_eq, html)
        # Card "Ativos hoje" / "Em Desmob." antigo (layout anterior, sem
        # estilos inline reduzidos) continua reconhecido como fallback, caso
        # algum HTML anterior a 22/08/2026 ainda esteja com esse formato.
        card_eq_legado = f'''<div class="kpirow">
              <div><div class="n">{d_eq["tot_real"]}</div><div class="l">Ativos hoje</div></div>
              <div><div class="n" style="color:var(--warn)">{n_em_desmob}</div><div class="l">Em Desmob.</div></div>
              <div><div class="n">{d_eq["parques_ativos"]}/{d_eq["parques_total"]}</div><div class="l">Parques</div></div>
            </div>'''
        html, _n_legado = re.subn(
            r'<div class="kpirow">\s*<div><div class="n">\d+</div><div class="l">Ativos hoje</div></div>.*?<div class="l">Parques</div></div>\s*</div>',
            card_eq_legado, html, count=1, flags=re.S)

    if d_fr is not None:
        conf_txt = f'{int(d_fr["conformidade_geral"])}%' if d_fr["conformidade_geral"] is not None else "-"
        n_pend_total = d_fr["n_prog"] + d_fr["n_emerg"] + d_fr["n_admin"]
        card_fr = f'''<div class="kpirow">
              <div><div class="n">{d_fr["n_total_veic"]}</div><div class="l">Veículos</div></div>
              <div><div class="n" style="color:var(--ok)">{conf_txt}</div><div class="l">Conformidade</div></div>
              <div><div class="n" style="color:var(--nok)">{n_pend_total}</div><div class="l">Ações pendentes</div></div>
            </div>'''
        html = sub_once(
            r'<div class="kpirow">\s*<div><div class="n">\d+</div><div class="l">Veículos</div></div>.*?<div class="l">Ações pendentes</div></div>\s*</div>',
            card_fr, html)

    if d_bo is not None:
        vb_venc, vc_venc = d_bo["indicadores"].get("vencer", (0, 0))
        vb_mult, vc_mult = d_bo["indicadores"].get("multas", (0, 0))
        vb_cont, vc_cont = d_bo["indicadores"].get("contestacao", (0, 0))
        n = lambda v: int(v) if isinstance(v, (int, float)) else 0
        card_bo = f'''<div class="kpirow">
              <div><div class="n">{n(vb_venc)}<span style="font-size:14px; color:var(--muted);"> / {n(vc_venc)}</span></div><div class="l">À vencer (U/L)</div></div>
              <div><div class="n" style="color:var(--nok)">{n(vb_mult)}<span style="font-size:14px; color:var(--muted);"> / {n(vc_mult)}</span></div><div class="l">Multas (U/L)</div></div>
              <div><div class="n" style="color:var(--warn)">{n(vb_cont)}<span style="font-size:14px; color:var(--muted);"> / {n(vc_cont)}</span></div><div class="l">Contestação (U/L)</div></div>
            </div>'''
        html = sub_once(
            r'<div class="kpirow">\s*<div><div class="n">\d+<span style="font-size:14px; color:var\(--muted\);"> / \d+</span></div><div class="l">À vencer \(U/L\)</div></div>.*?<div class="l">Contestação \(U/L\)</div></div>\s*</div>',
            card_bo, html)

    if d_en is not None:
        def n(v):
            return int(v) if isinstance(v, (int, float)) else 0
        aberto_total = d_en["aberto_total"] if d_en["aberto_total"] is not None else sum(q for (_, q) in d_en["aberto_rows"]) or 0
        card_en = f'''<div class="kpirow">
              <div><div class="n">{n(d_en["kpi_almox"])}</div><div class="l">Requisições 2026</div></div>
              <div><div class="n" style="color:var(--ok)">{n(d_en["kpi_entregues"])}</div><div class="l">Entregues (Log.)</div></div>
              <div><div class="n" style="color:var(--nok)">{n(aberto_total)}</div><div class="l">Em aberto</div></div>
            </div>'''
        html = sub_once(
            r'<div class="kpirow">\s*<div><div class="n">\d+</div><div class="l">Requisições 2026</div></div>.*?<div class="l">Em aberto</div></div>\s*</div>',
            card_en, html)

    if d_re is not None:
        mes_atual = d_re.get("mes_atual") or {}
        mes_val = mes_atual.get("valor")
        semana_val = d_re.get("semana_atual_valor")
        total_ano = d_re.get("total_ano")
        fmt_int_brl = lambda v: f"R$ {int(round(v)):,}".replace(",", ".") if isinstance(v, (int, float)) else "R$ -"
        card_re = f'''<div class="kpirow">
              <div><div class="n">{fmt_int_brl(mes_val)}</div><div class="l">Total mês atual</div></div>
              <div><div class="n">{fmt_int_brl(semana_val)}</div><div class="l">Semana atual (parcial)</div></div>
              <div><div class="n">{fmt_int_brl(total_ano)}</div><div class="l">Acumulado 2026</div></div>
            </div>'''
        html = sub_once(
            r'<div class="kpirow">\s*<div><div class="n">R\$ [\d.]+</div><div class="l">Total mês atual</div></div>.*?<div class="l">Acumulado 2026</div></div>\s*</div>',
            card_re, html)

    return html


# ============================================================
# BOLETOS
# ============================================================

# Ordem e rotulos das linhas 3-9 da aba DASHBOARD ATUALIZACAO (coluna A = rotulo,
# B = UNIDAS, C = LOCALIZA). "Proximo a Vencer" e uma data; as demais sao contagens.
BOLETOS_INDICADORES = [
    ("Boletos/Multas pendentes Download", "download", "n"),
    ("Boletos à Vencer", "vencer", "n"),
    ("Próximo à Vencer", "proximo", "d"),
    ("Multas em Aberto", "multas", "n"),
    ("Contestação", "contestacao", "n"),
    ("Contestar", "contestar", "n"),
    ("Sinistros em Aberto", "sinistros", "n"),
]


def _find_section_row(grid, max_row, must_contain):
    """Acha a linha cujo texto da coluna A contem todos os fragmentos de 'must_contain'
    (lista de substrings, comparacao case-insensitive). Retorna None se nao achar."""
    for r in range(1, max_row + 1):
        v = grid.get((r, 1))
        if isinstance(v, str) and v.strip():
            vu = v.strip().upper()
            if all(frag.upper() in vu for frag in must_contain):
                return r
    return None


def _read_detail_section(grid, title_row, max_row):
    """Le uma secao de detalhe no formato: linha de titulo, linha seguinte com
    cabecalhos (PLACA/LOCADORA/PARQUE/VENCIMENTO/MOTIVO/MOTIVO COBRANCA EXTRA),
    e entao linhas de dados ate a primeira linha com Placa vazia (ou ate o fim
    da planilha)."""
    if title_row is None:
        return []
    header_row = title_row + 1
    rows = []
    r = header_row + 1
    while r <= max_row:
        placa = grid.get((r, 1))
        if placa is None or (isinstance(placa, str) and not placa.strip()):
            break
        placa_s = str(placa).strip()
        venc_v = grid.get((r, 4))
        venc_s = fmt_date_br(venc_v) if isinstance(venc_v, datetime.datetime) else (str(venc_v) if venc_v else "")
        rows.append(dict(
            placa=placa_s,
            locadora=str(grid.get((r, 2)) or "").strip(),
            parque=str(grid.get((r, 3)) or "").strip(),
            venc=venc_s,
            motivo=str(grid.get((r, 5)) or "").strip(),
            motivo_extra=str(grid.get((r, 6)) or "").strip(),
        ))
        r += 1
    return rows


def _excel_serial(dt):
    epoch = datetime.datetime(1899, 12, 30)
    return (dt - epoch).days


def _calc_proximo_a_vencer(locadora, path):
    """Replica em Python a formula matricial '=MIN(IF(...))' da celula
    Proximo a Vencer (B6/C6 da aba DASHBOARD ATUALIZACAO), lendo direto a
    aba 'BOLETOS A PAGAR R-1' do arquivo indicado. Usado como fallback
    quando o LibreOffice nao recalcula essa formula (ver recalc_boletos_xlsx)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['BOLETOS À PAGAR R-1']
    hoje = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    datas = []
    for row in ws.iter_rows(min_row=2, max_row=2000):
        a = row[0].value
        c = row[2].value
        i = row[8].value
        j = row[9].value
        f = row[5].value
        if not a or not isinstance(c, str) or c.strip() != locadora.strip():
            continue
        try:
            i_v = float(i) if i is not None else 0
        except Exception:
            i_v = 0
        try:
            j_v = float(j) if j is not None else 0
        except Exception:
            j_v = 0
        if i_v > 0 and j_v == 0 and isinstance(f, datetime.datetime) and f >= hoje:
            datas.append(f)
    return min(datas) if datas else None


def _read_contestacao_detail(path):
    """Replica em Python o filtro do indicador 'Contestacao' (linha 8, colunas
    B/C da aba DASHBOARD ATUALIZACAO): COUNTIFS sobre a aba 'BOLETOS A PAGAR
    R-1' filtrando por Locadora, Valor Contestado (coluna J, indice 9) > 0 e
    Valor Pago (coluna H, indice 7) diferente do texto "RESOLVIDO". A aba
    resumo so guarda a CONTAGEM (formula COUNTIFS) -- nao existe secao de
    detalhe nativa na planilha para essa lista, entao lemos direto da aba
    fonte aqui (mesmo padrao usado por _calc_proximo_a_vencer) para poder
    exibir placa/motivo de cada contestacao no dashboard. Pode nao bater
    1:1 com o indicador em casos de borda (linhas que ainda nao mencionam
    "contestacao" no motivo mas ja tem valor contestado lancado) -- e
    esperado que o usuario confira visualmente no dashboard."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['BOLETOS À PAGAR R-1']
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=15, values_only=True):
        placa, parque, locadora, uso, ano, venc, mesb, valor_pago, valor_apagar, valor_contest, periodo, messerv, motivo, motivo_extra, obs = row
        if placa is None or (isinstance(placa, str) and not placa.strip()):
            continue
        try:
            j_v = float(valor_contest) if valor_contest is not None else 0
        except Exception:
            j_v = 0
        if j_v <= 0:
            continue
        resolvido = isinstance(valor_pago, str) and valor_pago.strip().upper() == "RESOLVIDO"
        if resolvido:
            continue
        venc_s = fmt_date_br(venc) if isinstance(venc, datetime.datetime) else (str(venc) if venc else "")
        rows.append(dict(
            placa=str(placa).strip(),
            locadora=str(locadora or "").strip(),
            parque=str(parque or "").strip(),
            venc=venc_s,
            motivo=str(motivo or "").strip(),
            motivo_extra=str(motivo_extra or "").strip(),
        ))
    # mais recentes primeiro (mesma ordem de exibicao das demais secoes de detalhe)
    rows.sort(key=lambda r: r["venc"], reverse=True)
    return rows


def _fix_proximo_a_vencer_xml(xlsx_path, unidas_dt, localiza_dt):
    """Edita o XML interno do xlsx (zip) para substituir o cache de erro
    (#VALUE!) das celulas B6/C6 (Proximo a Vencer) pelo valor calculado,
    SEM remover a formula matricial original -- preserva o recalculo
    automatico no Excel, so corrige o valor em cache usado pelo dashboard."""
    with zipfile.ZipFile(xlsx_path) as z:
        wb_xml = z.read('xl/workbook.xml').decode('utf-8')
        rels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')

    m = re.search(r'<sheet[^>]*name="DASHBOARD ATUALIZAÇÃO"[^>]*r:id="(rId\d+)"', wb_xml)
    if not m:
        raise ValueError("aba DASHBOARD ATUALIZAÇÃO nao encontrada no workbook.xml")
    rid = m.group(1)
    m2 = re.search(rf'<Relationship[^>]*Id="{rid}"[^>]*Target="([^"]+)"', rels_xml)
    if not m2:
        raise ValueError(f"relationship {rid} nao encontrado")
    target = m2.group(1)
    sheet_path = "xl/" + target if not target.startswith("/") else target.lstrip("/")

    with zipfile.ZipFile(xlsx_path) as z:
        sheet_xml = z.read(sheet_path).decode('utf-8')

    def fix_cell(xml, cell_ref, new_serial):
        pattern = re.compile(
            r'(<c r="' + cell_ref + r'" s="\d+")( t="e")(><f[^>]*>.*?</f>)<v>#VALUE!</v>(</c>)',
            re.S
        )
        def repl(mm):
            return mm.group(1) + ' t="n"' + mm.group(3) + f'<v>{new_serial}</v>' + mm.group(4)
        new_xml, n = pattern.subn(repl, xml)
        return new_xml, n

    n_total = 0
    if unidas_dt is not None:
        sheet_xml, n = fix_cell(sheet_xml, "B6", _excel_serial(unidas_dt))
        n_total += n
    if localiza_dt is not None:
        sheet_xml, n = fix_cell(sheet_xml, "C6", _excel_serial(localiza_dt))
        n_total += n

    if n_total == 0:
        return False

    tmp_out = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path) as zin, zipfile.ZipFile(tmp_out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_path:
                data = sheet_xml.encode('utf-8')
            zout.writestr(item, data)
    os.replace(tmp_out, xlsx_path)
    return True


def recalc_boletos_xlsx(path):
    """Recalcula as formulas da planilha de Boletos via LibreOffice headless
    e sobrescreve o arquivo original com os valores recalculados, preservando
    todas as formulas (nao apenas os valores). Trata o fallback conhecido do
    indicador 'Proximo a Vencer' (formula matricial que o LibreOffice nao
    recalcula corretamente). Faz backup do arquivo original antes de
    sobrescrever. Levanta excecao se algo alem desse fallback especifico
    vier invalido -- nesse caso o chamador deve preservar o arquivo original
    e so seguir lendo os indicadores como estavam (comportamento anterior).
    Retorna True se recalculou e sobrescreveu, False se nao foi possivel
    (ex: soffice indisponivel) -- nesse caso tambem preserva o original."""
    if shutil.which("soffice") is None:
        log("  [aviso] soffice (LibreOffice) nao encontrado nesta sessao -- pulando recalculo de formulas, lendo Boletos com o cache existente no arquivo.")
        return False

    with tempfile.TemporaryDirectory() as tmpdir:
        result = subprocess.run(
            ["soffice", "--headless", "--convert-to", "xlsx:Calc MS Excel 2007 XML", "--outdir", tmpdir, path],
            capture_output=True, text=True, timeout=120,
        )
        recalc_path = os.path.join(tmpdir, os.path.basename(path))
        if result.returncode != 0 or not os.path.isfile(recalc_path):
            log(f"  [aviso] recalculo de Boletos via LibreOffice falhou (returncode={result.returncode}) -- lendo Boletos com o cache existente no arquivo.")
            return False

        wb = openpyxl.load_workbook(recalc_path, data_only=True, read_only=True)
        ws = wb['DASHBOARD ATUALIZAÇÃO']
        erros = []
        for (r, label) in [(3, "Data do Último Boleto Atualizado"), (4, "Boletos/Multas pendentes Download"),
                            (5, "Boletos à Vencer"), (6, "Próximo à Vencer"), (7, "Multas em Aberto"),
                            (8, "Contestação"), (9, "Contestar"), (10, "Sinistros em Aberto")]:
            for col in (2, 3):
                v = ws.cell(row=r, column=col).value
                if isinstance(v, str) and v.startswith("#"):
                    erros.append((r, col, v))

        # Fallback conhecido: SO a celula Proximo a Vencer (linha 6) com erro,
        # e mais nenhuma outra -- calcula manualmente e corrige so o cache XML.
        outros_erros = [e for e in erros if e[0] != 6]
        erro_proximo = [e for e in erros if e[0] == 6]

        if outros_erros:
            log(f"  [aviso] Boletos: recalculo teve erro inesperado fora do fallback conhecido ({outros_erros}) -- mantendo arquivo original sem sobrescrever.")
            return False

        if erro_proximo:
            unidas_dt = _calc_proximo_a_vencer("UNIDAS ", path)
            localiza_dt = _calc_proximo_a_vencer("LOCALIZA", path)
            try:
                _fix_proximo_a_vencer_xml(recalc_path, unidas_dt, localiza_dt)
            except Exception as e:
                log(f"  [aviso] Boletos: falha ao aplicar fallback do indicador Proximo a Vencer ({e!r}) -- mantendo arquivo original sem sobrescrever.")
                return False

        # Validacao final: reabre a copia corrigida por completo (garante que
        # nao corrompeu o xlsx) e confirma que a contagem de formulas na aba
        # principal bate com o arquivo original (nenhuma formula perdida).
        try:
            openpyxl.load_workbook(recalc_path)
            def count_formulas(p):
                wbf = openpyxl.load_workbook(p, data_only=False, read_only=True)
                wsf = wbf['BOLETOS À PAGAR R-1']
                n = 0
                for row in wsf.iter_rows(min_row=2, max_row=50):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            n += 1
                return n
            if count_formulas(recalc_path) != count_formulas(path):
                log("  [aviso] Boletos: contagem de formulas divergente apos recalculo -- mantendo arquivo original sem sobrescrever.")
                return False
        except Exception as e:
            log(f"  [aviso] Boletos: copia recalculada invalida ({e!r}) -- mantendo arquivo original sem sobrescrever.")
            return False

        backup_path = path + f".bak_{datetime.date.today().isoformat()}"
        if not os.path.isfile(backup_path):
            shutil.copy(path, backup_path)
        shutil.copy(recalc_path, path)
        log("  Boletos: formulas recalculadas e arquivo original atualizado (fallback aplicado)" if erro_proximo else "  Boletos: formulas recalculadas e arquivo original atualizado.")
        return True


def build_boletos_data():
    wb = openpyxl.load_workbook(BOLETOS_XLSX, data_only=True, read_only=True)
    ws = wb['DASHBOARD ATUALIZAÇÃO']
    grid = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for c in row:
            if c.value is not None:
                grid[(c.row, c.column)] = c.value

    def get(r, c):
        return grid.get((r, c))

    # Busca os rotulos por texto (nao por numero de linha fixo) -- a planilha ja
    # mudou de layout mais de uma vez (linha nova inserida no topo, secoes de
    # detalhe adicionadas depois), entao e mais robusto varrer as primeiras
    # linhas procurando cada rotulo conhecido do que assumir numero de linha fixo.
    # Paramos de varrer indicadores assim que achamos qualquer titulo de secao
    # (linha com travessao "—", ex.: "PENDENCIAS EM ABERTO" ou "MULTAS EM ABERTO
    # — DETALHE") ou o antigo cabecalho "SEMANA" (formato antigo do log semanal).
    indicadores = {}
    ultimo_boleto = (None, None)
    for r in range(1, 40):
        label = get(r, 1)
        if not isinstance(label, str) or not label.strip():
            continue
        label_u = label.strip().upper()
        if label_u == "SEMANA" or "—" in label or "PENDÊNCIAS EM ABERTO" in label_u or "PENDENCIAS EM ABERTO" in label_u:
            break
        if label_u == "DATA DO ÚLTIMO BOLETO ATUALIZADO":
            ultimo_boleto = (get(r, 2), get(r, 3))
            continue
        match = next((key for (lbl, key, typ) in BOLETOS_INDICADORES if lbl.upper() == label_u), None)
        if match is not None:
            indicadores[match] = (get(r, 2), get(r, 3))

    # Tabela de log semanal manual (formato antigo, cabecalho "SEMANA"). A
    # planilha foi reestruturada e essa tabela pode nao existir mais -- nesse
    # caso log_table_found fica False e o HTML existente (com o historico
    # ja registrado) e preservado sem ser sobrescrito.
    log_header_row = None
    for r in range(1, ws.max_row + 1):
        label = get(r, 1)
        if isinstance(label, str) and label.strip().upper() == "SEMANA":
            log_header_row = r
            break

    log_rows = []
    log_table_found = log_header_row is not None
    if log_table_found:
        for r in range(log_header_row + 1, ws.max_row + 1):
            semana = get(r, 1)
            if not isinstance(semana, str) or not semana.strip():
                continue
            data_v = get(r, 2)
            data_str = fmt_date_br(data_v) if isinstance(data_v, datetime.datetime) else (str(data_v) if data_v else "")
            locadora = get(r, 3) or ""
            status = get(r, 4) or ""
            obs = get(r, 5) or ""
            log_rows.append(dict(semana=semana.strip(), data=data_str, locadora=str(locadora).strip(),
                                  status=str(status).strip(), obs=str(obs).strip()))

    # data do selo "Atualizado": a mais recente data do log (se existir), senao hoje.
    datas_validas = []
    if log_table_found:
        for r in range(log_header_row + 1, ws.max_row + 1):
            data_v = get(r, 2)
            if isinstance(data_v, datetime.datetime):
                datas_validas.append(data_v)
    upd_date = max(datas_validas) if datas_validas else None

    # Secoes de detalhe (novas) -- "Pendencias em Aberto": listas nominais de
    # placas por categoria, cada uma com titulo + cabecalho + linhas ate a
    # primeira Placa vazia.
    row_multas = _find_section_row(grid, ws.max_row, ["MULTAS EM ABERTO", "DETALHE"])
    row_contestar = _find_section_row(grid, ws.max_row, ["CONTESTAR", "DETALHE"])
    row_sinistros = _find_section_row(grid, ws.max_row, ["SINISTROS EM ABERTO", "DETALHE"])

    detalhe_multas = _read_detail_section(grid, row_multas, ws.max_row)
    detalhe_contestar = _read_detail_section(grid, row_contestar, ws.max_row)
    detalhe_sinistros = _read_detail_section(grid, row_sinistros, ws.max_row)

    # "Contestacao" (linha 8, indicador com contagem 2/1 tipico) nao tem secao
    # de detalhe propria na planilha -- e so uma formula COUNTIFS. Lemos a
    # lista direto da aba fonte (mesmo filtro da formula) para exibir no
    # dashboard; ver _read_contestacao_detail.
    try:
        detalhe_contestacao = _read_contestacao_detail(BOLETOS_XLSX)
    except Exception as e:
        log(f"  [aviso] Boletos: falha ao ler detalhe de Contestacao ({e!r}) -- secao ficara vazia.")
        detalhe_contestacao = []

    # Nota: a secao "BOLETOS/MULTAS PENDENTES DOWNLOAD — DETALHE (VERIFICACAO
    # AUTOMATICA)" existe na planilha mas e explicitamente marcada pela propria
    # planilha como possivelmente contendo falsos positivos (comparacao
    # automatica ainda nao conciliada com o indicador manual, que hoje mostra
    # 0/0) -- por isso NAO e lida/exibida aqui ainda.

    return dict(indicadores=indicadores, log_rows=log_rows, log_table_found=log_table_found,
                upd_date=upd_date, ultimo_boleto=ultimo_boleto,
                detalhe_multas=detalhe_multas, detalhe_contestar=detalhe_contestar,
                detalhe_sinistros=detalhe_sinistros, detalhe_contestacao=detalhe_contestacao)


def render_boletos(html, d):
    def sub_once(pattern, replacement, text, flags=re.S):
        new_text, n = re.subn(pattern, replacement, text, count=1, flags=flags)
        if n == 0:
            log(f"  [aviso] padrao nao encontrado (boletos): {pattern[:60]}...")
        return new_text

    hoje = (d["upd_date"] or datetime.datetime.now()).strftime('%d/%m/%Y')
    html = sub_once(r'<div class="badge-week" id="badgeAtualizado">Atualizado \d{2}/\d{2}/\d{4}</div>',
                     f'<div class="badge-week" id="badgeAtualizado">Atualizado {hoje}</div>', html)

    def fmt_data_cab(v):
        return v.strftime('%d/%m/%Y') if isinstance(v, datetime.datetime) else "-"

    vb_ult, vc_ult = d.get("ultimo_boleto", (None, None))
    ultimo_block = f'''<div class="locadoras-sm" id="ultimoBoletoSub" style="max-width:220px;">
        <div class="loc-badge-sm loc-unidas"><span class="n">{fmt_data_cab(vb_ult)}</span><span class="l">Unidas</span></div>
        <div class="loc-badge-sm loc-localiza"><span class="n">{fmt_data_cab(vc_ult)}</span><span class="l">Localiza</span></div>
      </div>'''
    html = sub_once(r'<div class="locadoras-sm" id="ultimoBoletoSub"[^>]*>.*?\n      </div>', ultimo_block, html)

    def fmt_val(v, typ):
        if v is None:
            return "-"
        if typ == "d" and isinstance(v, datetime.datetime):
            return v.strftime('%d/%m/%Y')
        if isinstance(v, (int, float)):
            return str(int(v)) if float(v).is_integer() else str(v)
        return str(v)

    # "Proximo a Vencer" nao tem mais card proprio -- vira uma linha pequena
    # dentro do card "Boletos a Vencer" (ver bloco dedicado logo abaixo).
    for (label, key, typ) in BOLETOS_INDICADORES:
        if key == "proximo":
            continue
        vb, vc = d["indicadores"].get(key, (None, None))
        block = f'''<div class="locadoras">
      <div class="loc-badge loc-unidas"><span class="n">{fmt_val(vb, typ)}</span><span class="l">Unidas</span></div>
      <div class="loc-badge loc-localiza"><span class="n">{fmt_val(vc, typ)}</span><span class="l">Localiza</span></div>'''
        pattern = r'(?i)(<div class="label">' + re.escape(label) + r'</div>\s*)<div class="locadoras">.*?(?=\n    </div>)'
        html = sub_once(pattern, lambda m, b=block: m.group(1) + b, html)

    vb_prox, vc_prox = d["indicadores"].get("proximo", (None, None))
    prox_block = f'''<div class="locadoras-sm" id="proximoSub">
      <div class="loc-badge-sm loc-unidas"><span class="n">{fmt_val(vb_prox, "d")}</span><span class="l">Próx. Unidas</span></div>
      <div class="loc-badge-sm loc-localiza"><span class="n">{fmt_val(vc_prox, "d")}</span><span class="l">Próx. Localiza</span></div>
    </div>'''
    html = sub_once(r'<div class="locadoras-sm" id="proximoSub">.*?\n    </div>', prox_block, html)

    # Log semanal manual (agora dentro do painel "Pendencias em Aberto -- Detalhe"):
    # se a tabela "SEMANA" ainda existir na planilha, mostra as linhas; a planilha
    # foi reestruturada e essa tabela pode nao existir mais -- nesse caso mostra o
    # mesmo estado vazio usado nas demais secoes de detalhe.
    if d.get("log_table_found") and d["log_rows"]:
        rows_html = ""
        for r in reversed(d["log_rows"]):
            rows_html += (f'    <tr><td>{r["semana"]}</td><td>{r["data"]}</td><td>{r["locadora"]}</td>'
                          f'<td><span class="tag tag-ok">{r["status"]}</span></td><td>{r["obs"]}</td></tr>\n')
    else:
        rows_html = '    <tr><td colspan="5" style="color:var(--muted); text-align:center;">Nenhuma pendência registrada</td></tr>\n'
    html = sub_once(r'<tbody id="logRows">\s*.*?\s*</tbody>', '<tbody id="logRows">\n' + rows_html + '    </tbody>', html)

    def fmt_detalhe_rows(rows):
        # cada linha de pendencia em aberto (multas/contestar/sinistros) leva
        # preenchimento amarelo (row-warn), inclusive novas linhas que aparecerem.
        if not rows:
            return '    <tr><td colspan="6" style="color:var(--muted); text-align:center;">Nenhuma pendência registrada</td></tr>\n'
        out = ""
        for r in rows:
            out += (f'    <tr class="row-warn"><td>{r["placa"]}</td><td>{r["locadora"]}</td><td>{r["parque"]}</td>'
                    f'<td>{r["venc"]}</td><td>{r["motivo"]}</td><td>{r["motivo_extra"]}</td></tr>\n')
        return out

    for tbody_id, key in (("detMultas", "detalhe_multas"), ("detContestar", "detalhe_contestar"),
                           ("detSinistros", "detalhe_sinistros"), ("detContestacao", "detalhe_contestacao")):
        rows_html = fmt_detalhe_rows(d.get(key, []))
        html = sub_once(rf'<tbody id="{tbody_id}">\s*.*?\s*</tbody>',
                         f'<tbody id="{tbody_id}">\n' + rows_html + '    </tbody>', html)

    return html


# ============================================================
# ENVIOS LOGISTICOS (requisicoes de materiais)
# ============================================================

def _find_label_row(grid, max_row, col, text_frags, max_scan=None):
    """Acha a primeira linha cuja celula (r, col) contem (case-insensitive)
    todos os fragmentos de texto informados."""
    limit = max_scan or max_row
    for r in range(1, limit + 1):
        v = grid.get((r, col))
        if isinstance(v, str) and v.strip():
            vu = v.strip().upper()
            if all(frag.upper() in vu for frag in text_frags):
                return r
    return None


def build_envios_data():
    # ---- KPIs / resumo em aberto / evolucao mensal: calculados direto da
    # aba "Sheet1" (fonte bruta das requisicoes). A aba "DASHBOARD (ROBSON)"
    # deixou de ter esse bloco pre-formatado (removido de proposito pela
    # rotina diaria de verificacao, que agora so mantem o Plano de Acao
    # nessa aba) -- entao calculamos aqui pra nao depender de layout fixo.
    #
    # Sheet1: header na linha 2 (linha 1 tem so os titulos-secao mesclados
    # "PREENCHIMENTO ALMOXARIFADO"/"PREENCHIMENTO LOGISTICA"), dados a partir
    # da linha 3. Colunas: A=ID, B=Data da Requisicao, C=Tipo, D=Responsavel,
    # E=Destinatario/Projeto, F=Status Almoxarifado, G=Data Prevista,
    # H=Data da Coleta, I=Forma de Entrega, J=Status Logistica,
    # K=Prazo solicitado, L=Entrega (data), M=Destinatario final,
    # N=Data do Recebimento Assinado.
    wb_raw = openpyxl.load_workbook(ENVIOS_XLSX, data_only=True, read_only=True)
    ws_raw = wb_raw['Sheet1']

    def norm_status(v):
        return v.strip().upper() if isinstance(v, str) else v

    ano_atual = datetime.date.today().year
    total_2026 = 0
    almox_2026 = 0     # todas as requisicoes de 2026 que passaram pelo Almoxarifado (nao vazio)
    entregues_2026 = 0
    cancel_2026 = 0
    aberto_sep = 0        # F = "A SEPARAR"
    aberto_coleta = 0     # F = "AGUARDANDO COLETA"
    aberto_trajeto = 0    # J = "EM TRAJETO"
    evo_by_month = defaultdict(lambda: [0, 0, 0])  # mes -> [requisicoes, entregues, canceladas]

    # Plano de Acao (Acoes Almoxarifado / Acoes Logistica): calculado direto
    # da Sheet1 a cada execucao (corrigido em 04/08/2026 -- antes vinha da
    # aba "DASHBOARD (ROBSON)", preenchida manualmente por outra rotina, que
    # ficava divergente da Sheet1 sempre que alguem esquecia de atualizar
    # uma linha ali; ex: requisicao Copel/ID 4074 "Em trajeto" que constava
    # no resumo mas nao aparecia na tabela). Ler direto da Sheet1 elimina
    # essa divergencia, pois e a mesma fonte usada nos KPIs/resumo acima.
    acoes_almox_raw = []
    acoes_log_raw = []

    # "Aguardando assinatura" (pedido do Robson em 05/08/2026): alem do status
    # textual "AGUARDANDO ASSINATURA" na coluna J (que na pratica nunca e
    # usado -- o Status Logistica normalmente ja vira "ENTREGUE" no momento da
    # entrega fisica), a planilha usa uma convencao visual manual: a celula da
    # coluna N (Data do Recebimento Assinado) fica com PREENCHIMENTO AMARELO
    # e sem data enquanto o comprovante assinado ainda nao voltou pro
    # Almoxarifado/Logistica. Isso vale independente do que a coluna J diga
    # (ex.: J="ENTREGUE" mas a assinatura do recebimento ainda esta pendente).
    # Detectamos isso pela cor de preenchimento (fgColor RGB) da celula N de
    # cada linha, nao por formula -- entao precisa reler a aba com o mesmo
    # arquivo (openpyxl consegue ler fill mesmo em modo read_only).
    aguardando_assinatura_raw = []

    for row in ws_raw.iter_rows(min_row=3):
        b_data_req = row[1].value if len(row) > 1 else None
        if not isinstance(b_data_req, datetime.datetime) or b_data_req.year < 2026:
            continue
        f_status = norm_status(row[5].value if len(row) > 5 else None)
        j_status = norm_status(row[9].value if len(row) > 9 else None)

        n_cell = row[13] if len(row) > 13 else None  # N = Data do Recebimento Assinado
        n_value = n_cell.value if n_cell is not None else None
        n_fill_rgb = None
        if n_cell is not None and n_cell.fill and n_cell.fill.fgColor:
            n_fill_rgb = n_cell.fill.fgColor.rgb
        n_vazio = n_value is None or (isinstance(n_value, str) and not n_value.strip())
        n_amarelo = isinstance(n_fill_rgb, str) and n_fill_rgb.upper().endswith("FFFF00")
        if n_amarelo and n_vazio and f_status != 'CANCELADO':
            projeto_e = row[4].value if len(row) > 4 else None
            responsavel_d = row[3].value if len(row) > 3 else None
            prazo_g = row[6].value if len(row) > 6 else None
            destinatario_m = row[12].value if len(row) > 12 else None
            data_entrega_l = row[11].value if len(row) > 11 else None
            aguardando_assinatura_raw.append([projeto_e, responsavel_d, prazo_g, destinatario_m, data_entrega_l])

        total_2026 += 1
        if f_status:
            almox_2026 += 1
        if f_status == 'ENTREGUE' or j_status == 'ENTREGUE':
            entregues_2026 += 1
        if f_status == 'CANCELADO':
            cancel_2026 += 1
        if f_status == 'A SEPARAR':
            aberto_sep += 1
        if f_status == 'AGUARDANDO COLETA':
            aberto_coleta += 1
        if j_status == 'EM TRAJETO':
            aberto_trajeto += 1

        if b_data_req.year == ano_atual:
            mes_idx = b_data_req.month - 1
            evo_by_month[mes_idx][0] += 1
            if f_status == 'ENTREGUE' or j_status == 'ENTREGUE':
                evo_by_month[mes_idx][1] += 1
            if f_status == 'CANCELADO':
                evo_by_month[mes_idx][2] += 1

        projeto = row[4].value if len(row) > 4 else None  # E = Destinatario/Projeto
        responsavel = row[3].value if len(row) > 3 else None  # D
        prazo = row[6].value if len(row) > 6 else None  # G = Data Prevista
        destinatario_final = row[12].value if len(row) > 12 else None  # M
        data_entrega = row[11].value if len(row) > 11 else None  # L = Entrega

        # Almoxarifado: mesmos status que contam como "em aberto" no resumo
        # acima (A SEPARAR / AGUARDANDO COLETA) -- "ENVIADO" e status
        # intermediario (ja saiu do almoxarifado), nao e pendencia dele.
        if f_status in ('A SEPARAR', 'AGUARDANDO COLETA'):
            acoes_almox_raw.append([projeto, responsavel, prazo, destinatario_final, row[5].value])

        # Logistica: EM TRAJETO (mesmo criterio do resumo) ou AGUARDANDO
        # ASSINATURA (etapa final antes de ENTREGUE, tambem pendente).
        if j_status in ('EM TRAJETO', 'AGUARDANDO ASSINATURA'):
            acoes_log_raw.append([projeto, responsavel, prazo, destinatario_final, data_entrega, row[9].value])

    kpi_total = total_2026
    kpi_almox = almox_2026
    kpi_entregues = entregues_2026
    kpi_cancel = cancel_2026

    aberto_rows = [
        ("Em separação", aberto_sep),
        ("Aguardando coleta", aberto_coleta),
        ("Em trajeto", aberto_trajeto),
    ]
    aberto_total = aberto_sep + aberto_coleta + aberto_trajeto

    evo_note = ("Base: mês da Data da Requisição (Sheet1, coluna B), calculado "
                "diretamente a cada execução.")
    evo_rows = []
    for mes_idx in range(12):
        if mes_idx in evo_by_month:
            req_v, ent_v, _cancel_v = evo_by_month[mes_idx]
            evo_rows.append((MESES_PT[mes_idx], req_v, ent_v))
    while evo_rows and evo_rows[-1][1] == 0 and evo_rows[-1][2] == 0:
        evo_rows.pop()

    def fmt_cell(v):
        if isinstance(v, datetime.datetime):
            return v.strftime('%d/%m/%Y')
        if v is None:
            return ""
        return str(v).strip()

    acoes_almox = [dict(projeto=fmt_cell(v[0]), responsavel=fmt_cell(v[1]), prazo=fmt_cell(v[2]),
                         destinatario=fmt_cell(v[3]), status=fmt_cell(v[4])) for v in acoes_almox_raw]
    acoes_log = [dict(projeto=fmt_cell(v[0]), responsavel=fmt_cell(v[1]), prazo=fmt_cell(v[2]),
                       destinatario=fmt_cell(v[3]), data_entrega=fmt_cell(v[4]), status=fmt_cell(v[5])) for v in acoes_log_raw]
    aguardando_assinatura = [dict(projeto=fmt_cell(v[0]), responsavel=fmt_cell(v[1]), prazo=fmt_cell(v[2]),
                                   destinatario=fmt_cell(v[3]), data_entrega=fmt_cell(v[4])) for v in aguardando_assinatura_raw]

    return dict(
        kpi_total=kpi_total, kpi_almox=kpi_almox, kpi_entregues=kpi_entregues, kpi_cancel=kpi_cancel,
        aberto_rows=aberto_rows, aberto_total=aberto_total,
        evo_note=evo_note, evo_rows=evo_rows,
        acoes_almox=acoes_almox, acoes_log=acoes_log,
        aguardando_assinatura=aguardando_assinatura,
    )


def render_envios(html, d):
    def sub_once(pattern, replacement, text, flags=re.S):
        new_text, n = re.subn(pattern, replacement, text, count=1, flags=flags)
        if n == 0:
            log(f"  [aviso] padrao nao encontrado (envios): {pattern[:60]}...")
        return new_text

    hoje = datetime.date.today().strftime('%d/%m/%Y')
    html = sub_once(r'<div class="badge-week" id="badgeAtualizado">Atualizado \d{2}/\d{2}/\d{4}</div>',
                     f'<div class="badge-week" id="badgeAtualizado">Atualizado {hoje}</div>', html)

    def n(v):
        if v is None:
            return "-"
        if isinstance(v, (int, float)):
            return str(int(v)) if float(v).is_integer() else str(v)
        return str(v)

    html = sub_once(r'(<div class="n" id="kpiTotal">)(?:\d+|-)(</div>)', lambda m: m.group(1) + n(d["kpi_total"]) + m.group(2), html)
    html = sub_once(r'(<div class="n" id="kpiAlmox">)(?:\d+|-)(</div>)', lambda m: m.group(1) + n(d["kpi_almox"]) + m.group(2), html)
    html = sub_once(r'(<div class="n" id="kpiEntregues">)(?:\d+|-)(</div>)', lambda m: m.group(1) + n(d["kpi_entregues"]) + m.group(2), html)
    html = sub_once(r'(<div class="n" id="kpiCanceladas">)(?:\d+|-)(</div>)', lambda m: m.group(1) + n(d["kpi_cancel"]) + m.group(2), html)

    # Resumo em aberto -- os 3 primeiros rotulos conhecidos vao para os IDs fixos
    # do template; qualquer rotulo extra/diferente e ignorado no badge fixo mas
    # ainda soma pro total.
    aberto_map = {label.strip().upper(): qtde for (label, qtde) in d["aberto_rows"]}
    v_sep = aberto_map.get("EM SEPARAÇÃO", aberto_map.get("EM SEPARACAO", 0))
    v_col = aberto_map.get("AGUARDANDO COLETA", 0)
    v_traj = aberto_map.get("EM TRAJETO", 0)
    total_ab = d["aberto_total"] if d["aberto_total"] is not None else (v_sep + v_col + v_traj)
    html = sub_once(r'(<div class="n" id="abEmSeparacao">)\d+(</div>)', lambda m: m.group(1) + n(v_sep) + m.group(2), html)
    html = sub_once(r'(<div class="n" id="abAguardandoColeta">)\d+(</div>)', lambda m: m.group(1) + n(v_col) + m.group(2), html)
    html = sub_once(r'(<div class="n" id="abEmTrajeto">)\d+(</div>)', lambda m: m.group(1) + n(v_traj) + m.group(2), html)
    html = sub_once(r'(<div class="n" id="abTotal">)\d+(</div>)', lambda m: m.group(1) + n(total_ab) + m.group(2), html)

    if d.get("evo_note"):
        note_txt = str(d["evo_note"]).strip()
        html = sub_once(r'(<p class="panel-note" id="evolucaoNote">).*?(</p>)',
                         lambda m: m.group(1) + note_txt + m.group(2), html)

    meses_js = js_str([mes for (mes, _, _) in d["evo_rows"]])
    req_js = json.dumps([int(req_v) if isinstance(req_v, (int, float)) else 0 for (_, req_v, _) in d["evo_rows"]])
    ent_js = json.dumps([int(ent_v) if isinstance(ent_v, (int, float)) else 0 for (_, _, ent_v) in d["evo_rows"]])
    html = sub_once(r'var evolucaoMeses = \[.*?\];', f'var evolucaoMeses = {meses_js};', html)
    html = sub_once(r'var evolucaoReq = \[.*?\];', f'var evolucaoReq = {req_js};', html)
    html = sub_once(r'var evolucaoEnt = \[.*?\];', f'var evolucaoEnt = {ent_js};', html)

    html = sub_once(r'(<div class="n" id="acaoCountAlmox">)\d+(</div>)', lambda m: m.group(1) + n(len(d["acoes_almox"])) + m.group(2), html)
    html = sub_once(r'(<div class="n" id="acaoCountLog">)\d+(</div>)', lambda m: m.group(1) + n(len(d["acoes_log"])) + m.group(2), html)

    if d["acoes_almox"]:
        rows_html = ""
        for r in d["acoes_almox"]:
            warn = ' class="row-warn"' if "divergência" in r["status"].lower() or "divergencia" in r["status"].lower() else ""
            rows_html += (f'    <tr{warn}><td>{r["projeto"]}</td><td>{r["responsavel"]}</td><td>{r["prazo"]}</td>'
                          f'<td>{r["destinatario"]}</td><td>{r["status"]}</td></tr>\n')
    else:
        rows_html = '    <tr><td colspan="5" style="color:var(--muted); text-align:center;">Nenhuma ação pendente</td></tr>\n'
    html = sub_once(r'<tbody id="acoesAlmox">\s*.*?\s*</tbody>', '<tbody id="acoesAlmox">\n' + rows_html + '    </tbody>', html)

    if d["acoes_log"]:
        rows_html = ""
        for r in d["acoes_log"]:
            status_l = r["status"].lower()
            warn = ' class="row-warn"' if ("em trajeto" in status_l or "aguardando assinatura" in status_l) else ""
            rows_html += (f'    <tr{warn}><td>{r["projeto"]}</td><td>{r["responsavel"]}</td><td>{r["prazo"]}</td>'
                          f'<td>{r["destinatario"]}</td><td>{r["data_entrega"]}</td><td>{r["status"]}</td></tr>\n')
    else:
        rows_html = '    <tr><td colspan="6" style="color:var(--muted); text-align:center;">Nenhuma ação pendente</td></tr>\n'
    html = sub_once(r'<tbody id="acoesLog">\s*.*?\s*</tbody>', '<tbody id="acoesLog">\n' + rows_html + '    </tbody>', html)

    # "Aguardando assinatura" (pedido do Robson em 05/08/2026): quadro proprio,
    # separado das Acoes Logistica -- entrega/coleta amarela sem data na
    # coluna "Data do Recebimento Assinado" (Sheet1, coluna N), independente
    # do que a coluna Status Logistica diga.
    html = sub_once(r'(<div class="n" id="acaoCountAssinatura">)\d+(</div>)',
                     lambda m: m.group(1) + n(len(d["aguardando_assinatura"])) + m.group(2), html)
    if d["aguardando_assinatura"]:
        rows_html = ""
        for r in d["aguardando_assinatura"]:
            rows_html += (f'    <tr class="row-warn"><td>{r["projeto"]}</td><td>{r["responsavel"]}</td><td>{r["prazo"]}</td>'
                          f'<td>{r["destinatario"]}</td><td>{r["data_entrega"]}</td></tr>\n')
    else:
        rows_html = '    <tr><td colspan="5" style="color:var(--muted); text-align:center;">Nenhuma pendência registrada</td></tr>\n'
    html = sub_once(r'<tbody id="acoesAssinatura">\s*.*?\s*</tbody>', '<tbody id="acoesAssinatura">\n' + rows_html + '    </tbody>', html)

    # Nota de rodape: Plano de Acao agora vem direto da Sheet1 (corrigido em
    # 04/08/2026), nao mais da aba DASHBOARD (ROBSON) preenchida a mao.
    html = sub_once(
        r'<div class="note" id="fonteNote">Fonte: planilha "Controle_Logistico_Requisicoes\.xlsx", aba (?:DASHBOARD \(ROBSON\)|Sheet1 \(dados brutos, calculado diretamente a cada execução\))\. ',
        '<div class="note" id="fonteNote">Fonte: planilha "Controle_Logistico_Requisicoes.xlsx", aba Sheet1 (dados brutos, calculado diretamente a cada execução). ',
        html)

    return html


# ============================================================
# REEMBOLSO (notas avulsas do dia-a-dia -- sem coluna de status na planilha,
# entao este dashboard mostra HISTORICO consolidado (mes atual, semana atual,
# ranking por responsavel/categoria, ultimos lancamentos), nao uma fila de
# pendencias. Por isso NAO entra em "Nossas Demandas" (decisao do Robson).
# ============================================================

MESES_FULL_PT = ['JANEIRO','FEVEREIRO','MARÇO','ABRIL','MAIO','JUNHO','JULHO',
                  'AGOSTO','SETEMBRO','OUTUBRO','NOVEMBRO','DEZEMBRO']


def fmt_brl(v):
    if v is None:
        return "-"
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    s = f"{v:,.2f}"
    s = s.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"R$ {s}"


def _find_week_blocks(grid, max_row, col_label=10):
    """Varre a coluna 'TOTAL X SEMANA' (col10/J) da aba REEMBOLSO procurando o
    rotulo de inicio de cada bloco semanal ('W12' OU so '12' -- a planilha
    tem essa mesma inconsistencia de rotulo que ja vimos na aba de Frota,
    numa faixa de semanas onde o 'W' foi digitado sem o prefixo). Devolve
    {numero_da_semana: linha_do_bloco}, pegando so a PRIMEIRA ocorrencia de
    cada semana (a linha de inicio do bloco, onde tambem fica a formula de
    FUNCIONARIOS ATIVOS na coluna ao lado)."""
    out = {}
    for r in range(1, max_row + 1):
        v = grid.get((r, col_label))
        wk = None
        if isinstance(v, str):
            m = re.match(r'^W?\s*(\d+)$', v.strip(), re.I)
            if m:
                wk = int(m.group(1))
        elif isinstance(v, (int, float)) and float(v).is_integer():
            wk = int(v)
        if wk is not None and 1 <= wk <= 60 and wk not in out:
            out[wk] = r
    return out


def _eval_simple_sum_formula(v):
    """Avalia com seguranca o valor da coluna FUNCIONARIOS ATIVOS: pode vir
    como numero literal (comum no Master de anos ja fechados) ou como formula
    de soma pura de literais tipo '=30+12+3' (comum no Master do ano corrente
    -- o cache dessas formulas some sempre que outra automacao resalva o
    arquivo via openpyxl, que nao recalcula formulas). So aceita digitos,
    + - . e espaco antes de avaliar; qualquer outro caractere (referencia de
    celula, funcao, texto) faz devolver None em vez de arriscar um eval
    perigoso ou inventar um numero."""
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str):
        return None
    s = v.strip()
    if s.startswith('='):
        s = s[1:]
    if not s or not re.fullmatch(r'[\d+\-.\s]+', s):
        return None
    try:
        return float(eval(s, {"__builtins__": {}}, {}))
    except Exception:
        return None


def _media_ano_anterior(ano):
    """Abre o Master de reembolso de um ano ANTERIOR (se o arquivo existir) e
    recalcula, direto dos lancamentos e blocos semanais brutos (nunca de
    formula/total pronto da planilha): media semanal de valor (R$) e media
    semanal de tecnicos ativos. Devolve None (sem inventar nada) se o arquivo
    ou a aba nao existir, ou se nao houver nenhum lancamento valido -- e o
    caso esperado pra anos sem Master (ex: 2024, se nao houver arquivo)."""
    try:
        base = _find_sibling("REEMBOLSO SEMANAL")
        path = os.path.join(base, f"REEMBOLSOS {ano}", f"Formulário de Reembolso Master ({ano}).xlsx")
        if not os.path.isfile(path):
            return None
        # Duas leituras do mesmo arquivo -- data_only=True pra tudo que e
        # literal (Valor, Semana, Projeto etc: usar True aqui pega o valor
        # certo mesmo nos raros lancamentos onde "Valor" acaba sendo uma
        # formula, ex: correcao lancada por referencia a outra celula -- usar
        # False faria esses lancamentos sumirem da soma, ja aconteceu e
        # gerou uma media errada por ~R$19/semana numa primeira versao desse
        # calculo) -- e data_only=False so pra coluna L (FUNCIONARIOS ATIVOS),
        # que pode ser formula de soma pura e cujo cache pode estar vazio.
        wb_true = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if "REEMBOLSO" not in wb_true.sheetnames:
            return None
        ws_true = wb_true["REEMBOLSO"]
        grid = sheet_to_grid(ws_true)
        header_row = None
        for r in range(1, 20):
            if grid.get((r, 2)) == "Projeto":
                header_row = r
                break
        if not header_row:
            return None
        total = 0.0
        semanas = set()
        for r in range(header_row + 1, ws_true.max_row + 1):
            valor = grid.get((r, 9))
            semana = grid.get((r, 8))
            if isinstance(valor, (int, float)):
                total += valor
                if isinstance(semana, (int, float)):
                    semanas.add(int(semana))
        if not semanas:
            return None
        media_valor = round(total / len(semanas), 2)

        wb_formula = openpyxl.load_workbook(path, data_only=False, read_only=True)
        ws_formula = wb_formula["REEMBOLSO"]
        grid_l = {}
        for row in ws_formula.iter_rows(min_row=header_row, max_row=ws_formula.max_row, min_col=10, max_col=12):
            j, l = row[0].value, row[2].value
            if j is not None or l is not None:
                grid_l[row[0].row] = (j, l)

        blocos = _find_week_blocks({(r, 10): v[0] for r, v in grid_l.items()}, ws_formula.max_row)
        tecnicos_vals = []
        for wk, row in blocos.items():
            t = _eval_simple_sum_formula(grid_l.get(row, (None, None))[1])
            if t is not None:
                tecnicos_vals.append(t)
        media_tecnicos = rhu(sum(tecnicos_vals) / len(tecnicos_vals)) if tecnicos_vals else None

        return dict(valor=media_valor, tecnicos=media_tecnicos)
    except Exception as e:
        log(f"  [risco] nao foi possivel calcular media de {ano} pro grafico de evolucao: {e!r}")
        return None


def _norm_parque_key(s):
    """Normaliza nome de parque/projeto pra cruzar "Projeto" (Reembolso) com
    "PARQUE" (aba DADOS do Diario de Bordo) mesmo quando a grafia varia um
    pouco entre planilhas (ex: "LDV1 E 2/PI" vs "LDV 1 e 2/PI", ou um sufixo
    " - SOLO" numa das duas). Maiuscula, tira sufixo "SOLO" no final, e
    remove tudo que nao for letra/numero/barra (espaco, hifen, etc)."""
    s = s.strip().upper()
    s = re.sub(r'\s*-?\s*SOLO\s*$', '', s)
    s = re.sub(r'[^A-Z0-9/]', '', s)
    return s


def _load_parque_cliente_map():
    """Le a aba DADOS do Diario de Bordo (historico de mobilizacao/
    desmobilizacao, que tem PARQUE e CLIENTE lado a lado) e monta um mapa
    parque_normalizado -> cliente mais frequente naquele parque. Usado pra
    identificar o cliente de cada lancamento de Reembolso a partir do campo
    "Projeto" (que normalmente e o nome do parque). Quando o mesmo parque
    aparece com mais de um cliente ao longo do historico (ex: TRAIRI/CE, que
    teve equipe GE e equipe SIEMENS em paralelo em periodos diferentes),
    ficamos com o cliente mais frequente -- e uma aproximacao, nao 100%
    precisa pra esses parques compartilhados."""
    try:
        wb = openpyxl.load_workbook(EQUIPES_XLSX, data_only=True, read_only=True)
        ws = wb['DADOS']
        agg = defaultdict(Counter)
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=10, max_col=11):
            parque, cliente = row[0].value, row[1].value
            if isinstance(parque, str) and parque.strip() and isinstance(cliente, str) and cliente.strip():
                agg[_norm_parque_key(parque)][cliente.strip()] += 1
        return {k: v.most_common(1)[0][0] for k, v in agg.items()}
    except Exception as e:
        log(f"  [risco] nao foi possivel carregar mapa parque->cliente (grafico Reembolso x Cliente): {e!r}")
        return {}


# Correcoes manuais do Robson pra projetos do Reembolso que nao batem com
# nenhum parque do historico do Diario de Bordo (grafia diferente, projeto
# novo ainda sem evento de mobilizacao/desmobilizacao registrado, etc.) --
# checadas antes do mapa automatico e do fallback por palavra-chave.
MANUAL_CLIENTE_OVERRIDES = {
    _norm_parque_key("VAM & VPA/RN"): "NORDEX / LWS",
    _norm_parque_key("GALPÃO SIMÕES FILHO/BA"): "SIEMENS",
    _norm_parque_key("LIVRAMENTO/RS"): "SIEMENS",
}


# ---- Regras de categorizacao das descricoes de lancamento de Reembolso
# (usadas so no 2o Pareto, estratificado dentro das 2 maiores sub-atividades).
# ORDEM IMPORTA: a primeira regra cujo padrao bater com a descricao decide a
# categoria (first-match-wins). Definida em rodada extensa com o Robson,
# incluindo resolucao explicita de conflitos:
#   - "Rolos de Laminacao" e "Fita Crepe" tem prioridade SOBRE "Espatulas
#     Celuloides" (ex: "Rolos de Laminacao e Espatulas" fica em Rolos de
#     Laminacao, nao em Espatulas) -- decisao explicita do Robson.
DESC_CATEGORY_RULES = [
    ("Rolos de Laminação", r"rolos?\s+de\s+lamin"),
    ("Fita Crepe", r"fita\s*crepe|fita\s*adesiva"),
    ("Espátulas Celuloídes", r"esp[aá]tulas?"),
    ("Translados Mobs ou Desmobs", r"translad|uber|t[aá]xi|passagem|taxa"),
    ("Hospedagens de Apoio Logístico", r"hospedage|hospedag"),
    ("Abastecimento de Combustível", r"abastec"),
    ("Pedágio", r"ped[aá]gio"),
    ("Material de Balanceamento", r"balancea"),
    ("Material de Limpeza", r"material\s+de\s+(el[ée]tric|limpez)|sacos?\s+de\s+(confeit|lixo)|vassoura|pinc[ée]is"),
    ("Utensílhos", r"[aá]lcool\s*gel|pilhas?|cola\b|peneira"),
    ("Consumíveis", r"discos?\s+de?\s*lixa"),
    ("Medicamentos", r"rem[ée]dio"),
    ("Confraternização Técnicos", r"caf[ée]\s+da\s+manh|caf[ée]\s+da\s+tarde|alimenta[çc][ãa]o|confraterniza"),
]
_DESC_CATEGORY_RULES_COMPILED = [(name, re.compile(pat, re.IGNORECASE)) for name, pat in DESC_CATEGORY_RULES]


def categorize_reembolso_desc(desc):
    """Aplica DESC_CATEGORY_RULES (first-match-wins) pra agrupar a descricao
    bruta do lancamento numa categoria consolidada. O que nao bater com
    nenhuma regra mantem a propria descricao original (categoria "de 1 item")."""
    if not desc:
        return "(sem descrição)"
    for name, pat in _DESC_CATEGORY_RULES_COMPILED:
        if pat.search(desc):
            return name
    return desc


def build_reembolso_data():
    # data_only=True: Semana/Valor/Projeto/rotulos/QTDD SEMANA sao literais
    # digitados a mao -- data_only=True le certo. ATENCAO: alguns poucos
    # lancamentos de "Valor" acabam sendo formula (ex: correcao referenciando
    # outra celula) -- usar data_only=False aqui faria esses lancamentos
    # sumirem da soma (ja aconteceu numa versao anterior deste calculo e
    # gerou um total ~R$300 menor). A coluna L (FUNCIONARIOS ATIVOS), essa
    # sim as vezes formula com cache vazio, e lida separadamente mais abaixo
    # com data_only=False, so pra ela.
    wb = openpyxl.load_workbook(REEMBOLSO_XLSX, data_only=True, read_only=True)
    ws = wb['REEMBOLSO']
    grid = sheet_to_grid(ws)

    # ---- Tabela "CONTROLE MENSAL": so aproveitamos dela o ROTULO do mes (col I)
    # e a "QTDD SEMANA" (col K) -- essa ultima e um numero digitado a mao, NAO
    # formula. O "Valor" (col J) e a "MEDIA SEMANAL" (col L) SAO formulas em
    # cadeia (J{mes} = K{semana1}+K{semana2}+...; cada K{semana} por sua vez e
    # SUM(...) das linhas daquela semana) -- e o openpyxl NAO recalcula
    # formulas. Toda vez que outra automacao (ex: "fechar-semana-reembolso")
    # abre e resalva esse arquivo via openpyxl, o cache do valor calculado
    # dessas formulas some e data_only=True passa a ler None/0 ali. Por isso
    # NUNCA confiamos em J/L -- os totais de valor sao sempre recalculados
    # aqui a partir dos lancamentos brutos (coluna "Semana" + "Valor", que sao
    # digitados a mao, nao formula, e por isso imunes a esse problema).
    meses_qtd = []  # [(rotulo, qtd_semanas), ...] na ordem em que aparecem
    for r in range(1, 40):
        label = grid.get((r, 9))
        if not isinstance(label, str):
            continue
        lbl = label.strip().upper()
        if lbl in MESES_FULL_PT:
            qtd = grid.get((r, 11))
            meses_qtd.append((lbl.capitalize(), qtd if isinstance(qtd, (int, float)) else None))

    # ---- Lancamentos brutos (linha de cabecalho "Projeto" na coluna B, dados
    # logo abaixo ate o fim da planilha) ----
    header_row = None
    for r in range(1, 20):
        if grid.get((r, 2)) == "Projeto":
            header_row = r
            break
    rows = []
    if header_row:
        for r in range(header_row + 1, ws.max_row + 1):
            valor = grid.get((r, 9))
            if not isinstance(valor, (int, float)):
                continue
            projeto = grid.get((r, 2))
            nome = grid.get((r, 4))
            resp = grid.get((r, 5))
            sub = grid.get((r, 6))
            desc = grid.get((r, 7))
            semana = grid.get((r, 8))
            rows.append(dict(
                projeto=str(projeto).strip() if projeto else "-",
                nome=str(nome).strip() if nome else "-",
                resp=str(resp).strip() if isinstance(resp, str) and resp.strip() else "-",
                sub=str(sub).strip() if isinstance(sub, str) and sub.strip() else "-",
                desc=str(desc).strip() if isinstance(desc, str) else "",
                semana=semana if isinstance(semana, (int, float)) else None,
                valor=float(valor),
            ))

    n_lancamentos = len(rows)

    # ---- Reembolso por Cliente (pedido do Robson: pizza de % de reembolso
    # por cliente em 2026). O campo "Projeto" do Reembolso normalmente e o
    # nome do parque -- cruzamos com o mapa parque->cliente construido a
    # partir do historico da aba DADOS do Diario de Bordo. "EXTREME" e um
    # projeto interno (nao e parque de cliente nenhum) -- a pedido do Robson,
    # esses lancamentos entram na conta da GE. O que sobra sem match no mapa
    # tenta reconhecer o nome do cliente direto no texto do projeto (ex: "VDB
    # NORDEX/BA"); o que nao for identificado de nenhuma forma vira "Nao
    # identificado" -- preferimos isso a adivinhar e atribuir errado.
    parque_cliente_map = _load_parque_cliente_map()
    cliente_valor = defaultdict(float)
    cliente_qtd = defaultdict(int)
    cliente_projetos = defaultdict(set)
    for row in rows:
        proj_key = _norm_parque_key(row["projeto"]) if row["projeto"] and row["projeto"] != "-" else ""
        if proj_key == "EXTREME":
            cliente = "GE"
        elif proj_key in MANUAL_CLIENTE_OVERRIDES:
            cliente = MANUAL_CLIENTE_OVERRIDES[proj_key]
        elif proj_key in parque_cliente_map:
            cliente = parque_cliente_map[proj_key]
        else:
            cliente = "Não identificado"
            proj_upper = row["projeto"].upper()
            for kw in CLIENTES_TEC:
                kw_check = "NORDEX" if kw == "NORDEX / LWS" else kw
                if kw_check in proj_upper:
                    cliente = kw
                    break
        cliente_valor[cliente] += row["valor"]
        cliente_qtd[cliente] += 1
        cliente_projetos[cliente].add(row["projeto"])
    total_cliente_geral = sum(cliente_valor.values()) or 1
    ranking_cliente_reemb = sorted(
        ({"nome": k, "valor": v, "qtd": cliente_qtd[k], "pct": round(100 * v / total_cliente_geral, 1),
          "n_projetos": len(cliente_projetos[k])} for k, v in cliente_valor.items()),
        key=lambda x: -x["valor"])

    # ---- Monta os totais mensais somando os lancamentos brutos por faixa de
    # semana (1-indexado), usando a "QTDD SEMANA" (literal) pra saber quantas
    # semanas cada mes cobre -- ex: Jan=semanas 1-4, Fev=5-9, etc. O mes mais
    # recente cobre ate a ultima semana com lancamento (pode estar em
    # andamento -- a semana atual, ainda incompleta, ja fica incluida nele,
    # que e como a planilha original tambem calculava).
    semana_valor = defaultdict(float)
    for row in rows:
        if row["semana"] is not None:
            semana_valor[int(row["semana"])] += row["valor"]
    max_semana = max(semana_valor.keys(), default=None)

    monthly = []
    semana_cursor = 1
    for mes, qtd in meses_qtd:
        if qtd is None:
            continue
        semana_ini, semana_fim = semana_cursor, semana_cursor + int(qtd) - 1
        valor_mes = sum(v for wk, v in semana_valor.items() if semana_ini <= wk <= semana_fim)
        monthly.append(dict(mes=mes, valor=valor_mes, qtd=int(qtd),
                             media=round(valor_mes / qtd, 2) if qtd else None,
                             semana_ini=semana_ini, semana_fim=semana_fim))
        semana_cursor = semana_fim + 1

    mes_atual = monthly[-1] if monthly else None
    semana_atual_label = f"W{max_semana}" if max_semana is not None else None
    semana_atual_valor = semana_valor.get(max_semana, 0) if max_semana is not None else 0
    # total do ano = soma de TODOS os lancamentos (equivalente a soma dos
    # meses, ja que as faixas de semana cobrem tudo) -- nao soma a semana
    # atual "por fora": ela ja esta dentro do mes corrente, tanto aqui quanto
    # na tabela original da planilha (a linha "Wxx" ali e so uma exibicao
    # isolada da semana em andamento, nao um valor adicional).
    total_ano = sum(row["valor"] for row in rows)

    # ---- Tecnicos ativos por semana (coluna FUNCIONARIOS ATIVOS, col12/L) --
    # lida via deteccao de bloco semanal (rotulo "Wxx"/"xx" na col10/J) +
    # avaliacao segura da formula/literal, pedido do Robson pra acompanhar
    # junto do valor no grafico. Recalculado aqui em vez de confiar no cache
    # da planilha, que fica None sempre que outra automacao resalva o arquivo
    # via openpyxl.
    week_blocks = _find_week_blocks(grid, ws.max_row)
    # Leitura separada em modo "formula" (data_only=False) so pra coluna L --
    # 1 unica varredura sequencial da coluna inteira (rapida, mesmo padrao de
    # sheet_to_grid), nao uma busca por linha de cada vez (acesso aleatorio
    # e muito lento em modo read_only, ja vimos isso travar em outra aba).
    wb_formula = openpyxl.load_workbook(REEMBOLSO_XLSX, data_only=False, read_only=True)
    ws_formula = wb_formula['REEMBOLSO']
    l_col = {}
    for cell_row in ws_formula.iter_rows(min_row=1, max_row=ws_formula.max_row, min_col=12, max_col=12):
        if cell_row[0].value is not None:
            l_col[cell_row[0].row] = cell_row[0].value

    tecnicos_semana = {}
    for wk, row in week_blocks.items():
        t = _eval_simple_sum_formula(l_col.get(row))
        if t is not None:
            tecnicos_semana[wk] = t

    # Media de tecnicos ativos por mes (pedido do Robson, tabela "Evolucao
    # Mensal 2026") -- media das semanas daquele mes, mesma logica usada no
    # grafico de evolucao por trimestre, so que por mes.
    for m in monthly:
        vals_mes = [t for wk, t in tecnicos_semana.items() if m["semana_ini"] <= wk <= m["semana_fim"]]
        m["tecnicos_media"] = rhu(sum(vals_mes) / len(vals_mes)) if vals_mes else None
        # Valor medio por tecnico = media semanal (R$) dividido pela qtd media
        # de tecnicos ativos no mes -- pedido do Robson pra ultima coluna da
        # tabela "Evolucao Mensal 2026".
        m["valor_por_tecnico"] = (round(m["media"] / m["tecnicos_media"], 2)
                                   if m["tecnicos_media"] else None)

    # ---- Grafico de evolucao (pedido do Robson): telescopia no tempo --
    # o ano ANTERIOR fechado (ano_atual-1) vira 1 barra de media semanal de
    # valor + media semanal de tecnicos ativos; trimestres FECHADOS do ano
    # corrente (todos os 3 meses ja presentes na CONTROLE MENSAL) viram 1
    # barra cada (media de valor E media de tecnicos ativos das semanas do
    # trimestre); o trimestre ainda EM ANDAMENTO (meses incompletos) mostra 1
    # barra POR SEMANA ate a ultima semana lancada -- assim que esse
    # trimestre fechar (proxima execucao, quando os 3 meses dele estiverem
    # completos), ele vira automaticamente 1 barra so, e o trimestre seguinte
    # comeca a mostrar semana a semana. Nada aqui e hardcoded por ano/
    # trimestre -- e recalculado a cada execucao a partir da data de hoje e
    # do que existir na planilha. Ano anterior-anterior (ex: 2024) foi
    # removido a pedido do Robson -- nunca houve arquivo/dado confiavel pra
    # ele (referencia quebrada na planilha de origem).
    ano_atual = datetime.date.today().year
    evolucao = []
    nota_evolucao = None

    ano_hist = ano_atual - 1
    media_ant = _media_ano_anterior(ano_hist)
    if media_ant is None:
        evolucao.append(dict(label=f"Média {ano_hist}", valor=None, tecnicos=None, tipo="ano"))
        nota_evolucao = (f"Média {ano_hist} sem dado disponível (nenhum arquivo de Master de reembolso "
                          f"encontrado para {ano_hist}). ")
    else:
        evolucao.append(dict(label=f"Média {ano_hist}", valor=media_ant["valor"],
                              tecnicos=media_ant["tecnicos"], tipo="ano"))

    def _media_tecnicos(semanas_ini, semanas_fim):
        vals = [t for wk, t in tecnicos_semana.items() if semanas_ini <= wk <= semanas_fim]
        return rhu(sum(vals) / len(vals)) if vals else None

    TRI_LABELS = ["1º Tri", "2º Tri", "3º Tri", "4º Tri"]
    for qi in range(0, len(monthly), 3):
        grupo = monthly[qi:qi + 3]
        tri_num = qi // 3
        if len(grupo) == 3:
            valor_tri = sum(g["valor"] for g in grupo)
            qtd_semanas_tri = sum(g["qtd"] for g in grupo)
            media_tri = round(valor_tri / qtd_semanas_tri, 2) if qtd_semanas_tri else None
            tecnicos_tri = _media_tecnicos(grupo[0]["semana_ini"], grupo[-1]["semana_fim"])
            evolucao.append(dict(label=f"{TRI_LABELS[tri_num]} {ano_atual}", valor=media_tri,
                                  tecnicos=tecnicos_tri, tipo="trimestre"))
        else:
            # trimestre em andamento -- 1 barra por semana, da 1a semana desse
            # trimestre ate a ultima semana com lancamento na planilha.
            ini = grupo[0]["semana_ini"]
            for wk in range(ini, (max_semana or ini) + 1):
                evolucao.append(dict(label=f"W{wk}", valor=round(semana_valor.get(wk, 0), 2),
                                      tecnicos=(rhu(tecnicos_semana[wk]) if wk in tecnicos_semana else None),
                                      tipo="semana"))

    # ---- Ranking por responsavel ----
    resp_valor = defaultdict(float)
    resp_qtd = defaultdict(int)
    for row in rows:
        resp_valor[row["resp"]] += row["valor"]
        resp_qtd[row["resp"]] += 1
    total_rows_valor = sum(row["valor"] for row in rows) or 1
    ranking_resp = sorted(
        ({"nome": k, "valor": v, "qtd": resp_qtd[k], "pct": round(100 * v / total_rows_valor, 1)}
         for k, v in resp_valor.items()),
        key=lambda x: -x["valor"])

    # ---- Ranking por sub-atividade (top 10, com % e % acumulado pra Pareto) --
    # % e % acumulado sao calculados sobre o TOTAL de todas as categorias (nao
    # so as 10 exibidas), entao a curva acumulada mostra corretamente que fatia
    # do total geral aquelas top 10 representam (pode nao chegar a 100% se
    # houver categorias fora do top 10 -- isso e esperado).
    sub_valor = defaultdict(float)
    sub_qtd = defaultdict(int)
    for row in rows:
        sub_valor[row["sub"]] += row["valor"]
        sub_qtd[row["sub"]] += 1
    total_sub_geral = sum(sub_valor.values()) or 1
    ranking_sub_full = sorted(
        ({"nome": k, "valor": v, "qtd": sub_qtd[k]} for k, v in sub_valor.items()),
        key=lambda x: -x["valor"])
    ranking_sub = []
    cum = 0.0
    for r in ranking_sub_full[:10]:
        cum += r["valor"]
        r["pct"] = round(100 * r["valor"] / total_sub_geral, 1)
        r["cum_pct"] = round(100 * cum / total_sub_geral, 1)
        ranking_sub.append(r)

    # ---- Pareto estratificado: top 10 CATEGORIAS (descricao do lancamento
    # agrupada por regex) dentro das 2 sub-atividades de maior valor (as
    # mesmas 2 que lideram o Pareto acima) -- pedido do Robson pra abrir o
    # que mais pesa dentro delas. As descricoes brutas (593 valores distintos)
    # foram agrupadas manualmente com o Robson numa longa rodada de revisao,
    # priorizando regras mais especificas por cima de mais genericas (a
    # PRIMEIRA regra que bater, na ordem abaixo, decide a categoria -- por
    # isso a ORDEM da lista DESC_CATEGORY_RULES importa e nao deve ser
    # reordenada sem confirmar com ele; conflitos explicitos ja foram
    # resolvidos com o Robson, ver comentarios inline).
    # % e % acumulado aqui sao sobre o total combinado dessas 2 sub-atividades
    # (nao sobre o total geral de todas as categorias).
    top2_sub_nomes = {r["nome"] for r in ranking_sub_full[:2]}
    desc_valor = defaultdict(float)
    desc_qtd = defaultdict(int)
    for row in rows:
        if row["sub"] in top2_sub_nomes:
            key = categorize_reembolso_desc(row["desc"])
            desc_valor[key] += row["valor"]
            desc_qtd[key] += 1
    total_top2_sub = sum(desc_valor.values()) or 1
    ranking_desc_top2_full = sorted(
        ({"nome": k, "valor": v, "qtd": desc_qtd[k]} for k, v in desc_valor.items()),
        key=lambda x: -x["valor"])
    ranking_desc_top2 = []
    cum = 0.0
    for r in ranking_desc_top2_full[:10]:
        cum += r["valor"]
        r["pct"] = round(100 * r["valor"] / total_top2_sub, 1)
        r["cum_pct"] = round(100 * cum / total_top2_sub, 1)
        ranking_desc_top2.append(r)
    top2_sub_labels = [r["nome"] for r in ranking_sub_full[:2]]

    # "Ultimos lancamentos" foi retirado do dashboard a pedido do Robson --
    # nao computamos mais essa lista.

    n_responsaveis = len(resp_valor)
    n_categorias = len(sub_valor)

    return dict(
        monthly=monthly, mes_atual=mes_atual,
        semana_atual_label=semana_atual_label, semana_atual_valor=semana_atual_valor,
        total_ano=total_ano, n_lancamentos=n_lancamentos,
        ranking_resp=ranking_resp, ranking_sub=ranking_sub,
        ranking_cliente_reemb=ranking_cliente_reemb,
        ranking_desc_top2=ranking_desc_top2, top2_sub_labels=top2_sub_labels,
        n_responsaveis=n_responsaveis, n_categorias=n_categorias,
        evolucao=evolucao, nota_evolucao=nota_evolucao,
    )


def render_reembolso(html, d):
    def sub_once(pattern, replacement, text, flags=re.S):
        new_text, n = re.subn(pattern, replacement, text, count=1, flags=flags)
        if n == 0:
            log(f"  [aviso] padrao nao encontrado (reembolso): {pattern[:60]}...")
        return new_text

    hoje = datetime.date.today().strftime('%d/%m/%Y')
    html = sub_once(r'<div class="badge-week" id="badgeAtualizado">Atualizado \d{2}/\d{2}/\d{4}</div>',
                     f'<div class="badge-week" id="badgeAtualizado">Atualizado {hoje}</div>', html)

    # Linha de resumo (total acumulado/semana atual/lancamentos/etc.) foi
    # retirada do topo do dashboard a pedido do Robson.

    # Notas explicativas (evolucaoNote, nota de cliente/pareto, fonteNote) foram
    # removidas do HTML a pedido do Robson -- nao ha mais substituicao aqui.

    evo = d.get("evolucao") or []
    def js_val(v):
        return "null" if v is None else str(v)
    labels_js = "[" + ",".join(js_str(e["label"]) for e in evo) + "]"
    valores_js = "[" + ",".join(js_val(e["valor"]) for e in evo) + "]"
    tecnicos_js = "[" + ",".join(js_val(e.get("tecnicos")) for e in evo) + "]"
    tipos_js = "[" + ",".join(js_str(e["tipo"]) for e in evo) + "]"
    html = sub_once(r'var reembolsoEvoLabels = \[.*?\];', f'var reembolsoEvoLabels = {labels_js};', html)
    html = sub_once(r'var reembolsoEvoValores = \[.*?\];', f'var reembolsoEvoValores = {valores_js};', html)
    html = sub_once(r'var reembolsoEvoTecnicos = \[.*?\];', f'var reembolsoEvoTecnicos = {tecnicos_js};', html)
    html = sub_once(r'var reembolsoEvoTipos = \[.*?\];', f'var reembolsoEvoTipos = {tipos_js};', html)

    rcli = d.get("ranking_cliente_reemb") or []
    cli_labels_js = "[" + ",".join(js_str(r["nome"]) for r in rcli) + "]"
    cli_valores_js = "[" + ",".join(js_val(round(r["valor"], 2)) for r in rcli) + "]"
    cli_pct_js = "[" + ",".join(js_val(r["pct"]) for r in rcli) + "]"
    cli_nproj_js = "[" + ",".join(js_val(r["n_projetos"]) for r in rcli) + "]"
    html = sub_once(r'var reembolsoCliLabels = \[.*?\];', f'var reembolsoCliLabels = {cli_labels_js};', html)
    html = sub_once(r'var reembolsoCliValores = \[.*?\];', f'var reembolsoCliValores = {cli_valores_js};', html)
    html = sub_once(r'var reembolsoCliPct = \[.*?\];', f'var reembolsoCliPct = {cli_pct_js};', html)
    html = sub_once(r'var reembolsoCliNProjetos = \[.*?\];', f'var reembolsoCliNProjetos = {cli_nproj_js};', html)

    # Painel "Por Responsável" foi removido do dashboard a pedido do Robson --
    # ranking_resp continua calculado acima (usado no resumo "N responsaveis"),
    # so a renderizacao da tabela em si foi tirada.

    rsub = d.get("ranking_sub") or []
    sub_labels_js = "[" + ",".join(js_str(r["nome"]) for r in rsub) + "]"
    sub_valores_js = "[" + ",".join(js_val(round(r["valor"], 2)) for r in rsub) + "]"
    sub_qtd_js = "[" + ",".join(js_val(r["qtd"]) for r in rsub) + "]"
    sub_cum_js = "[" + ",".join(js_val(r["cum_pct"]) for r in rsub) + "]"
    html = sub_once(r'var reembolsoSubLabels = \[.*?\];', f'var reembolsoSubLabels = {sub_labels_js};', html)
    html = sub_once(r'var reembolsoSubValores = \[.*?\];', f'var reembolsoSubValores = {sub_valores_js};', html)
    html = sub_once(r'var reembolsoSubQtd = \[.*?\];', f'var reembolsoSubQtd = {sub_qtd_js};', html)
    html = sub_once(r'var reembolsoSubCumPct = \[.*?\];', f'var reembolsoSubCumPct = {sub_cum_js};', html)

    rdesc2 = d.get("ranking_desc_top2") or []
    desc2_labels_js = "[" + ",".join(js_str(r["nome"]) for r in rdesc2) + "]"
    desc2_valores_js = "[" + ",".join(js_val(round(r["valor"], 2)) for r in rdesc2) + "]"
    desc2_qtd_js = "[" + ",".join(js_val(r["qtd"]) for r in rdesc2) + "]"
    desc2_cum_js = "[" + ",".join(js_val(r["cum_pct"]) for r in rdesc2) + "]"
    top2_sub_labels = d.get("top2_sub_labels") or []
    top2_sub_txt = " + ".join(top2_sub_labels) if top2_sub_labels else "-"
    html = sub_once(r'var reembolsoDescTop2Labels = \[.*?\];', f'var reembolsoDescTop2Labels = {desc2_labels_js};', html)
    html = sub_once(r'var reembolsoDescTop2Valores = \[.*?\];', f'var reembolsoDescTop2Valores = {desc2_valores_js};', html)
    html = sub_once(r'var reembolsoDescTop2Qtd = \[.*?\];', f'var reembolsoDescTop2Qtd = {desc2_qtd_js};', html)
    html = sub_once(r'var reembolsoDescTop2CumPct = \[.*?\];', f'var reembolsoDescTop2CumPct = {desc2_cum_js};', html)
    html = sub_once(r'(<p class="panel-note" id="paretoDescTop2Note">).*?(</p>)',
                     lambda m: m.group(1) + f'Top 10 categorias de lançamento (descrições agrupadas) dentro das 2 sub-atividades de maior valor ({top2_sub_txt}), acumulado 2026.' + m.group(2), html)

    if d["monthly"]:
        rows_html = ""
        for m in reversed(d["monthly"]):
            qtd_txt = str(int(m["qtd"])) if isinstance(m["qtd"], (int, float)) else "-"
            tec_txt = str(m["tecnicos_media"]) if m.get("tecnicos_media") is not None else "-"
            vpt_txt = fmt_brl(m["valor_por_tecnico"]) if m.get("valor_por_tecnico") is not None else "-"
            rows_html += (f'    <tr><td>{m["mes"]}</td><td class="num">{qtd_txt}</td>'
                          f'<td class="num">{fmt_brl(m["media"])}</td><td class="num">{tec_txt}</td>'
                          f'<td class="num">{vpt_txt}</td></tr>\n')
    else:
        rows_html = '    <tr><td colspan="5" style="color:var(--muted); text-align:center;">Nenhum dado encontrado</td></tr>\n'
    html = sub_once(r'<tbody id="monthlyRows">\s*.*?\s*</tbody>', '<tbody id="monthlyRows">\n' + rows_html + '    </tbody>', html)

    # Painel "Ultimos Lancamentos" foi retirado a pedido do Robson.

    return html


# ============================================================
# NOSSAS DEMANDAS (copia consolidada -- nao le planilha nenhuma, so reaproveita
# os dicts ja construidos pelos 4 build_*_data() acima)
# ============================================================

def build_demandas_data(d_eq, d_fr, d_bo, d_en):
    hoje = datetime.date.today()

    # BUG corrigido (09/08/2026): quando uma fonte falha na leitura (ex.:
    # ENVIOS_XLSX nao encontrado -- ver [ERRO] em main()), o parametro
    # correspondente (d_eq/d_fr/d_bo/d_en) chega aqui como None. Antes disso
    # era tratado exatamente igual a "fonte leu certo e nao tem pendencia
    # nenhuma": a secao ficava com lista vazia, sem nenhuma distincao visual,
    # e o HTML era regravado assim mesmo -- apagando (no app, via
    # sync_to_mobile_www) uma secao que na sessao anterior podia estar cheia
    # e correta. "source_failed" carrega essa distincao adiante pra
    # render_demandas(), que agora PULA a regravacao do tbody da secao cuja
    # fonte falhou, preservando o HTML anterior em vez de zerar.
    source_failed = dict(equipes=d_eq is None, frota=d_fr is None, boletos=d_bo is None, envios=d_en is None)

    # ---- Equipes: desmobilizacoes ja decididas (toda a lista ja e "pendente
    # de saida", nao precisa filtrar mais nada) ----
    equipes_rows = []
    if d_eq is not None:
        for t in d_eq.get("em_desmobilizacao", []):
            data_str = ""
            if t.get("data_finalizacao"):
                try:
                    data_str = datetime.date.fromisoformat(str(t["data_finalizacao"])[:10]).strftime('%d/%m/%Y')
                except Exception:
                    data_str = str(t["data_finalizacao"])
            equipes_rows.append(dict(mat=t.get("mat", ""), nome=t.get("nome", ""), parque=t.get("parque", ""),
                                      cliente=t.get("cliente", ""), data=data_str))

    # ---- Frota: acoes com prazo em hoje (ou vencido) OU marcadas "!!!"
    # (urgente). Regra ATUALIZADA pelo Robson em 18/08/2026 (substitui a de
    # 05/08/2026, documentada no SKILL.md da tarefa agendada e no
    # REGRAS_AUTOMACAO.md -- so agora implementada de fato, confirmado com o
    # Robson em 22/08/2026 antes desta edicao):
    #   - status em branco (sem prazo definido) volta a contar como "vence
    #     hoje" -- mesmo criterio ja usado no proprio dashboard de Frota
    #     (normalizarStatusVencidoOuSemData, no JS do dashboard-frota-
    #     manutencao.html), agora espelhado aqui para Nossas Demandas.
    #   - itens marcados "Acompanhar"/"Stand by" continuam de fora desta
    #     consolidacao (regra de 05/08/2026 mantida neste ponto) -- ficam
    #     visiveis so no proprio dashboard de Frota (tag-warn amarelo).
    frota_rows = []
    if d_fr is not None:
        for a in d_fr.get("acao_rows", []):
            status = (a.get("status") or "").strip()
            if re.search(r'conclu[ií]do', status, re.I):
                continue
            if re.search(r'acompanhar|stand by', status, re.I):
                continue
            is_urgente = status == "!!!"
            is_sem_data = status == ""
            is_hoje_ou_vencido = False
            m = re.match(r'^(\d{2})/(\d{2})/(\d{4})$', status)
            if m:
                try:
                    dt = datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
                    is_hoje_ou_vencido = dt <= hoje
                except Exception:
                    pass
            if is_urgente or is_sem_data or is_hoje_ou_vencido:
                if is_urgente:
                    status_label = "Urgente"
                elif is_sem_data:
                    status_label = hoje.strftime('%d/%m/%Y')
                else:
                    status_label = status
                frota_rows.append(dict(acao=a.get("acao", ""), tipo=a.get("tipo", ""), parque=a.get("parque", ""),
                                        quem=a.get("quem", ""), status=status_label,
                                        obs=a.get("obs", "")))

    # ---- Boletos: tudo que ja esta nas 3 secoes de "pendencia em aberto"
    # (multas / contestar / sinistros) -- essas secoes na planilha origem ja
    # so trazem o que esta em aberto hoje, entao entra tudo. ----
    boletos_rows = []
    if d_bo is not None:
        for tipo, key in (("Multa", "detalhe_multas"), ("Contestar", "detalhe_contestar"), ("Sinistro", "detalhe_sinistros")):
            for r in d_bo.get(key, []):
                boletos_rows.append(dict(tipo=tipo, placa=r.get("placa", ""), locadora=r.get("locadora", ""),
                                          parque=r.get("parque", ""), venc=r.get("venc", ""), motivo=r.get("motivo", "")))

    # ---- Envios Logisticos: em transito/aguardando assinatura na Logistica.
    # Itens do Almoxarifado (divergencias) ficam de fora a pedido do Robson --
    # so aparecem no proprio dashboard de Envios, nao entram na consolidacao.
    # "Aguardando Assinatura" (quadro separado, detectado por preenchimento
    # amarelo sem data na coluna N -- ver build_envios_data) tambem entra
    # aqui, pedido do Robson em 05/08/2026, mesmo criterio ja usado pra
    # trazer itens em amarelo dos outros dashboards pra esta consolidacao. ----
    # BUG corrigido (09/08/2026): uma mesma requisicao pode satisfazer DOIS
    # criterios ao mesmo tempo -- status textual "EM TRAJETO"/"AGUARDANDO
    # ASSINATURA" na coluna J (capturado em acoes_log) E o criterio visual
    # de celula N vazia com fill amarelo (capturado separadamente em
    # aguardando_assinatura, ver build_envios_data) -- e entrava duas vezes
    # em envios_rows, uma vinda de cada lista. Dedup por chave estavel
    # (projeto+responsavel+prazo+destinatario) antes de retornar; quando ha
    # colisao, mantemos a entrada de aguardando_assinatura (criterio mais
    # confiavel/manual) sobre a de acoes_log.
    envios_seen = {}
    for r in d_en.get("acoes_log", []) if d_en is not None else []:
        status_l = (r.get("status") or "").lower()
        if "em trajeto" in status_l or "aguardando assinatura" in status_l:
            key = (r.get("projeto", ""), r.get("responsavel", ""), r.get("prazo", ""), r.get("destinatario", ""))
            envios_seen[key] = dict(setor="Logística", projeto=r.get("projeto", ""), responsavel=r.get("responsavel", ""),
                                     prazo=r.get("prazo", ""), destinatario=r.get("destinatario", ""), status=r.get("status", ""))
    for r in d_en.get("aguardando_assinatura", []) if d_en is not None else []:
        key = (r.get("projeto", ""), r.get("responsavel", ""), r.get("prazo", ""), r.get("destinatario", ""))
        envios_seen[key] = dict(setor="Logística", projeto=r.get("projeto", ""), responsavel=r.get("responsavel", ""),
                                 prazo=r.get("prazo", ""), destinatario=r.get("destinatario", ""), status="Aguardando assinatura")
    envios_rows = list(envios_seen.values())

    return dict(hoje=hoje, equipes_rows=equipes_rows, frota_rows=frota_rows,
                boletos_rows=boletos_rows, envios_rows=envios_rows,
                source_failed=source_failed)


def render_demandas(html, d):
    def sub_once(pattern, replacement, text, flags=re.S):
        new_text, n = re.subn(pattern, replacement, text, count=1, flags=flags)
        if n == 0:
            log(f"  [aviso] padrao nao encontrado (demandas): {pattern[:60]}...")
        return new_text

    hoje_str = d["hoje"].strftime('%d/%m/%Y')
    html = sub_once(r'<div class="badge-week" id="badgeData">Data: \d{2}/\d{2}/\d{4}</div>',
                     f'<div class="badge-week" id="badgeData">Data: {hoje_str}</div>', html)

    # BUG corrigido (09/08/2026): quando source_failed[chave] e True, a fonte
    # daquela secao falhou nesta execucao (ver build_demandas_data) -- pulamos
    # a atualizacao do contador E do tbody dessa secao inteira, preservando
    # o que ja estava gravado no HTML anterior em vez de mostrar zero/vazio
    # (que antes era indistinguivel de "genuinamente sem pendencias").
    sf = d.get("source_failed", {})
    if sf.get("equipes") or sf.get("frota") or sf.get("boletos") or sf.get("envios"):
        falhas = [k for k in ("equipes", "frota", "boletos", "envios") if sf.get(k)]
        log(f"  [aviso] Nossas Demandas: secao(oes) preservada(s) do HTML anterior por falha de leitura: {', '.join(falhas)}.")

    n_eq, n_fr, n_bo, n_en = len(d["equipes_rows"]), len(d["frota_rows"]), len(d["boletos_rows"]), len(d["envios_rows"])

    # cntTotal/pillTotal precisam ser lidos de volta do HTML atual para as
    # secoes preservadas, ja que n_eq/n_fr/n_bo/n_en aqui vem vazias quando a
    # fonte falhou (nao refletem o que continua exibido na secao antiga).
    def current_count(span_id, fallback):
        m = re.search(rf'<[^>]*id="{span_id}">(\d+)</', html)
        return int(m.group(1)) if m else fallback

    eff_eq = current_count("cntEquipes", n_eq) if sf.get("equipes") else n_eq
    eff_fr = current_count("cntFrota", n_fr) if sf.get("frota") else n_fr
    eff_bo = current_count("cntBoletos", n_bo) if sf.get("boletos") else n_bo
    eff_en = current_count("cntEnvios", n_en) if sf.get("envios") else n_en
    n_total = eff_eq + eff_fr + eff_bo + eff_en

    for span_id, val, skip in (
        ("cntEquipes", n_eq, sf.get("equipes")), ("cntFrota", n_fr, sf.get("frota")),
        ("cntBoletos", n_bo, sf.get("boletos")), ("cntEnvios", n_en, sf.get("envios")),
        ("cntTotal", n_total, False),
        ("pillEquipes", n_eq, sf.get("equipes")), ("pillFrota", n_fr, sf.get("frota")),
        ("pillBoletos", n_bo, sf.get("boletos")), ("pillEnvios", n_en, sf.get("envios")),
    ):
        if skip:
            continue
        html = sub_once(rf'(<[^>]*id="{span_id}">)\d+(</)', lambda m, v=val: m.group(1) + str(v) + m.group(2), html)

    if not sf.get("equipes"):
        if d["equipes_rows"]:
            rows_html = "".join(
                f'    <tr class="row-warn"><td>{r["mat"]}</td><td>{r["nome"]}</td><td>{r["parque"]}</td>'
                f'<td>{r["cliente"]}</td><td>{r["data"]}</td></tr>\n' for r in d["equipes_rows"])
        else:
            rows_html = '    <tr><td colspan="5" class="empty-note">Nenhuma demanda</td></tr>\n'
        html = sub_once(r'<tbody id="bodyEquipes">.*?</tbody>', '<tbody id="bodyEquipes">\n' + rows_html + '    </tbody>', html)

    if not sf.get("frota"):
        if d["frota_rows"]:
            rows_html = "".join(
                f'    <tr class="row-warn"><td>{r["acao"]}</td><td>{r["tipo"]}</td><td>{r["parque"]}</td>'
                f'<td>{r["quem"]}</td><td>{r["status"]}</td><td>{r["obs"]}</td></tr>\n' for r in d["frota_rows"])
        else:
            rows_html = '    <tr><td colspan="6" class="empty-note">Nenhuma demanda</td></tr>\n'
        html = sub_once(r'<tbody id="bodyFrota">.*?</tbody>', '<tbody id="bodyFrota">\n' + rows_html + '    </tbody>', html)

    if not sf.get("boletos"):
        if d["boletos_rows"]:
            rows_html = "".join(
                f'    <tr class="row-warn"><td>{r["tipo"]}</td><td>{r["placa"]}</td><td>{r["locadora"]}</td>'
                f'<td>{r["parque"]}</td><td>{r["venc"]}</td><td>{r["motivo"]}</td></tr>\n' for r in d["boletos_rows"])
        else:
            rows_html = '    <tr><td colspan="6" class="empty-note">Nenhuma demanda</td></tr>\n'
        html = sub_once(r'<tbody id="bodyBoletos">.*?</tbody>', '<tbody id="bodyBoletos">\n' + rows_html + '    </tbody>', html)

    if not sf.get("envios"):
        if d["envios_rows"]:
            rows_html = "".join(
                f'    <tr class="row-warn"><td>{r["setor"]}</td><td>{r["projeto"]}</td><td>{r["responsavel"]}</td>'
                f'<td>{r["prazo"]}</td><td>{r["destinatario"]}</td><td>{r["status"]}</td></tr>\n' for r in d["envios_rows"])
        else:
            rows_html = '    <tr><td colspan="6" class="empty-note">Nenhuma demanda</td></tr>\n'
        html = sub_once(r'<tbody id="bodyEnvios">.*?</tbody>', '<tbody id="bodyEnvios">\n' + rows_html + '    </tbody>', html)

    return html


def main():
    log(f"=== Atualizando dashboards Extreme Wind - {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')} ===")
    ok_eq = ok_fr = False
    d_eq_ok = d_fr_ok = None

    try:
        log("Lendo Diario de Bordo...")
        if not os.path.isfile(EQUIPES_XLSX):
            raise FileNotFoundError(EQUIPES_XLSX)
        d_eq = build_equipes_data()
        log(f"  Equipes: real={d_eq['tot_real']} plano={d_eq['tot_plano']} parques_ativos={d_eq['parques_ativos']}/{d_eq['parques_total']} total_eventos={d_eq['n_total']}")

        try:
            d_eq["tecnicos_busca"] = build_tecnicos_data()
            log(f"  Busca por tecnico: {len(d_eq['tecnicos_busca'])} tecnicos indexados.")
        except Exception as e:
            log(f"  [ERRO] Falha ao montar dados de busca por tecnico: {e!r}. Aba 'Buscar Tecnico' mantida como estava.")
            d_eq["tecnicos_busca"] = None

        with open(EQUIPES_HTML, encoding="utf-8") as f:
            html_eq = f.read()
        html_eq_new = render_equipes(html_eq, d_eq)
        with open(EQUIPES_HTML, "w", encoding="utf-8") as f:
            f.write(html_eq_new)
        log(f"  Gravado: {EQUIPES_HTML}")
        ok_eq = True
        d_eq_ok = d_eq
    except Exception as e:
        log(f"  [ERRO] Falha ao atualizar dashboard de Equipes: {e!r}. Arquivo mantido como estava (nao sobrescrito).")

    try:
        log("Lendo Controle de Frota...")
        if not os.path.isfile(FROTA_XLSX):
            raise FileNotFoundError(FROTA_XLSX)
        d_fr = build_frota_data()
        log(f"  Frota: semana={d_fr['week_label']} veiculos={d_fr['n_total_veic']} conformidade={d_fr['conformidade_geral']} acoes={d_fr['n_acoes']}")
        if d_fr["uncertain_weeks"]:
            log(f"  [risco] Semanas com rotulo original inconsistente: {sorted(d_fr['uncertain_weeks'])}")

        with open(FROTA_HTML, encoding="utf-8") as f:
            html_fr = f.read()
        html_fr_new = render_frota(html_fr, d_fr)
        with open(FROTA_HTML, "w", encoding="utf-8") as f:
            f.write(html_fr_new)
        log(f"  Gravado: {FROTA_HTML}")
        ok_fr = True
        d_fr_ok = d_fr
    except Exception as e:
        log(f"  [ERRO] Falha ao atualizar dashboard de Frota: {e!r}. Arquivo mantido como estava (nao sobrescrito).")

    d_bo_ok = None
    try:
        log("Lendo Controle de Boletos (Locadoras)...")
        if not wait_for_file(BOLETOS_XLSX):
            raise FileNotFoundError(BOLETOS_XLSX)
        try:
            recalc_boletos_xlsx(BOLETOS_XLSX)
        except Exception as e:
            log(f"  [aviso] Boletos: recalculo de formulas nao concluido ({e!r}) -- lendo indicadores com o cache existente no arquivo.")
        d_bo = build_boletos_data()
        log(f"  Boletos: {len(d_bo['indicadores'])}/7 indicadores lidos, "
            f"{len(d_bo['detalhe_multas'])} multas / {len(d_bo['detalhe_contestar'])} contestar / "
            f"{len(d_bo['detalhe_sinistros'])} sinistros / {len(d_bo['detalhe_contestacao'])} contestacao no detalhe"
            + (f", {len(d_bo['log_rows'])} linhas de log semanal" if d_bo.get("log_table_found") else ", tabela de log semanal nao encontrada (preservado)"))

        with open(BOLETOS_HTML, encoding="utf-8") as f:
            html_bo = f.read()
        html_bo_new = render_boletos(html_bo, d_bo)
        with open(BOLETOS_HTML, "w", encoding="utf-8") as f:
            f.write(html_bo_new)
        log(f"  Gravado: {BOLETOS_HTML}")
        d_bo_ok = d_bo
    except Exception as e:
        log(f"  [ERRO] Falha ao atualizar dashboard de Boletos: {e!r}. Arquivo mantido como estava (nao sobrescrito).")

    d_en_ok = None
    try:
        log("Lendo Controle Logistico de Requisicoes (Envios)...")
        if not wait_for_file(ENVIOS_XLSX):
            raise FileNotFoundError(ENVIOS_XLSX)
        d_en = build_envios_data()
        aberto_total = d_en["aberto_total"] if d_en["aberto_total"] is not None else sum(q for (_, q) in d_en["aberto_rows"])
        log(f"  Envios: total={d_en['kpi_total']} almox={d_en['kpi_almox']} entregues={d_en['kpi_entregues']} "
            f"canceladas={d_en['kpi_cancel']} em_aberto={aberto_total} acoes_almox={len(d_en['acoes_almox'])} "
            f"acoes_log={len(d_en['acoes_log'])} aguardando_assinatura={len(d_en['aguardando_assinatura'])}")

        with open(ENVIOS_HTML, encoding="utf-8") as f:
            html_en = f.read()
        html_en_new = render_envios(html_en, d_en)
        with open(ENVIOS_HTML, "w", encoding="utf-8") as f:
            f.write(html_en_new)
        log(f"  Gravado: {ENVIOS_HTML}")
        d_en_ok = d_en
    except Exception as e:
        log(f"  [ERRO] Falha ao atualizar dashboard de Envios Logisticos: {e!r}. Arquivo mantido como estava (nao sobrescrito).")

    d_re_ok = None
    try:
        log("Lendo Formulario de Reembolso...")
        if not os.path.isfile(REEMBOLSO_XLSX):
            raise FileNotFoundError(REEMBOLSO_XLSX)
        d_re = build_reembolso_data()
        mes_nome = d_re["mes_atual"]["mes"] if d_re["mes_atual"] else "?"
        log(f"  Reembolso: mes_atual={mes_nome} valor_mes={d_re['mes_atual']['valor'] if d_re['mes_atual'] else '-'} "
            f"semana_atual={d_re['semana_atual_label']} total_ano={round(d_re['total_ano'],2)} "
            f"lancamentos={d_re['n_lancamentos']} responsaveis={d_re['n_responsaveis']} categorias={d_re['n_categorias']}")

        if not os.path.isfile(REEMBOLSO_HTML):
            raise FileNotFoundError(REEMBOLSO_HTML)
        with open(REEMBOLSO_HTML, encoding="utf-8") as f:
            html_re = f.read()
        html_re_new = render_reembolso(html_re, d_re)
        with open(REEMBOLSO_HTML, "w", encoding="utf-8") as f:
            f.write(html_re_new)
        log(f"  Gravado: {REEMBOLSO_HTML}")
        d_re_ok = d_re
    except Exception as e:
        log(f"  [ERRO] Falha ao atualizar dashboard de Reembolso: {e!r}. Arquivo mantido como estava (nao sobrescrito).")

    try:
        log("Consolidando Nossas Demandas...")
        d_dm = build_demandas_data(d_eq_ok, d_fr_ok, d_bo_ok, d_en_ok)
        n_total = len(d_dm["equipes_rows"]) + len(d_dm["frota_rows"]) + len(d_dm["boletos_rows"]) + len(d_dm["envios_rows"])
        log(f"  Demandas: equipes={len(d_dm['equipes_rows'])} frota={len(d_dm['frota_rows'])} "
            f"boletos={len(d_dm['boletos_rows'])} envios={len(d_dm['envios_rows'])} total={n_total}")

        if not os.path.isfile(DEMANDAS_HTML):
            raise FileNotFoundError(DEMANDAS_HTML)
        with open(DEMANDAS_HTML, encoding="utf-8") as f:
            html_dm = f.read()
        html_dm_new = render_demandas(html_dm, d_dm)
        with open(DEMANDAS_HTML, "w", encoding="utf-8") as f:
            f.write(html_dm_new)
        log(f"  Gravado: {DEMANDAS_HTML}")
    except Exception as e:
        log(f"  [ERRO] Falha ao atualizar Nossas Demandas: {e!r}. Arquivo mantido como estava (nao sobrescrito).")

    try:
        if not VAGAS_MESTRE_XLSX or not os.path.isfile(VAGAS_MESTRE_XLSX):
            raise FileNotFoundError(f"VAGAS_MESTRE nao encontrada: {VAGAS_MESTRE_XLSX}")
        if not os.path.isfile(EQUIPES_XLSX):
            raise FileNotFoundError(f"EQUIPES (Diario de Bordo) nao encontrada: {EQUIPES_XLSX}")
        if not os.path.isfile(EFETIVO_HTML):
            raise FileNotFoundError(f"Dashboard template nao encontrado: {EFETIVO_HTML}")

        from efetivo_pipeline import build_equipe_by_mat, build_historico_by_mat, build_efetivo_data, render_efetivo
        equipe_by_mat = build_equipe_by_mat(EQUIPES_XLSX)
        historico_por_mat = build_historico_by_mat(EQUIPES_XLSX)
        d_ef_ok = build_efetivo_data(VAGAS_MESTRE_XLSX, equipe_by_mat=equipe_by_mat)

        log(f"  Controle de Efetivo: {len(d_ef_ok.get('pessoas', []))} pessoas, "
            f"disponivel={d_ef_ok.get('total_disponivel', 0)}, "
            f"bdh={d_ef_ok.get('bdh', 0)}, "
            f"aguardando={d_ef_ok.get('aguardando_convocacao', 0)}")

        with open(EFETIVO_HTML, encoding="utf-8") as f:
            html_ef = f.read()
        html_ef_new = render_efetivo(html_ef, d_ef_ok, historico_por_mat)
        with open(EFETIVO_HTML, "w", encoding="utf-8") as f:
            f.write(html_ef_new)
        log(f"  Gravado: {EFETIVO_HTML}")
    except FileNotFoundError as e:
        log(f"  [SKIP] Arquivo nao encontrado para Controle de Efetivo: {e}")
        d_ef_ok = None
    except Exception as e:
        log(f"  [ERRO] Falha ao atualizar Controle de Efetivo: {e!r}. Arquivo mantido como estava (nao sobrescrito).")
        d_ef_ok = None

    if d_eq_ok is not None or d_fr_ok is not None or d_bo_ok is not None or d_en_ok is not None or d_re_ok is not None:
        try:
            if not os.path.isfile(INDEX_HTML):
                raise FileNotFoundError(INDEX_HTML)
            with open(INDEX_HTML, encoding="utf-8") as f:
                html_idx = f.read()
            html_idx_new = render_index(html_idx, d_eq_ok, d_fr_ok, d_bo_ok, d_en_ok, d_re_ok)
            with open(INDEX_HTML, "w", encoding="utf-8") as f:
                f.write(html_idx_new)
            log(f"  Gravado: {INDEX_HTML}")
        except Exception as e:
            log(f"  [ERRO] Falha ao atualizar pagina inicial (index.html): {e!r}. Arquivo mantido como estava (nao sobrescrito).")

    try:
        sync_to_mobile_www()
    except Exception as e:
        log(f"  [ERRO] Falha ao sincronizar com Mobile/www/: {e!r}.")

    log("=== Concluido ===" if (ok_eq and ok_fr) else "=== Concluido com falhas (ver [ERRO] acima) ===")
    return ok_eq and ok_fr


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)


